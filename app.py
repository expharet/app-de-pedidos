import streamlit as st
import pandas as pd
import json
import os
import uuid
from datetime import datetime, date, timedelta
import io
import hashlib
from typing import Dict, List, Optional
from PIL import Image

# ─── PAGE CONFIG ─────────────────────────────────────────────
st.set_page_config(
    page_title="Export Haret - Pedidos v5.0",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── CONSTANTS ───────────────────────────────────────────────
ORDEN_ESTADOS = ["Recibido","Confirmado","Preparando","Enviado","Entregado","Cancelado"]
ESTADO_ICONS = {
    "Recibido":"📬","Confirmado":"✅","Preparando":"📦",
    "Enviado":"🚚","Entregado":"✨","Cancelado":"❌",
}
DATA_FILE      = "precios_data.json"
CLIENTS_FILE   = "clientes.json"
HIST_FILE      = "precio_historial.json"
EMAIL_LOG_FILE = "email_log.json"
PEDIDOS_FILE   = "pedidos_data.json"

USERS = {
    "admin@exportharet.com":  {"pwd": hashlib.md5(b"admin123").hexdigest(),  "rol": "admin",   "nombre": "Admin"},
    "ventas@exportharet.com": {"pwd": hashlib.md5(b"ventas123").hexdigest(), "rol": "ventas",  "nombre": "Ventas"},
}

# ─── DATA HELPERS ─────────────────────────────────────────────
def _load_json(path, default):
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return default

def _save_json(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        st.error(f"Error guardando {path}: {e}")
        return False

@st.cache_data(ttl=60)
def load_precios():
    return _load_json(DATA_FILE, {"products": [], "config": {"destinos": {}, "grupos": {}, "minimos": {}}, "pedidos": []})

def load_clients():
    return _load_json(CLIENTS_FILE, {})

def load_pedidos():
    return _load_json(PEDIDOS_FILE, [])

def load_historial():
    return _load_json(HIST_FILE, [])

def load_email_log():
    return _load_json(EMAIL_LOG_FILE, [])

def save_precios(data):
    _save_json(DATA_FILE, data)
    st.cache_data.clear()

def save_clients(clients):
    _save_json(CLIENTS_FILE, clients)

def save_pedidos(pedidos):
    _save_json(PEDIDOS_FILE, pedidos)

def save_historial(hist):
    _save_json(HIST_FILE, hist)

def save_email_log(log):
    _save_json(EMAIL_LOG_FILE, log)


# ─── AUTH ─────────────────────────────────────────────────────
def check_login():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "user_email" not in st.session_state:
        st.session_state.user_email = ""
    if "user_rol" not in st.session_state:
        st.session_state.user_rol = ""
    if "user_nombre" not in st.session_state:
        st.session_state.user_nombre = ""

def login_page():
    st.markdown("""
    <div style='text-align:center; padding:40px 0 20px 0;'>
        <h1>🚀 Export Haret</h1>
        <h3 style='color:#666;'>Sistema de Gestión de Pedidos v5.0</h3>
    </div>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.5, 1])
    with col2:
        st.markdown("### 🔐 Iniciar Sesión")
        email = st.text_input("Email", placeholder="usuario@exportharet.com")
        pwd   = st.text_input("Contraseña", type="password", placeholder="••••••••")

        if st.button("Entrar →", use_container_width=True, type="primary"):
            pwd_hash = hashlib.md5(pwd.encode()).hexdigest()
            if email in USERS and USERS[email]["pwd"] == pwd_hash:
                st.session_state.logged_in  = True
                st.session_state.user_email = email
                st.session_state.user_rol   = USERS[email]["rol"]
                st.session_state.user_nombre = USERS[email]["nombre"]
                st.rerun()
            else:
                st.error("❌ Email o contraseña incorrectos")

        st.markdown("---")
        st.caption("👤 admin@exportharet.com / admin123")
        st.caption("👤 ventas@exportharet.com / ventas123")


# ─── SEGMENTACIÓN DE CLIENTES ─────────────────────────────────
def segmentar_cliente(email, clients):
    c = clients.get(email, {})
    pedidos = c.get("pedidos_ids", [])
    todos_pedidos = load_pedidos()
    mis_pedidos = [p for p in todos_pedidos if p.get("client_email") == email]

    if not mis_pedidos:
        return {"segmento": "Nuevo", "descuento": 0.0, "credito_max": 10000, "badge": "🆕 Nuevo"}

    hoy = datetime.now()
    pedidos_30d = [p for p in mis_pedidos if (hoy - datetime.fromisoformat(p.get("fecha", hoy.isoformat()))).days <= 30]
    factura_30d = sum(p.get("total_usd", 0) for p in pedidos_30d)

    if factura_30d >= 5000 or len(pedidos_30d) >= 10:
        return {"segmento": "VIP", "descuento": 0.05, "credito_max": 50000, "badge": "⭐ VIP"}
    elif len(mis_pedidos) >= 2:
        return {"segmento": "Regular", "descuento": 0.02, "credito_max": 25000, "badge": "⚫ Regular"}
    else:
        return {"segmento": "Nuevo", "descuento": 0.0, "credito_max": 10000, "badge": "🆕 Nuevo"}

# ─── PRECIO HELPERS ───────────────────────────────────────────
def get_precio_producto(codigo, destino_key, data):
    for prod in data.get("products", []):
        if prod.get("codigo") == codigo:
            precio_base = prod.get("precio_cif_usd", 0)
            config = data.get("config", {})
            destinos = config.get("destinos", {})
            dest_info = destinos.get(destino_key, {})
            factor = dest_info.get("factor", 1.0)
            return round(precio_base * factor, 2)
    return 0.0

def registrar_cambio_precio(codigo, antes, despues, motivo="Manual"):
    if antes == despues:
        return
    cambio_pct = ((despues - antes) / antes * 100) if antes > 0 else 0
    hist = load_historial()
    hist.append({
        "id": f"CHG-{len(hist)+1:05d}",
        "fecha": datetime.now().isoformat(),
        "producto": codigo,
        "antes": antes,
        "despues": despues,
        "cambio_pct": round(cambio_pct, 2),
        "usuario": st.session_state.get("user_email", "sistema"),
        "motivo": motivo
    })
    save_historial(hist)
    if abs(cambio_pct) > 20:
        st.warning(f"⚠️ Cambio > 20% en {codigo}: {cambio_pct:+.1f}%")


# ─── EXCEL EXPORT ─────────────────────────────────────────────
def exportar_excel(pedidos):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        st.error("Instalar openpyxl")
        return None

    wb = Workbook()

    # Hoja 1: Resumen
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1["A1"] = "EXPORT HARET - REPORTE DE PEDIDOS"
    ws1["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws1["A1"].fill = PatternFill(start_color="003E8C", end_color="003E8C", fill_type="solid")
    ws1.merge_cells("A1:E1")
    ws1["A2"] = f"Generado: {date.today().strftime('%d/%m/%Y')}"

    estados = {}
    for p in pedidos:
        est = p.get("estado", "Recibido")
        if est not in estados:
            estados[est] = {"count": 0, "total": 0}
        estados[est]["count"] += 1
        estados[est]["total"] += p.get("total_usd", 0)

    ws1.append(["ESTADO", "PEDIDOS", "TOTAL USD", "% TOTAL"])
    total_gen = sum(d["total"] for d in estados.values()) or 1
    for est, d in sorted(estados.items()):
        ws1.append([est, d["count"], round(d["total"], 2), round(d["total"]/total_gen*100, 1)])

    # Hoja 2: Detalle Pedidos
    ws2 = wb.create_sheet("Pedidos")
    ws2.append(["ID", "CLIENTE", "EMAIL", "ESTADO", "DESTINO", "TOTAL USD", "FECHA", "NOTAS"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003E8C", end_color="003E8C", fill_type="solid")
    for p in sorted(pedidos, key=lambda x: x.get("fecha", ""), reverse=True):
        ws2.append([
            p.get("id", "").upper(), p.get("client_name", ""), p.get("client_email", ""),
            p.get("estado", ""), p.get("destino", ""), round(p.get("total_usd", 0), 2),
            p.get("fecha", "")[:10], p.get("notas", "")[:80]
        ])

    # Hoja 3: Productos
    ws3 = wb.create_sheet("Productos")
    ws3.append(["CÓDIGO", "PRODUCTO", "CAJAS", "PALLETS", "PRECIO USD", "TOTAL"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003E8C", end_color="003E8C", fill_type="solid")
    for p in pedidos:
        for prod in p.get("productos", []):
            total_prod = prod.get("cajas", 0) * prod.get("precio_usd", 0)
            ws3.append([prod.get("codigo",""), prod.get("producto",""), prod.get("cajas",0), round(prod.get("pallets",0),2), round(prod.get("precio_usd",0),2), round(total_prod,2)])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# ─── SLA ──────────────────────────────────────────────────────
def calcular_sla(pedidos):
    metas = {"Recibido_Confirmado": 4, "Confirmado_Preparando": 2, "Preparando_Enviado": 48, "Enviado_Entregado": 168}
    slas = []
    for p in pedidos:
        hist = sorted(p.get("historial_estados", []), key=lambda x: x.get("fecha", ""))
        for i in range(len(hist) - 1):
            try:
                de = hist[i].get("estado", "")
                a  = hist[i+1].get("estado", "")
                t1 = datetime.fromisoformat(hist[i]["fecha"])
                t2 = datetime.fromisoformat(hist[i+1]["fecha"])
                horas = (t2 - t1).total_seconds() / 3600
                meta = metas.get(f"{de}_{a}")
                if meta:
                    slas.append({"pedido": p.get("id",""), "trans": f"{de}→{a}", "horas": round(horas,1), "meta": meta, "ok": horas <= meta})
            except:
                pass
    total = len(slas) or 1
    cumple = sum(1 for s in slas if s["ok"])
    return slas, {"pct": round(cumple/total*100, 1), "criticos": total - cumple, "total": len(slas), "prom": round(sum(s["horas"] for s in slas)/total, 1)}


# ─── SIDEBAR ─────────────────────────────────────────────────────────────
def render_sidebar():
    st.sidebar.markdown("# 🚀 Export Haret v5.0")
    st.sidebar.markdown(f"**{st.session_state.user_nombre}** ({st.session_state.user_rol})")
    st.sidebar.markdown("---")
    pedidos = load_pedidos()
    clients = load_clients()
    st.sidebar.markdown("### 📊 Resumen")
    st.sidebar.metric("📦 Pedidos", len(pedidos))
    fac = sum(p.get('total_usd',0) for p in pedidos)
    st.sidebar.metric("💵 Facturación", f"${fac:,.0f}")
    st.sidebar.metric("👥 Clientes", len(clients))
    pendientes = len([p for p in pedidos if p.get('estado') in ['Recibido','Confirmado','Preparando']])
    st.sidebar.metric("⏳ En proceso", pendientes)
    st.sidebar.markdown("---")
    if st.sidebar.button("🚪 Cerrar Sesión", use_container_width=True):
        st.session_state.logged_in = False
        st.rerun()
    st.sidebar.markdown("---")
    st.sidebar.caption("Export Haret © 2026")

# ─── TAB DASHBOARD ─────────────────────────────────────────────
def tab_dashboard():
    st.markdown("## 📊 Dashboard Ejecutivo")
    pedidos = load_pedidos()
    clients = load_clients()
    data = load_precios()

    # KPIs
    c1,c2,c3,c4 = st.columns(4)
    total_fac = sum(p.get('total_usd',0) for p in pedidos)
    hoy_str = str(date.today())
    pedidos_hoy = len([p for p in pedidos if p.get('fecha','')[:10] == hoy_str])
    vip_count = sum(1 for e in clients if segmentar_cliente(e, clients)['segmento'] == 'VIP')
    c1.metric('📦 Total Pedidos', f'{len(pedidos):,}')
    c2.metric('💵 Facturación', f'${total_fac:,.0f}', 'USD')
    c3.metric('👥 Clientes', f'{len(clients):,}', f'{vip_count} VIP')
    c4.metric('📬 Hoy', pedidos_hoy, 'nuevos')

    st.markdown('---')
    # Estados
    st.markdown('### 📋 Pedidos por Estado')
    est_counts = {}
    for p in pedidos:
        est = p.get('estado','Recibido')
        est_counts[est] = est_counts.get(est,0) + 1
    if est_counts:
        cols = st.columns(len(ORDEN_ESTADOS))
        for i, est in enumerate(ORDEN_ESTADOS):
            cols[i].metric(f"{ESTADO_ICONS.get(est,'')} {est}", est_counts.get(est,0))
    else:
        st.info('ℹ️ No hay pedidos aun. Crea tu primer pedido en el tab “Hacer Pedido”')

    st.markdown('---')
    # SLA
    st.markdown('### ⏱ SLA - Tiempos de Proceso')
    _, sla_stats = calcular_sla(pedidos)
    s1,s2,s3,s4 = st.columns(4)
    s1.metric('✅ Cumplimiento', f"{sla_stats['pct']:.1f}%", 'Meta: 95%')
    s2.metric('⚠️ Críticos', sla_stats['criticos'])
    s3.metric('⏱ Prom. Horas', f"{sla_stats['prom']:.1f}h")
    s4.metric('📊 Transiciones', sla_stats['total'])
    if sla_stats['pct'] < 95 and sla_stats['total'] > 0:
        st.warning('SLA por debajo del 95%. Revisar pedidos críticos.')

    st.markdown('---')
    # Segmentacion
    st.markdown('### ⭐ Segmentación de Clientes')
    segs = {'VIP':0,'Regular':0,'Nuevo':0,'Inactivo':0}
    for e in clients:
        seg = segmentar_cliente(e, clients)['segmento']
        segs[seg] = segs.get(seg,0) + 1
    sg1,sg2,sg3,sg4 = st.columns(4)
    sg1.metric('⭐ VIP', segs.get('VIP',0), '+5% desc.')
    sg2.metric('⚫ Regular', segs.get('Regular',0), '+2% desc.')
    sg3.metric('🆕 Nuevo', segs.get('Nuevo',0))
    sg4.metric('⚠️ Inactivo', segs.get('Inactivo',0))

    # Productos activos
    st.markdown('---')
    st.markdown('### 📦 Productos Activos')
    prods = data.get('products',[])
    if prods:
        df_p = pd.DataFrame([{'Código': p.get('codigo',''), 'Producto': p.get('descripcion',''), 'Precio CIF': f"${p.get('precio_cif_usd',0):.2f}"} for p in prods[:10]])
        st.dataframe(df_p, use_container_width=True, hide_index=True)
    else:
        st.info('💡 Sube tu archivo Cotizaciones.xlsx en el tab Cotización para cargar productos y precios')

# ─── TAB HACER PEDIDO ───────────────────────────────────────────────
def tab_hacer_pedido():
    st.markdown('## 🛒 Crear Nuevo Pedido')
    data = load_precios()
    clients = load_clients()
    prods = data.get('products', [])
    config = data.get('config', {})
    destinos_cfg = config.get('destinos', {})

    if not prods:
        st.warning('⚠️ No hay productos cargados. Ve al tab **Cotización** y sube tu archivo Excel primero.')
        return

    # PASO 1: CLIENTE
    st.markdown('### 1️⃣ Datos del Cliente')
    col1, col2 = st.columns(2)
    with col1:
        client_email = st.text_input('📧 Email del cliente', placeholder='cliente@empresa.com', key='ped_email')
    with col2:
        client_name = st.text_input('👤 Nombre del cliente', placeholder='Nombre Empresa', key='ped_nombre')

    # Autocompletar si existe
    seg_info = None
    if client_email and client_email in clients:
        c = clients[client_email]
        seg_info = segmentar_cliente(client_email, clients)
        st.success(f"✅ Cliente encontrado: {c.get('nombre','')} | {seg_info['badge']} | Descuento: {seg_info['descuento']*100:.0f}%")
        if not client_name:
            client_name = c.get('nombre', '')
    elif client_email:
        st.info('🆕 Cliente nuevo - se registrará al guardar el pedido')

    # PASO 2: DESTINO
    st.markdown('### 2️⃣ Destino')
    dest_options = list(destinos_cfg.keys()) if destinos_cfg else ['Madrid/España','París/Francia','Londres/UK','Miami/USA']
    destino = st.selectbox('🌍 Destino', dest_options, key='ped_destino')
    dest_info = destinos_cfg.get(destino, {})
    moneda = dest_info.get('moneda', dest_info.get('currency', 'USD'))
    st.caption(f'Moneda: {moneda}')

    # PASO 3: PRODUCTOS
    st.markdown('### 3️⃣ Seleccionar Productos')
    if 'carrito' not in st.session_state:
        st.session_state.carrito = []

    prod_codigos = [p.get('codigo','') + ' - ' + p.get('descripcion','') for p in prods]
    col_prod, col_cajas, col_btn = st.columns([3, 1, 1])
    with col_prod:
        prod_sel_str = st.selectbox('Producto', [''] + prod_codigos, key='ped_prod_sel')
    with col_cajas:
        cajas_sel = st.number_input('Cajas', min_value=1, value=100, step=50, key='ped_cajas')
    with col_btn:
        st.markdown('<br>', unsafe_allow_html=True)
        if st.button('➕ Agregar', key='btn_add_prod') and prod_sel_str:
            codigo = prod_sel_str.split(' - ')[0]
            prod_data = next((p for p in prods if p.get('codigo') == codigo), {})
            precio = get_precio_producto(codigo, destino, data)
            if seg_info:
                precio = round(precio * (1 - seg_info['descuento']), 2)
            cajas_x_pallet = prod_data.get('cajas_pallet', 200)
            pallets = round(cajas_sel / cajas_x_pallet, 2) if cajas_x_pallet else 0
            item = {'codigo': codigo, 'producto': prod_data.get('descripcion',''), 'cajas': cajas_sel, 'pallets': pallets, 'precio_usd': precio, 'total': round(cajas_sel * precio, 2)}
            # Verificar si ya esta en carrito
            existing = next((i for i, x in enumerate(st.session_state.carrito) if x['codigo'] == codigo), None)
            if existing is not None:
                st.session_state.carrito[existing]['cajas'] += cajas_sel
                st.session_state.carrito[existing]['total'] = round(st.session_state.carrito[existing]['cajas'] * precio, 2)
            else:
                st.session_state.carrito.append(item)
            st.rerun()

    # MOSTRAR CARRITO
    if st.session_state.carrito:
        st.markdown('#### 🛒 Carrito')
        df_carrito = pd.DataFrame(st.session_state.carrito)
        st.dataframe(df_carrito[['codigo','producto','cajas','pallets','precio_usd','total']], use_container_width=True, hide_index=True)
        total_ped = sum(i['total'] for i in st.session_state.carrito)
        total_cajas = sum(i['cajas'] for i in st.session_state.carrito)
        total_pallets = sum(i['pallets'] for i in st.session_state.carrito)
        tc1,tc2,tc3 = st.columns(3)
        tc1.metric('📦 Total Cajas', f'{total_cajas:,}')
        tc2.metric('📍 Pallets', f'{total_pallets:.1f}')
        tc3.metric('💰 Total USD', f'${total_ped:,.2f}')

        if st.button('🗑️ Vaciar Carrito', key='vaciar_carrito'):
            st.session_state.carrito = []
            st.rerun()

    # PASO 4: NOTAS Y CONFIRMACION
    st.markdown('### 4️⃣ Notas y Confirmar')
    notas = st.text_area('Notas adicionales', placeholder='Instrucciones especiales, referencias...', key='ped_notas')

    if st.button('📤 GUARDAR PEDIDO', type='primary', use_container_width=True, key='btn_guardar_ped'):
        if not client_email:
            st.error('❌ Ingresa el email del cliente')
        elif not client_name:
            st.error('❌ Ingresa el nombre del cliente')
        elif not st.session_state.carrito:
            st.error('❌ Agrega al menos un producto al carrito')
        else:
            # Crear pedido
            pedido_id = 'PED-' + datetime.now().strftime('%Y%m%d-%H%M%S')
            total = sum(i['total'] for i in st.session_state.carrito)
            nuevo_pedido = {
                'id': pedido_id,
                'client_email': client_email,
                'client_name': client_name,
                'destino': destino,
                'moneda': moneda,
                'productos': list(st.session_state.carrito),
                'total_usd': round(total, 2),
                'estado': 'Recibido',
                'fecha': datetime.now().isoformat(),
                'notas': notas,
                'historial_estados': [{'estado': 'Recibido', 'fecha': datetime.now().isoformat(), 'usuario': st.session_state.user_email}],
                'creado_por': st.session_state.user_email,
            }
            # Guardar pedido
            todos = load_pedidos()
            todos.append(nuevo_pedido)
            save_pedidos(todos)
            # Registrar/actualizar cliente
            if client_email not in clients:
                clients[client_email] = {'nombre': client_name, 'email': client_email, 'fecha_registro': datetime.now().isoformat(), 'pedidos_ids': []}
            clients[client_email]['pedidos_ids'] = clients[client_email].get('pedidos_ids',[]) + [pedido_id]
            save_clients(clients)
            # Log email
            elog = load_email_log()
            elog.append({'id': f'EMAIL-{len(elog)+1:05d}', 'destinatario': client_email, 'asunto': f'Pedido {pedido_id} recibido', 'tipo': 'confirmacion', 'fecha': datetime.now().isoformat(), 'estado': 'simulado'})
            save_email_log(elog)
            # Limpiar carrito
            st.session_state.carrito = []
            st.success(f'✅ Pedido {pedido_id} creado exitosamente por ${total:,.2f} USD')
            st.balloons()
            st.cache_data.clear()

# ─── TAB GESTIÓN PEDIDOS ────────────────────────────────────────────
def tab_gestion_pedidos():
    st.markdown('## 📦 Gestión de Pedidos')
    pedidos = load_pedidos()

    # Filtros
    f1,f2,f3 = st.columns([2,2,2])
    filt_estado = f1.selectbox('Estado', ['Todos'] + ORDEN_ESTADOS, key='gp_estado')
    filt_cliente = f2.text_input('Buscar cliente/ID', key='gp_cliente')
    destinos_uniq = sorted(set(p.get('destino','') for p in pedidos if p.get('destino')))
    filt_dest = f3.selectbox('Destino', ['Todos'] + destinos_uniq, key='gp_dest')

    # Aplicar filtros
    filtrados = [p for p in pedidos
        if (filt_estado == 'Todos' or p.get('estado') == filt_estado)
        and (not filt_cliente or filt_cliente.lower() in (p.get('client_name','') + p.get('id','')).lower())
        and (filt_dest == 'Todos' or p.get('destino') == filt_dest)]

    st.markdown(f'**✅ {len(filtrados)} pedidos encontrados**')

    # Exportar Excel
    col_exp1, col_exp2 = st.columns([1,3])
    with col_exp1:
        if filtrados:
            excel_bytes = exportar_excel(filtrados)
            if excel_bytes:
                st.download_button('📥 Descargar Excel', data=excel_bytes, file_name=f'pedidos_{date.today()}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Listado con acciones
    st.markdown('---')
    for ped in sorted(filtrados, key=lambda x: x.get('fecha',''), reverse=True)[:50]:
        icon = ESTADO_ICONS.get(ped.get('estado',''), '📦')
        label = f"{icon} #{ped.get('id','').upper()} • {ped.get('client_name','N/A')} • {ped.get('destino','')} • ${ped.get('total_usd',0):,.2f} • {ped.get('estado','')}"
        with st.expander(label):
            cl1,cl2,cl3 = st.columns(3)
            cl1.markdown(f"**Cliente:** {ped.get('client_name','')}");    cl1.markdown(f"**Email:** {ped.get('client_email','')}")
            cl2.markdown(f"**Destino:** {ped.get('destino','')}");        cl2.markdown(f"**Fecha:** {ped.get('fecha','')[:10]}")
            cl3.markdown(f"**Total:** ${ped.get('total_usd',0):,.2f}");   cl3.markdown(f"**Estado:** {ped.get('estado','')}")
            if ped.get('productos'):
                st.dataframe(pd.DataFrame(ped['productos'])[['codigo','producto','cajas','pallets','precio_usd','total']], use_container_width=True, hide_index=True)
            if ped.get('notas'):
                st.markdown(f"**Notas:** {ped.get('notas','')}")
            # Cambiar estado
            st.markdown('**Cambiar Estado:**')
            act_col1, act_col2 = st.columns([2,1])
            with act_col1:
                nuevo_estado = st.selectbox('Nuevo estado', ORDEN_ESTADOS, index=ORDEN_ESTADOS.index(ped.get('estado','Recibido')) if ped.get('estado') in ORDEN_ESTADOS else 0, key=f'est_{ped["id"]}')
            with act_col2:
                st.markdown('<br>', unsafe_allow_html=True)
                if st.button('✅ Actualizar', key=f'upd_{ped["id"]}'):
                    todos = load_pedidos()
                    for i, p in enumerate(todos):
                        if p.get('id') == ped.get('id'):
                            todos[i]['estado'] = nuevo_estado
                            hist = todos[i].get('historial_estados', [])
                            hist.append({'estado': nuevo_estado, 'fecha': datetime.now().isoformat(), 'usuario': st.session_state.user_email})
                            todos[i]['historial_estados'] = hist
                            break
                    save_pedidos(todos)
                    st.success(f'✅ Estado actualizado a {nuevo_estado}')
                    st.cache_data.clear()
                    st.rerun()

# ─── TAB CLIENTES ────────────────────────────────────────────────────
def tab_clientes():
    st.markdown('## 👥 Base de Datos de Clientes')
    clients = load_clients()
    pedidos = load_pedidos()

    if not clients:
        st.info('🔎 No hay clientes registrados. Se crean automáticamente al hacer pedidos.')
        return

    # Segmentacion
    seg_data = []
    for email, c in clients.items():
        seg = segmentar_cliente(email, clients)
        mis_pedidos = [p for p in pedidos if p.get('client_email') == email]
        fac_total = sum(p.get('total_usd',0) for p in mis_pedidos)
        ultimo_pedido = max((p.get('fecha','')[:10] for p in mis_pedidos), default='-')
        seg_data.append({
            'Email': email,
            'Nombre': c.get('nombre',''),
            'Segmento': seg['badge'],
            'Pedidos': len(mis_pedidos),
            'Facturación': f"${fac_total:,.2f}",
            'Descuento': f"{seg['descuento']*100:.0f}%",
            'Crédito Máx': f"${seg['credito_max']:,.0f}",
            'Último Pedido': ultimo_pedido,
        })
    df = pd.DataFrame(seg_data)
    st.dataframe(df, use_container_width=True, hide_index=True)

    # Detalle de cliente
    st.markdown('---')
    st.markdown('### 🔎 Detalle de Cliente')
    sel_email = st.selectbox('Seleccionar cliente', [''] + list(clients.keys()), key='cli_sel')
    if sel_email:
        c = clients[sel_email]
        seg = segmentar_cliente(sel_email, clients)
        mis_peds = [p for p in pedidos if p.get('client_email') == sel_email]
        c1,c2 = st.columns(2)
        c1.markdown(f"**Nombre:** {c.get('nombre','')}")
        c1.markdown(f"**Email:** {sel_email}")
        c1.markdown(f"**Segmento:** {seg['badge']}")
        c2.markdown(f"**Pedidos totales:** {len(mis_peds)}")
        c2.markdown(f"**Facturación total:** ${sum(p.get('total_usd',0) for p in mis_peds):,.2f}")
        c2.markdown(f"**Descuento:** {seg['descuento']*100:.0f}%")
        if mis_peds:
            st.markdown('**Historial de Pedidos:**')
            df_peds = pd.DataFrame([{'ID': p.get('id',''), 'Destino': p.get('destino',''), 'Total': f"${p.get('total_usd',0):,.2f}", 'Estado': p.get('estado',''), 'Fecha': p.get('fecha','')[:10]} for p in sorted(mis_peds, key=lambda x: x.get('fecha',''), reverse=True)])
            st.dataframe(df_peds, use_container_width=True, hide_index=True)

# ─── TAB PRECIOS ──────────────────────────────────────────────────────
def tab_precios():
    st.markdown('## ⚙️ Administración de Precios')
    data = load_precios()
    prods = data.get('products', [])

    if not prods:
        st.info('⚠️ No hay productos. Sube el archivo Excel en el tab Cotización.')
        return

    st.markdown('### 📄 Tabla de Precios - Editable')
    df_edit = pd.DataFrame([{'Código': p.get('codigo',''), 'Descripción': p.get('descripcion',''), 'Precio CIF USD': p.get('precio_cif_usd', 0)} for p in prods])
    edited = st.data_editor(df_edit, use_container_width=True, hide_index=True, key='precios_editor', num_rows='fixed')

    if st.button('💾 Guardar Cambios de Precios', type='primary', key='btn_save_precios'):
        cambios = 0
        for i, row in edited.iterrows():
            codigo = row['Código']
            nuevo_precio = float(row['Precio CIF USD'])
            prod_orig = next((p for p in prods if p.get('codigo') == codigo), None)
            if prod_orig:
                precio_orig = prod_orig.get('precio_cif_usd', 0)
                if abs(nuevo_precio - precio_orig) > 0.001:
                    registrar_cambio_precio(codigo, precio_orig, nuevo_precio)
                    prod_orig['precio_cif_usd'] = nuevo_precio
                    cambios += 1
        if cambios > 0:
            save_precios(data)
            st.success(f'✅ {cambios} precios actualizados')
        else:
            st.info('No hubo cambios')

    # Historial de cambios
    st.markdown('---')
    st.markdown('### 📈 Historial de Cambios de Precios')
    hist = load_historial()
    if hist:
        df_h = pd.DataFrame(hist[-30:][::-1])
        df_h['fecha'] = pd.to_datetime(df_h['fecha']).dt.strftime('%d/%m/%Y %H:%M')
        st.dataframe(df_h[['fecha','producto','antes','despues','cambio_pct','usuario','motivo']].rename(columns={'fecha':'Fecha','producto':'Producto','antes':'Antes','despues':'Después','cambio_pct':'Cambio %','usuario':'Usuario','motivo':'Motivo'}), use_container_width=True, hide_index=True)
        h1,h2,h3 = st.columns(3)
        h1.metric('📝 Total Cambios', len(hist))
        h2.metric('📈 Aumentos', sum(1 for h in hist if h.get('cambio_pct',0) > 0))
        h3.metric('⚠️ Sospechosos >20%', sum(1 for h in hist if abs(h.get('cambio_pct',0)) > 20))
    else:
        st.info('Sin cambios de precios registrados')

    # Log de Emails
    st.markdown('---')
    st.markdown('### 📧 Log de Emails')
    elog = load_email_log()
    if elog:
        df_e = pd.DataFrame(elog[-20:][::-1])
        st.dataframe(df_e[['id','destinatario','asunto','tipo','fecha','estado']].rename(columns={'id':'ID','destinatario':'Para','asunto':'Asunto','tipo':'Tipo','fecha':'Fecha','estado':'Estado'}), use_container_width=True, hide_index=True)
    else:
        st.info('Sin emails registrados')

# ─── TAB COTIZACIÓN ───────────────────────────────────────────────────
def tab_cotizacion():
    st.markdown('## 📄 Cotización - Gestión de Datos')

    # Subir Excel
    st.markdown('### 📤 Cargar Archivo Excel')
    st.info('Sube tu archivo Cotizaciones.xlsx con las hojas: CONFIGURACION, TABLA PRECIOS, TODOS DESTINOS')
    uploaded = st.file_uploader('Selecciona tu archivo Excel', type=['xlsx','xls'], key='excel_uploader')

    if uploaded:
        try:
            # Guardar el archivo
            with open('Cotizaciones.xlsx', 'wb') as f:
                f.write(uploaded.getvalue())
            xl = pd.ExcelFile('Cotizaciones.xlsx')
            st.success(f'✅ Archivo cargado. Hojas: {", ".join(xl.sheet_names)}')

            # Leer configuración
            products = []
            destinos_cfg = {}
            grupos_cfg = {}

            sheet_names_lower = {s.lower(): s for s in xl.sheet_names}

            # Hoja de precios
            precio_sheet = sheet_names_lower.get('tabla precios', sheet_names_lower.get('precios', xl.sheet_names[0]))
            df_precios = pd.read_excel('Cotizaciones.xlsx', sheet_name=precio_sheet)
            df_precios.columns = [str(c).strip().lower() for c in df_precios.columns]
            for _, row in df_precios.dropna(subset=[df_precios.columns[0]]).iterrows():
                codigo = str(row.get('codigo', row.iloc[0])).strip()
                if not codigo or codigo == 'nan':
                    continue
                desc_col = next((c for c in df_precios.columns if 'desc' in c or 'producto' in c or 'nombre' in c), df_precios.columns[1] if len(df_precios.columns)>1 else 'descripcion')
                precio_col = next((c for c in df_precios.columns if 'precio' in c or 'cif' in c or 'usd' in c), df_precios.columns[2] if len(df_precios.columns)>2 else 'precio')
                cajas_col = next((c for c in df_precios.columns if 'caja' in c or 'pallet' in c), None)
                products.append({
                    'codigo': codigo,
                    'descripcion': str(row.get(desc_col, '')).strip(),
                    'precio_cif_usd': float(row.get(precio_col, 0) or 0),
                    'cajas_pallet': int(row.get(cajas_col, 200) or 200) if cajas_col else 200,
                    'activo': True,
                })

            # Hoja destinos
            dest_sheet_key = next((k for k in sheet_names_lower if 'destino' in k), None)
            if dest_sheet_key:
                df_dest = pd.read_excel('Cotizaciones.xlsx', sheet_name=sheet_names_lower[dest_sheet_key])
                df_dest.columns = [str(c).strip().lower() for c in df_dest.columns]
                for _, row in df_dest.dropna(subset=[df_dest.columns[0]]).iterrows():
                    dest_name = str(row.iloc[0]).strip()
                    if not dest_name or dest_name == 'nan': continue
                    moneda_col = next((c for c in df_dest.columns if 'moneda' in c or 'currency' in c), None)
                    factor_col = next((c for c in df_dest.columns if 'factor' in c or 'cif' in c or 'precio' in c), None)
                    destinos_cfg[dest_name] = {
                        'moneda': str(row.get(moneda_col, 'USD')).strip() if moneda_col else 'USD',
                        'factor': float(row.get(factor_col, 1.0) or 1.0) if factor_col else 1.0,
                    }

            # Guardar datos procesados
            data_nueva = {'products': products, 'config': {'destinos': destinos_cfg, 'grupos': grupos_cfg, 'minimos': {}}, 'pedidos': []}
            save_precios(data_nueva)
            st.success(f'✅ {len(products)} productos y {len(destinos_cfg)} destinos importados exitosamente!')
            st.rerun()

        except Exception as e:
            st.error(f'❌ Error procesando Excel: {e}')

    # Mostrar datos actuales
    data = load_precios()
    prods = data.get('products', [])
    destinos = data.get('config', {}).get('destinos', {})

    if prods:
        st.markdown('---')
        st.markdown(f'### 📋 Productos Identificados ({len(prods)}) - EDITABLE')
        df_p = pd.DataFrame([{'Código': p.get('codigo',''), 'Descripción': p.get('descripcion',''), 'Precio CIF USD': p.get('precio_cif_usd',0), 'Cajas/Pallet': p.get('cajas_pallet',200)} for p in prods])
        st.dataframe(df_p, use_container_width=True, hide_index=True)

    if destinos:
        st.markdown(f'### 🌍 Destinos Identificados ({len(destinos)}) - EDITABLE')
        df_d = pd.DataFrame([{'Destino': k, 'Moneda': v.get('moneda','USD'), 'Factor CIF': v.get('factor',1.0)} for k,v in destinos.items()])
        st.dataframe(df_d, use_container_width=True, hide_index=True)

# ─── TAB REPORTES ────────────────────────────────────────────────────
def tab_reportes():
    st.markdown('## 📊 Reportes y Analytics')
    pedidos = load_pedidos()
    clients = load_clients()

    if not pedidos:
        st.info('📊 No hay pedidos aun para analizar. Crea pedidos primero.')
        return

    # Filtro de fechas
    st.markdown('### 📅 Rango de Fechas')
    fc1, fc2 = st.columns(2)
    fecha_desde = fc1.date_input('Desde', value=date.today() - timedelta(days=30), key='rep_desde')
    fecha_hasta = fc2.date_input('Hasta', value=date.today(), key='rep_hasta')

    ped_filtro = [p for p in pedidos if str(fecha_desde) <= p.get('fecha','')[:10] <= str(fecha_hasta)]

    st.markdown(f'**{len(ped_filtro)} pedidos en el período seleccionado**')

    # KPIs del periodo
    k1,k2,k3,k4 = st.columns(4)
    fac_periodo = sum(p.get('total_usd',0) for p in ped_filtro)
    clientes_uniq = len(set(p.get('client_email','') for p in ped_filtro))
    ticket_prom = fac_periodo / len(ped_filtro) if ped_filtro else 0
    entregados = len([p for p in ped_filtro if p.get('estado') == 'Entregado'])
    k1.metric('💰 Facturación', f'${fac_periodo:,.0f}')
    k2.metric('📦 Pedidos', len(ped_filtro))
    k3.metric('🎫 Ticket Promedio', f'${ticket_prom:,.0f}')
    k4.metric('✅ Entregados', entregados)

    st.markdown('---')

    # Ventas por Destino
    st.markdown('### 🌍 Ventas por Destino')
    dest_data = {}
    for p in ped_filtro:
        dest = p.get('destino','Otros')
        dest_data[dest] = dest_data.get(dest, 0) + p.get('total_usd',0)
    if dest_data:
        df_dest = pd.DataFrame(list(dest_data.items()), columns=['Destino','Total USD']).sort_values('Total USD', ascending=False)
        st.dataframe(df_dest, use_container_width=True, hide_index=True)

    # Top clientes
    st.markdown('---')
    st.markdown('### 👑 Top 10 Clientes')
    cli_fac = {}
    for p in ped_filtro:
        email = p.get('client_email','')
        cli_fac[email] = cli_fac.get(email, 0) + p.get('total_usd',0)
    if cli_fac:
        top10 = sorted(cli_fac.items(), key=lambda x: x[1], reverse=True)[:10]
        df_top = pd.DataFrame([{'Cliente': clients.get(e,{}).get('nombre',e), 'Email': e, 'Facturación': f'${v:,.2f}'} for e,v in top10])
        st.dataframe(df_top, use_container_width=True, hide_index=True)

    # Productos mas vendidos
    st.markdown('---')
    st.markdown('### 📦 Productos Más Vendidos')
    prod_ventas = {}
    for p in ped_filtro:
        for prod in p.get('productos',[]):
            key = prod.get('codigo','') + ' - ' + prod.get('producto','')
            if key not in prod_ventas:
                prod_ventas[key] = {'cajas': 0, 'total': 0}
            prod_ventas[key]['cajas'] += prod.get('cajas',0)
            prod_ventas[key]['total'] += prod.get('total',0)
    if prod_ventas:
        top_prods = sorted(prod_ventas.items(), key=lambda x: x[1]['total'], reverse=True)[:10]
        df_pp = pd.DataFrame([{'Producto': k, 'Cajas': v['cajas'], 'Total USD': f'${v["total"]:,.2f}'} for k,v in top_prods])
        st.dataframe(df_pp, use_container_width=True, hide_index=True)

    # Tendencia por estado
    st.markdown('---')
    st.markdown('### 📈 Distribución por Estado')
    estado_res = {est: 0 for est in ORDEN_ESTADOS}
    for p in ped_filtro:
        est = p.get('estado','Recibido')
        estado_res[est] = estado_res.get(est,0) + 1
    df_est = pd.DataFrame([{'Estado': k, 'Cantidad': v, 'Icono': ESTADO_ICONS.get(k,'')} for k,v in estado_res.items() if v > 0])
    if not df_est.empty:
        st.dataframe(df_est, use_container_width=True, hide_index=True)

    # SLA en periodo
    st.markdown('---')
    st.markdown('### ⏱ SLA del Período')
    _, sla_stats = calcular_sla(ped_filtro)
    sl1,sl2,sl3,sl4 = st.columns(4)
    sl1.metric('✅ Cumplimiento', f"{sla_stats['pct']:.1f}%")
    sl2.metric('⚠️ Críticos', sla_stats['criticos'])
    sl3.metric('⏱ Prom. Horas', f"{sla_stats['prom']:.1f}h")
    sl4.metric('📊 Transiciones', sla_stats['total'])

    # Exportar reporte
    st.markdown('---')
    if ped_filtro:
        excel_rpt = exportar_excel(ped_filtro)
        if excel_rpt:
            st.download_button('📥 Exportar Reporte Excel (3 hojas)', data=excel_rpt, file_name=f'reporte_{fecha_desde}_{fecha_hasta}.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', type='primary')

# ─── MAIN ─────────────────────────────────────────────────────────────
def main():
    check_login()

    if not st.session_state.logged_in:
        login_page()
        return

    # Header principal
    st.markdown("""
    <div style='background: linear-gradient(90deg, #003E8C 0%, #0066CC 100%); padding: 16px 24px; border-radius: 8px; margin-bottom: 20px;'>
        <h2 style='color:white; margin:0;'>🚀 EXPORT HARET - Sistema de Gestión de Pedidos v5.0</h2>
    </div>
    """, unsafe_allow_html=True)

    render_sidebar()

    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        '📊 Dashboard',
        '🛒 Hacer Pedido',
        '📦 Mis Pedidos',
        '👥 Clientes',
        '⚙️ Precios',
        '📄 Cotización',
        '📊 Reportes',
    ])

    with tab1: tab_dashboard()
    with tab2: tab_hacer_pedido()
    with tab3: tab_gestion_pedidos()
    with tab4: tab_clientes()
    with tab5: tab_precios()
    with tab6: tab_cotizacion()
    with tab7: tab_reportes()

    # Footer
    st.markdown('---')
    st.markdown('<div style="text-align:center;color:#888;"><small>🚀 Export Haret v5.0 © 2026 | Sistema Profesional de Gestión de Pedidos</small></div>', unsafe_allow_html=True)

if __name__ == '__main__':
    main()
