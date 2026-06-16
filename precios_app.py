import streamlit as st
import pandas as pd
import json
import os
import io
import html
import hashlib
import uuid
import logging
from datetime import datetime, date, timedelta
from typing import Dict, List, Optional
from PIL import Image
try:
    import finanzas_sync  # sincronización con el hub Finanzas (cliente/cotización/envío)
except Exception:
    finanzas_sync = None
try:
    import outbox  # cola durable de pedidos (Gist) para recuperación cuando el Mac estaba apagado
except Exception:
    outbox = None
try:
    import theme  # sistema de diseño central (tema visual)
except Exception:
    theme = None
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ─── LOGGING SETUP ───────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(name)s: %(message)s')
logger = logging.getLogger('exportharet')

# ─── PAGE CONFIG ─────────────────────────────────────────────────────────────
# Logo dinamico: usar logo.png si existe, sino emoji generico
import os as _os_init
_PAGE_ICON = 'logo.png' if _os_init.path.exists('logo.png') else '📋'
st.set_page_config(
    page_title="Export Haret - Sistema de Pedidos",
    page_icon=_PAGE_ICON,
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ─── CONSTANTES ───────────────────────────────────────────────────────────────
ORDEN_ESTADOS = ["Recibido","Confirmado","Preparando","Enviado","Entregado","Cancelado"]
ESTADO_ICONS = {"Recibido":"📤","Confirmado":"✅","Preparando":"📦","Enviado":"🚚","Entregado":"✨","Cancelado":"❌"}

# Tramos de descuento por volumen (pallets)
TRAMOS_VOLUMEN = [
    {"min": 1,  "max": 2,  "descuento": 0.00, "label": "1-2 Pallets"},
    {"min": 3,  "max": 5,  "descuento": 0.10, "label": "3-5 Pallets (-10%)"},
    {"min": 6,  "max": 9,  "descuento": 0.12, "label": "6-9 Pallets (-12%)"},
    {"min": 10, "max": 19, "descuento": 0.14, "label": "10-19 Pallets (-14%)"},
    {"min": 20, "max": 9999, "descuento": 0.15, "label": "20+ Pallets (-15%)"},
]

def get_descuento_volumen(total_pallets):
    """Retorna el descuento (0.00-0.15) segun total de pallets."""
    try:
        _p = max(1, int(total_pallets))
    except Exception:
        _p = 1
    for _t in TRAMOS_VOLUMEN:
        if _t['min'] <= _p <= _t['max']:
            return float(_t.get('descuento', 0.0) or 0.0)
    return 0.0

def cajas_y_pallets(qty, unidad, cxp):
    """Fuente ÚNICA de la conversión cantidad↔(cajas, pallets).
    La usan el pre-pass, el bucle de productos y los tests, para que las tres
    cuenten exactamente igual (evita desfases). `unidad` es el valor canónico
    'Pallets' o 'Cajas'; `cxp` = cajas por pallet del grupo.
    Devuelve (cajas:int, pallets:float).
    """
    try:
        _cxp = max(int(cxp or 1), 1)
    except (TypeError, ValueError):
        _cxp = 1
    try:
        _q = float(qty or 0)
    except (TypeError, ValueError):
        _q = 0.0
    if _q <= 0:
        return 0, 0.0
    if unidad == 'Pallets':
        return int(_q * _cxp), float(_q)
    _cj = int(_q)
    return _cj, round(_cj / _cxp, 2)

MONEDAS = ["USD", "EUR", "GBP", "CHF", "AED", "CAD", "MXN", "BRL", "COP"]
MONEDA_SIMBOLO = {"USD": "$", "EUR": "€", "GBP": "£", "CHF": "Fr", "AED": "د.إ", "CAD": "CA$", "MXN": "MX$", "BRL": "R$", "COP": "COP$", "PEN": "S/", "CLP": "CLP$", "ARS": "AR$"}
# -- IDIOMA / LANGUAGE TRANSLATIONS --
LANG_TEXTS = {
    'es': {
        'step1': '### 1️⃣ Tus datos de contacto',
        'step2': '### 2️⃣ ¿Cómo y a dónde lo enviamos?',
        'step3': '### 3️⃣ Elige tus frutas y cantidades',
        'step4': '### 4️⃣ Revisa y envía tu pedido',
        'email_label': 'Tu correo electrónico',
        'email_ph': 'tu@empresa.com',
        'no_catalog': '⚠️ Catálogo no disponible. Contacte a order@exportharet.com',
        'enter_email': '**Ingresa tu correo electrónico arriba para continuar**',
        'welcome_back': 'Bienvenido de vuelta, **{name}**!',
        'not_registered': '✨ ¡Bienvenido! Eres nuevo aquí — completa tus datos para crear tu cuenta y empezar.',
        'edit_data': 'Editar mis datos',
        'client_ready': 'Listo para pedir',
        'edit_shipping': 'Cambiar modalidad de envío',
        'cart_restored': '🛒 Retomamos tu pedido donde lo dejaste — revisa las cantidades.',
        'err_email_friendly': 'Revisa el correo: parece que falta la @ o el dominio (ej. nombre@empresa.com).',
        'order_received_kick': 'Pedido recibido',
        'post_steps': 'Recibido|Confirmamos (24 h)|Preparación|Envío',
        'post_copy_hint': 'Guarda tu comprobante y, si quieres, confírmanos por WhatsApp o email:',
        'ship_fob_title': 'Precio en Ecuador (FOB)',
        'ship_fob_sub': 'Tú coordinas el transporte desde Quito/Guayaquil — sale más económico.',
        'ship_cif_title': 'Puesto en {dest} (flete incluido)',
        'ship_cif_sub': 'Nosotros enviamos hasta tu ciudad. Flete: ${flete:.2f} USD/kg.',
        'nombre_label': 'Nombre completo *',
        'empresa_label': 'Empresa (opcional)',
        'telefono_label': 'Teléfono / WhatsApp (opcional)',
        'telefono_ph': '+34 600 000 000',
        'pais_label': 'País',
        'auto_register': '🔒 Al guardar el pedido, tu cuenta quedará registrada automáticamente.',
        'tab_datos': 'Mis Datos',
        'tab_pedidos': '📦 Mis Pedidos ({n})',
        'no_orders': '📦 Aún no tienes pedidos. ¡Haz tu primer pedido a continuación!',
        'price_type_label': '¿Cómo quieres el precio?',
        'price_type_help': '«Puesto en tu ciudad» incluye el flete hasta tu destino (lo más común). «En Ecuador» es más barato pero tú organizas el transporte.',
        'dest_label': '¿A qué ciudad o país?',
        'no_dest': '⚠️ No hay destinos configurados',
        'flete_caption': '🛫 Flete ya incluido: **${flete:.2f} USD/Kilo** | {orig} → {dest}',
        'cif_info': '✅ El precio que ves **ya incluye el flete** hasta **{dest}**. Embarcamos desde **{orig}**.',
        'fob_info': '📦 **Precio en Ecuador** — no incluye el flete; tú coordinas el transporte desde Ecuador (sale más económico).',
 'price_update_notice': 'ℹ️ Los precios del catálogo se actualizan cada **martes**. Dudas: order@exportharet.com', 'btn_validate': 'Continuar →', 'btn_acceder': 'Continuar →', 'progress_step1': 'Tus datos', 'progress_step2': 'Entrega', 'progress_step3': 'Productos', 'progress_step4': 'Enviar', 'save_data_btn': '💾 Guardar datos', 'data_saved': 'Datos guardados', 'header_subtitle': 'Sistema de Pedidos — Frutas Exóticas Premium | order@exportharet.com', 'admin_access': '🔒 Acceso administración', 'download_catalog': '📥 Descargar Catálogo', 'sidebar_subtitle': 'Portal de Pedidos', 'sidebar_footer': 'Export Haret © 2026 | order@exportharet.com', 'footer_text': 'Export Haret © 2026 | order@exportharet.com | Frutas Exóticas Premium', 'cart_label': 'Tu pedido', 'cart_products': 'fruta(s)', 'min_progress_zero': '📋 Pedido mínimo: 3 pallets — Añade productos para comenzar', 'min_progress_short': '⚠️ {curr:.1f}/{min} pallets — Faltan {needed:.1f} pallets para el mínimo', 'min_progress_ok': '{curr:.0f} pallets ✓ (mín. {min})', 'min_progress_next': ' | Con {n}+ pallets el precio baja aún más 🚀', 'cart_summary_label': 'Pedido', 'group_summary': '📦 Resumen por grupo de embalaje', 'group_full_pallets': 'pallet(s) completo(s)', 'group_partial_boxes': 'cj parciales', 'group_total_pallets': '📊 Total pallets agrupados',
        'order_total_label': 'Total del pedido',
        'unit_pallets': 'pallets',
        'unit_boxes': 'cajas',
        'unit_products': 'producto(s)',
        'unit_kg_per_box': 'kg/caja', 'group_missing': '🎁 Te faltan <b>{n} cj</b> para completar 1 pallet del Grupo {g} y mejorar el precio', 'group_complete_with': ' · Completalo con: <b>{names}</b>', 'col_product': 'Producto', 'col_price': 'Precio/cja', 'col_qty': 'Cantidad', 'col_unit': 'Unidad', 'col_boxes': 'Cajas', 'unit_pallets': 'Pallets', 'unit_boxes': 'Cajas', 'min_from': 'Desde', 'added_mark': '✓ Agregado', 'saved_per_box': '💰 Ahorras', 'min_warning': '⚠️ Mínimo: <b>{n} {u}</b> para {p}', 'calc_title': '🧮 Calculadora de costo en otra moneda (referencial)', 'calc_convert_to': 'Convertir a:', 'calc_total_usd': 'Total en USD:', 'calc_total_dest': 'Total en {m}:', 'calc_rate': 'Tasa aplicada:', 'calc_per_pallet': 'Costo aprox. por pallet:', 'calc_note': 'ℹ Valor referencial. La transacción se realiza en USD.', 'order_summary_title': '📝 Resumen del Pedido', 'order_lbl_client': 'Cliente:', 'order_lbl_company': 'Empresa:', 'order_lbl_country': 'País:', 'order_lbl_mode': 'Modalidad:', 'order_lbl_payment': 'T. pago:', 'order_lbl_pending': 'Por confirmar', 'order_savings': '💰 Ahorro por volumen: ${s:,.2f} USD', 'order_total_label': 'TOTAL DEL PEDIDO', 'order_n_pallets': '📦 {n:.2f} pallets', 'order_n_boxes': '📋 {n:,} cajas', 'order_weight': '⚖️ {n:,.0f} kg net.', 'order_weight_note': '* kg net. de fruta — no incluye embalaje', 'savings_per_box': '💰 Ahorras ${d:.2f}/cj', 'unlock_better_price': '🎉 ¡Desbloqueaste mejor precio! Tramo: {label}', 'min_order_alert': '⚠️ Pedido mínimo: 3 pallets. Tienes {curr:.1f} pallets — añade {miss:.1f} pallets más para poder confirmar.', 'order_sent_email': 'Tu pedido ha sido enviado a <b>order@exportharet.com</b> para su confirmación. Nuestro equipo te contactará en 24-48h. Sigue el estado en la pestaña <b>Mis Pedidos</b> ↑', 'post_order_h3': 'Pedido {pid} enviado', 'post_order_thanks': '<b>Gracias {name}</b>, hemos recibido tu pedido correctamente.<br>📞 Nuestro equipo se pondrá en contacto contigo en <b>menos de 24 horas</b> para confirmar disponibilidad, precios finales y logística.<br>También recibirás confirmación por email cuando el pedido sea procesado.', 'wa_confirm': '💬 Confirmar por WhatsApp', 'em_send': 'Enviar por Email', 'wa_msg_greeting': 'Estimado equipo de Export Haret,', 'wa_msg_intro': 'Acabo de realizar el siguiente pedido a través del portal:', 'wa_msg_order': '📋 *Pedido', 'wa_msg_client': 'Cliente:', 'wa_msg_company': 'Empresa:', 'wa_msg_country': 'País:', 'wa_msg_details': '*Detalle de productos:*', 'wa_msg_total': '💰 *TOTAL', 'wa_msg_closing': 'Quedo a disposición para coordinar los detalles.', 'wa_msg_regards': 'Saludos.', 'em_subj': 'Pedido', 'em_body_intro': 'Adjunto el albarán del pedido', 'em_body_data': 'Datos del pedido:', 'em_body_client': '• Cliente:', 'em_body_company': 'Empresa:', 'em_body_incoterm': '• Incoterm:', 'em_body_total': '• Total:', 'em_body_products': 'Productos:', 'em_body_closing': 'Quedo a disposición para cualquier consulta.', 'em_body_regards': 'Saludos.', 'share_portal_title': '🔗 Compartir Portal Clientes', 'share_portal_caption': 'Envía este enlace a tus clientes:', 'share_msg': 'Hola, te invitamos a usar nuestro portal de pedidos Export Haret:', 'destination_fob': 'FOB (desde Ecuador)', 'order_status': 'Estado:', 'order_type': 'Tipo:', 'order_destination': 'Destino:', 'order_date': 'Fecha:', 'order_total_lbl': 'Total:', 'tracking_title': '📍 Seguimiento:', 'full_history': '📜 Historial completo', 'products_label': 'Productos:', 'btn_repeat': '🔄 Repetir', 'btn_repeat_help': 'Cargar al pedido', 'btn_cancel': '🗑️ Cancelar', 'order_repeat_loaded': 'Pedido {pid} cargado. Revisa y confirma.', 'order_no_products': 'Este pedido no tiene productos registrados.', 'confirm_cancel': '⚠️ ¿Confirmas la cancelación del pedido <b>{pid}</b>? Se notificará a nuestro equipo.', 'btn_yes_cancel': 'Sí, cancelar', 'btn_no': '❌ No', 'order_cancelled': 'Pedido {pid} cancelado. Notificación enviada a order@exportharet.com', 'enter_valid_email': 'Ingresa un correo electrónico válido para continuar.', 'enter_full_name': 'Ingresa tu nombre completo para continuar.', 'err_invalid_email': '❌ Formato de email inválido', 'fob_freight_caption': 'Flete: ${flete:.2f} USD/Kilo', 'price_in_dest': '💱 <b>Equiv. {m}</b> (ref.): <b style="color:#0c6e51">{sym}{tot:,.2f} {m}</b> <small style="color:#888">1 USD = {rate:.4f} {m}</small>', 'rate_info': '💱 1 USD = <b>{sym}{rate:.4f} {m}</b> — {live} | Fuente: {src} | Actualizado: {ts}', 'rate_info_sub': 'Precios en {m} son de referencia. La transacción se realiza en USD.', 'rate_live': '🟢 En vivo', 'rate_approx': '⚪ Aprox.', 'btn_change_client': '🔄 Cambiar cliente', 'err_email_format': '❌ Formato de email inválido', 'msg_click_validate': '👆 Haz clic en validar para continuar', 'order_cleared': 'Pedido vaciado',       'fob_origin': '📌 Origen de embarque: **Quito o Guayaquil, Ecuador**',
        'min_order_empty': '📋 <strong>Pedido mínimo: 3 pallets</strong> — Añade productos para comenzar tu pedido',
        'min_order_short': '📋 <strong>Pedido mínimo: 3 pallets</strong> — Tienes {plt:.1f} plt. Añade más productos.',
        'cart_fob': '📦 Precios FOB — El flete corre por tu cuenta desde Quito/Guayaquil, Ecuador',
        'clear_cart': '🗑️ Vaciar pedido',
        'notes_label': '📝 Notas / instrucciones especiales',
        'notes_ph': 'Ej: Entrega en almén frigorífico, embalaje especial...',
        'payment_terms': '📋 Términos de pago (opcional)',
        'confirm_btn': '📤 CONFIRMAR Y ENVIAR PEDIDO',
        'err_email': '❌ Ingresa tu correo electrónico',
        'err_nombre': '❌ Ingresa tu nombre completo',
        'err_cart': '❌ Añade al menos una fruta a tu pedido',
        'err_destino': '❌ Selecciona un destino para precio CIF',
        'order_confirmed': '**Pedido {pid} confirmado y enviado**',
        'order_email_sent': 'Recibirás confirmación en **{email}** en 24-48h.',
        'download_pdf': '⬇️ Descargar Albarán PDF',
        'new_order_btn': '🔄 Nuevo Pedido',
        'quote_expander': '💬 Solicitar cotización especial',
        'quote_intro': '**¿Necesitas presupuesto personalizado?** Rellena el formulario y te contactamos en 24h.',
        'quote_name_lbl': 'Tu nombre / empresa',
        'quote_name_ph': 'Nombre o empresa',
        'quote_msg_lbl': 'Mensaje adicional',
        'quote_msg_ph': 'Condiciones especiales...',
        'send_quote': '📤 Enviar solicitud',
        'price_update_notice': 'ℹ️ Los precios del catálogo se actualizan cada **martes**. Dudas: order@exportharet.com',
    },
    'en': {
        'step1': '### 1️⃣ Your Details',
        'step2': '### 2️⃣ Price Type & Destination',
        'step3': '### 3️⃣ Select your Products',
        'step4': '### 4️⃣ Confirm Order',
        'email_label': 'Your email address',
        'email_ph': 'you@company.com',
        'no_catalog': '⚠️ Catalogue not available. Contact order@exportharet.com',
        'enter_email': '**Enter your email address above to continue**',
        'welcome_back': 'Welcome back, **{name}**!',
        'not_registered': '✨ Welcome! You are new here — complete your details to create your account and get started.',
        'edit_data': 'Edit my details',
        'client_ready': 'Ready to order',
        'edit_shipping': 'Change shipping option',
        'cart_restored': '🛒 We picked up your order where you left off — review the quantities.',
        'err_email_friendly': 'Check the email: it looks like the @ or domain is missing (e.g. name@company.com).',
        'order_received_kick': 'Order received',
        'post_steps': 'Received|We confirm (24 h)|Preparation|Shipping',
        'post_copy_hint': 'Keep your receipt and, if you wish, confirm via WhatsApp or email:',
        'ship_fob_title': 'Price in Ecuador (FOB)',
        'ship_fob_sub': 'You arrange transport from Quito/Guayaquil — more economical.',
        'ship_cif_title': 'Delivered to {dest} (freight included)',
        'ship_cif_sub': 'We ship to your city. Freight: ${flete:.2f} USD/kg.',
        'nombre_label': 'Full name *',
        'empresa_label': 'Company',
        'telefono_label': 'Phone / WhatsApp',
        'telefono_ph': '+1 000 000 0000',
        'pais_label': 'Country',
        'auto_register': '🔒 By placing your order, your account will be registered automatically.',
        'tab_datos': 'My Details',
        'tab_pedidos': '📦 My Orders ({n})',
        'no_orders': '📦 No orders yet. Place your first order below!',
        'price_type_label': '💲 Price type',
        'price_type_help': 'FOB = Ex-works price (no freight). CIF = Price includes freight to destination.',
        'dest_label': 'Destination',
        'no_dest': '⚠️ No destinations configured',
        'flete_caption': '🛫 Freight included: **${flete:.2f} USD/Kilo** | {orig} → {dest}',
        'cif_info': '📍 **Incoterm CIF** — Price includes cost + freight to **{dest}**. We ship from **{orig}**.',
        'fob_info': '📦 **FOB (Free On Board)** — Price **does not include freight**. You arrange transport from Ecuador.',
        'fob_origin': '📌 Port of origin: **Quito or Guayaquil, Ecuador**',
        'min_order_empty': '📋 <strong>Minimum order: 3 pallets</strong> — Add products to start your order',
        'min_order_short': '📋 <strong>Minimum order: 3 pallets</strong> — You have {plt:.1f} plt. Add more products.',
        'cart_fob': '📦 FOB prices — Freight is your responsibility from Quito/Guayaquil, Ecuador',
        'clear_cart': '🗑️ Clear order',
        'notes_label': '📝 Notes / special instructions',
        'notes_ph': 'e.g. Delivery to cold storage, special packaging...',
        'payment_terms': '📋 Payment terms (optional)',
        'confirm_btn': '📤 CONFIRM AND SEND ORDER',
        'err_email': '❌ Please enter your email address',
        'err_nombre': '❌ Please enter your full name',
        'err_cart': '❌ Please add at least one fruit to your order',
        'err_destino': '❌ Please select a destination for CIF pricing',
        'order_confirmed': '**Order {pid} confirmed and sent**',
        'order_email_sent': 'Confirmation will be sent to **{email}** within 24-48h.',
        'download_pdf': '⬇️ Download Order PDF',
        'new_order_btn': '🔄 New Order',
        'quote_expander': '💬 Request special quote',
        'quote_intro': '**Need a custom quote?** Fill in the form and we will contact you within 24h.',
        'quote_name_lbl': 'Your name / company',
        'quote_name_ph': 'Your name or company',
        'quote_msg_lbl': 'Additional message',
        'quote_msg_ph': 'Special conditions...',
        'send_quote': '📤 Send request',
        'price_update_notice': 'ℹ️ Catalogue prices are updated every **Tuesday**. Questions: order@exportharet.com', 'btn_validate': 'Validate / Access', 'btn_change_client': '🔄 Change client', 'err_email_format': '❌ Invalid email format', 'msg_click_validate': '👆 Click validate to continue', 'order_cleared': 'Order cleared', 'btn_validate': 'Validate', 'btn_acceder': 'Validate', 'progress_step1': 'Your Details', 'progress_step2': 'Price & Destination', 'progress_step3': 'Products', 'progress_step4': 'Confirm', 'save_data_btn': '💾 Save details', 'data_saved': 'Details saved', 'header_subtitle': 'Order System — Premium Exotic Fruits | order@exportharet.com', 'admin_access': '🔒 Admin access', 'download_catalog': '📥 Download Catalog', 'sidebar_subtitle': 'Order Portal', 'sidebar_footer': 'Export Haret © 2026 | order@exportharet.com', 'footer_text': 'Export Haret © 2026 | order@exportharet.com | Premium Exotic Fruits', 'cart_label': 'Your order', 'cart_products': 'fruit(s)', 'min_progress_zero': '📋 Minimum order: 3 pallets — Add products to start', 'min_progress_short': '⚠️ {curr:.1f}/{min} pallets — {needed:.1f} more pallets needed', 'min_progress_ok': '{curr:.0f} pallets ✓ (mín. {min})', 'min_progress_next': ' | With {n}+ pallets the price goes even lower 🚀', 'cart_summary_label': 'Order', 'group_summary': '📦 Summary by packaging group', 'group_full_pallets': 'full pallet(s)', 'group_partial_boxes': 'partial boxes', 'group_total_pallets': '📊 Total grouped pallets',
        'order_total_label': 'Order Total',
        'unit_pallets': 'pallets',
        'unit_boxes': 'boxes',
        'unit_products': 'product(s)',
        'unit_kg_per_box': 'kg/box', 'group_missing': '🎁 You need <b>{n} bx</b> more to complete 1 pallet in Group {g} and improve the price', 'group_complete_with': ' · Complete with: <b>{names}</b>', 'col_product': 'Product', 'col_price': 'Price/box', 'col_qty': 'Quantity', 'col_unit': 'Unit', 'col_boxes': 'Boxes', 'unit_pallets': 'Pallets', 'unit_boxes': 'Boxes', 'min_from': 'From', 'added_mark': '✓ Added', 'saved_per_box': '💰 You save', 'min_warning': '⚠️ Minimum: <b>{n} {u}</b> for {p}', 'calc_title': '🧮 Cost calculator in another currency (reference)', 'calc_convert_to': 'Convert to:', 'calc_total_usd': 'Total in USD:', 'calc_total_dest': 'Total in {m}:', 'calc_rate': 'Applied rate:', 'calc_per_pallet': 'Approx. cost per pallet:', 'calc_note': 'ℹ Reference value. Transactions are made in USD.', 'order_summary_title': '📝 Order Summary', 'order_lbl_client': 'Customer:', 'order_lbl_company': 'Company:', 'order_lbl_country': 'Country:', 'order_lbl_mode': 'Mode:', 'order_lbl_payment': 'Payment:', 'order_lbl_pending': 'To be confirmed', 'order_savings': '💰 Volume savings: ${s:,.2f} USD', 'order_total_label': 'ORDER TOTAL', 'order_n_pallets': '📦 {n:.2f} pallets', 'order_n_boxes': '📋 {n:,} boxes', 'order_weight': '⚖️ {n:,.0f} kg net', 'order_weight_note': '* kg net of fruit — does not include packaging', 'savings_per_box': '💰 You save ${d:.2f}/bx', 'unlock_better_price': '🎉 You unlocked a better price! Tier: {label}', 'min_order_alert': '⚠️ Minimum order: 3 pallets. You have {curr:.1f} pallets — add {miss:.1f} more pallets to confirm.', 'order_sent_email': 'Your order has been sent to <b>order@exportharet.com</b> for confirmation. Our team will contact you within 24-48h. Track the status in the <b>My Orders</b> tab ↑', 'post_order_h3': 'Order {pid} sent', 'post_order_thanks': '<b>Thank you {name}</b>, we have received your order correctly.<br>📞 Our team will contact you within <b>less than 24 hours</b> to confirm availability, final prices and logistics.<br>You will also receive an email confirmation when the order is processed.', 'wa_confirm': '💬 Confirm via WhatsApp', 'em_send': 'Send by Email', 'wa_msg_greeting': 'Dear Export Haret team,', 'wa_msg_intro': 'I have just placed the following order through the portal:', 'wa_msg_order': '📋 *Order', 'wa_msg_client': 'Customer:', 'wa_msg_company': 'Company:', 'wa_msg_country': 'Country:', 'wa_msg_details': '*Product details:*', 'wa_msg_total': '💰 *TOTAL', 'wa_msg_closing': 'I remain available to coordinate the details.', 'wa_msg_regards': 'Best regards.', 'em_subj': 'Order', 'em_body_intro': 'Attached the delivery note for order', 'em_body_data': 'Order data:', 'em_body_client': '• Customer:', 'em_body_company': 'Company:', 'em_body_incoterm': '• Incoterm:', 'em_body_total': '• Total:', 'em_body_products': 'Products:', 'em_body_closing': 'I remain available for any question.', 'em_body_regards': 'Best regards.', 'share_portal_title': '🔗 Share Customer Portal', 'share_portal_caption': 'Send this link to your customers:', 'share_msg': 'Hello, we invite you to use our Export Haret order portal:', 'destination_fob': 'FOB (from Ecuador)', 'order_status': 'Status:', 'order_type': 'Type:', 'order_destination': 'Destination:', 'order_date': 'Date:', 'order_total_lbl': 'Total:', 'tracking_title': '📍 Tracking:', 'full_history': '📜 Full history', 'products_label': 'Products:', 'btn_repeat': '🔄 Repeat', 'btn_repeat_help': 'Load to order', 'btn_cancel': '🗑️ Cancel', 'order_repeat_loaded': 'Order {pid} loaded. Review and confirm.', 'order_no_products': 'This order has no registered products.', 'confirm_cancel': '⚠️ Confirm cancellation of order <b>{pid}</b>? Our team will be notified.', 'btn_yes_cancel': 'Yes, cancel', 'btn_no': '❌ No', 'order_cancelled': 'Order {pid} cancelled. Notification sent to order@exportharet.com', 'enter_valid_email': 'Please enter a valid email address to continue.', 'enter_full_name': 'Please enter your full name to continue.', 'err_invalid_email': '❌ Invalid email format', 'fob_freight_caption': 'Freight: ${flete:.2f} USD/Kilo', 'price_in_dest': '💱 <b>Equiv. {m}</b> (ref.): <b style="color:#0c6e51">{sym}{tot:,.2f} {m}</b> <small style="color:#888">1 USD = {rate:.4f} {m}</small>', 'rate_info': '💱 1 USD = <b>{sym}{rate:.4f} {m}</b> — {live} | Source: {src} | Updated: {ts}', 'rate_info_sub': 'Prices in {m} are for reference. Transactions are made in USD.', 'rate_live': '🟢 Live', 'rate_approx': '⚪ Approx.',
    }
}

# ── CSS personalizado y responsive ───────────────────────────────────────────
CUSTOM_CSS = """
<style>
/* === RESPONSIVE MÓVIL === */
@media (max-width: 768px) {
    .stApp { font-size: 14px !important; }
    div[data-testid="column"] { min-width: 0 !important; }
    .stNumberInput input { font-size: 16px !important; }
    .stSelectbox select { font-size: 16px !important; }
    .product-row { flex-wrap: wrap !important; }
}
/* === MINI CARRITO FLOTANTE === */
.cart-badge {
    background: #0c6e51;
    color: white;
    border-radius: 12px;
    padding: 2px 10px;
    font-weight: bold;
    font-size: 0.85em;
}
/* === PROGRESS BAR MÍNIMO === */
.min-progress {
    height: 8px;
    background: #e0e0e0;
    border-radius: 4px;
    margin: 4px 0 8px 0;
}
.min-progress-fill {
    height: 100%;
    background: linear-gradient(90deg, #10a37a, #0c6e51);
    border-radius: 4px;
    transition: width 0.3s ease;
}
/* === BOTONES === */
.stButton > button[kind="primary"] { background: #0c6e51 !important; border-color: #0c6e51 !important; }
.stButton > button[kind="primary"]:hover { background: #084a37 !important; }
/* === RESUMEN PEDIDO === */
.order-summary-box {
    background: #f8faff;
    border: 1.5px solid #0c6e51;
    border-radius: 10px;
    padding: 16px;
    margin: 12px 0;
}
/* === PRODUCT ROW === */
.product-header { font-weight: 600; color: #555; font-size: 0.8em; border-bottom: 1px solid #eee; padding-bottom: 4px; margin-bottom: 4px; }
/* === LANG BUTTONS === */
div[data-testid="column"]:has(> div > div > button[title="Español"]),
div[data-testid="column"]:has(> div > div > button[title="English"]) {
    padding: 0 2px !important;
}
</style>
"""

PEDIDOS_FILE = "pedidos_data.json"
CLIENTS_FILE = "clientes.json"
HIST_FILE    = "precio_historial.json"
EMAIL_FILE   = "email_log.json"
DATA_FILE    = "precios_data.json"
APP_CONFIG_FILE = "app_config.json"
ACCESOS_FILE   = "accesos_log.json"
MIN_LOG_FILE   = "min_cambios_log.json"

USERS = {
    "admin@exportharet.com":  {"pwd": hashlib.md5(b"admin123").hexdigest(),  "rol": "admin",  "nombre": "Administrador"},
    "ventas@exportharet.com": {"pwd": hashlib.md5(b"ventas123").hexdigest(), "rol": "ventas", "nombre": "Ventas"},
}

# ─── DATA HELPERS ───────────────────────────────────────────────────
def _load(path, default):
    try:
        if os.path.exists(path):
            with open(path,'r',encoding='utf-8') as f: return json.load(f)
    except Exception as e:
        logger.warning(f'Error loading {path}: {e}')
    return default

def _save(path, data):
    # C2: Atomic write to prevent data corruption on interrupted writes
    import tempfile as _tmpf
    try:
        _dirn = os.path.dirname(os.path.abspath(path)) or '.'
        _fd, _tmp_path = _tmpf.mkstemp(prefix='.eh_', suffix='.tmp', dir=_dirn)
        try:
            with os.fdopen(_fd, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
                f.flush()
                try: os.fsync(f.fileno())
                except Exception: pass
            os.replace(_tmp_path, path)
            return True
        except Exception:
            try: os.unlink(_tmp_path)
            except Exception: pass
            raise
    except Exception as e:
        logger.error(f'Error guardando {path}: {e}')
        try: st.error(f'Error guardando: {e}')
        except Exception: pass
        return False

@st.cache_data(ttl=60)
def load_data():
    return _load(DATA_FILE, {'products':[],'config':{'destinos':{},'grupos':{},'minimos':{}},'pedidos':[]})

def load_clients(): return _load(CLIENTS_FILE, {})
def load_pedidos(): return _load(PEDIDOS_FILE, [])
def load_historial(): return _load(HIST_FILE, [])
def load_email_log(): return _load(EMAIL_FILE, [])
def load_app_config():
    return _load(APP_CONFIG_FILE, {"app_title": "📊 Export Haret — Panel de Administración"})
def save_app_config(cfg): _save(APP_CONFIG_FILE, cfg)
def save_data(d): _save(DATA_FILE,d); st.cache_data.clear()
def save_clients(c): _save(CLIENTS_FILE,c)
def save_pedidos(p):
    _save(PEDIDOS_FILE, p)
    if outbox:
        try:
            outbox.publish(p)  # vuelca a la cola durable (Gist); no-op si no está configurado
        except Exception as e:
            logger.warning(f'outbox falló: {e}')
def sync_finanzas(ped, todos=None):
    """Empuja el pedido al hub Finanzas. Nunca rompe el flujo si falla."""
    if not finanzas_sync:
        return
    try:
        finanzas_sync.sync_pedido(ped)
        if todos is not None:
            save_pedidos(todos)  # re-persistir con los ids de Finanzas
    except Exception as e:
        logger.warning(f'sync Finanzas falló: {e}')
def hidratar_pedidos_gist():
    """#3: tras un reinicio de Streamlit Cloud (disco efímero) recupera los pedidos
    desde el Gist, para que el cliente vuelva a ver 'Mis Pedidos'. Una vez por sesión."""
    if not outbox or st.session_state.get('_hidratado_gist'):
        return
    st.session_state['_hidratado_gist'] = True
    try:
        # 1) Pedidos
        remoto = outbox.fetch()  # None si no hay gist configurado; basta el gist_id para leer
        if remoto:
            local = load_pedidos()
            ids = {p.get('id') for p in local}
            nuevos = [p for p in remoto if p.get('id') and p.get('id') not in ids]
            if nuevos:
                local.extend(nuevos)
                _save(PEDIDOS_FILE, local)
                st.cache_data.clear()
                logger.info(f'hidratados {len(nuevos)} pedidos desde el Gist')
        # 2) Padrón de clientes del portal (incluye los pre-registrados por el admin)
        remoto_cli = outbox.fetch_clients()
        if remoto_cli:
            local_cli = load_portal_clients()
            faltan = {e: v for e, v in remoto_cli.items() if e not in local_cli}
            if faltan:
                local_cli.update(faltan)
                _save(PORTAL_CLIENTS_FILE, local_cli)
                logger.info(f'hidratados {len(faltan)} clientes desde el Gist')
        # 3) Carritos pendientes: el Gist es la verdad durable (sobrevive a reinicios).
        #    Permite "sigue donde lo dejaste" aunque el contenedor de Cloud se reinicie.
        if hasattr(outbox, 'fetch_carts'):
            remoto_carts = outbox.fetch_carts()
            if remoto_carts is not None:
                _save(PORTAL_CARTS_FILE, remoto_carts)
                logger.info(f'hidratados {len(remoto_carts)} carritos desde el Gist')
    except Exception as e:
        logger.warning(f'hidratar gist falló: {e}')
def _esc(s):
    """#5: escapa texto del usuario antes de meterlo en HTML/email (evita XSS)."""
    return html.escape(str(s if s is not None else ''))
def save_historial(h2): _save(HIST_FILE,h2)
def save_email_log(e): _save(EMAIL_FILE,e)
def load_accesos(): return _load(ACCESOS_FILE, [])
def save_accesos(data): _save(ACCESOS_FILE, data)
def registrar_acceso(email, nombre, rol):
    acc = load_accesos()
    acc.append({"fecha_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "email": email, "nombre": nombre, "rol": rol})
    save_accesos(acc)


def load_min_log(): return _load(MIN_LOG_FILE, [])
def save_min_log(data): _save(MIN_LOG_FILE, data)

# ─── SESIÓN ADMIN PERSISTENTE (sobrevive al refresco; caduca por inactividad) ──
ADMIN_SESSIONS_FILE = 'admin_sessions.json'
_ADMIN_SESSION_TTL_S = 5 * 3600   # 5 h de inactividad

def _admin_sessions_load():
    return _load(ADMIN_SESSIONS_FILE, {})

def _admin_sessions_purge(s):
    """Quita sesiones caducadas por inactividad."""
    _now = datetime.now()
    out = {}
    for _k, _v in (s or {}).items():
        try:
            if (_now - datetime.fromisoformat(_v.get('last', ''))).total_seconds() < _ADMIN_SESSION_TTL_S:
                out[_k] = _v
        except (ValueError, TypeError):
            continue
    return out

def admin_session_new(email, nombre, rol):
    """Crea un token de sesión y lo guarda. Devuelve el token (va en la URL)."""
    import uuid as _uuid
    tok = _uuid.uuid4().hex
    s = _admin_sessions_purge(_admin_sessions_load())
    s[tok] = {'email': email, 'nombre': nombre, 'rol': rol, 'last': datetime.now().isoformat()}
    _save(ADMIN_SESSIONS_FILE, s)
    return tok

def admin_session_resume(tok):
    """Si el token existe y no caducó por inactividad, devuelve sus datos; si no, None."""
    if not tok:
        return None
    s = _admin_sessions_load()
    rec = s.get(tok)
    if not isinstance(rec, dict):
        return None
    try:
        _last = datetime.fromisoformat(rec.get('last', ''))
    except (ValueError, TypeError):
        return None
    if (datetime.now() - _last).total_seconds() >= _ADMIN_SESSION_TTL_S:
        s.pop(tok, None); _save(ADMIN_SESSIONS_FILE, s)
        return None
    return rec

def admin_session_touch(tok):
    """Renueva la última actividad (desliza la ventana de 5 h). No escribe más de
    una vez por minuto para no machacar el disco en cada rerun."""
    if not tok:
        return
    import streamlit as _st
    _now = datetime.now()
    _last_touch = _st.session_state.get('_admin_tok_touched')
    if _last_touch and (_now - _last_touch).total_seconds() < 60:
        return
    s = _admin_sessions_load()
    if tok in s:
        s[tok]['last'] = _now.isoformat()
        _save(ADMIN_SESSIONS_FILE, s)
        _st.session_state['_admin_tok_touched'] = _now

def admin_session_end(tok):
    if not tok:
        return
    s = _admin_sessions_load()
    if s.pop(tok, None) is not None:
        _save(ADMIN_SESSIONS_FILE, s)


# ─── AUTH ────────────────────────────────────────────────────────────────
def init_session():
    defaults = {'logged_in':False,'user_email':'','user_rol':'','user_nombre':'','carrito':[]}
    for k,v in defaults.items():
        if k not in st.session_state: st.session_state[k] = v

def login_page():
    # Login premium centrado (coherente con el portal Pro).
    import os as _osL
    _l1, _l2, _l3 = st.columns([1, 1.5, 1])
    with _l2:
        st.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
        if _osL.path.exists('logo.png'):
            _lc1, _lc2, _lc3 = st.columns([1, 2, 1])
            with _lc2:
                st.image('logo.png', use_container_width=True)
        else:
            st.markdown('<div style="text-align:center"><h1 style="margin:0;color:#084a37">Export Haret</h1></div>', unsafe_allow_html=True)
        st.markdown(
            '<div style="text-align:center;margin:8px 0 16px">'
            '<div style="font-size:.72rem;letter-spacing:2px;text-transform:uppercase;color:#0c6e51;font-weight:700">Panel de administración</div>'
            '<div style="font-weight:800;color:#14201a;font-size:1.5rem;letter-spacing:-.5px;margin-top:4px">Inicia sesión</div>'
            '</div>', unsafe_allow_html=True)
        with st.form('admin_login_form', clear_on_submit=False):
            email = st.text_input('Email')
            pwd = st.text_input('Contraseña', type='password')
            _ok = st.form_submit_button('Entrar →', use_container_width=True, type='primary')
        if _ok:
            h = hashlib.md5(pwd.encode()).hexdigest()
            if email in USERS and USERS[email]['pwd'] == h:
                st.session_state.logged_in = True
                st.session_state.user_email = email
                st.session_state.user_rol = USERS[email]['rol']
                st.session_state.user_nombre = USERS[email]['nombre']
                # Sesión persistente: token en la URL para sobrevivir al refresco
                _tok = admin_session_new(email, USERS[email]['nombre'], USERS[email]['rol'])
                st.session_state['_admin_tok'] = _tok
                st.query_params['view'] = 'admin'
                st.query_params['s'] = _tok
                registrar_acceso(email=email, nombre=USERS[email]['nombre'], rol=USERS[email]['rol'])
                st.rerun()
            else:
                st.error('❌ Email o contraseña incorrectos')
        st.markdown('<div style="text-align:center;color:#8a978f;font-size:.8rem;margin:10px 0 4px">🔒 Acceso restringido al personal autorizado</div>', unsafe_allow_html=True)
        if st.button('← Volver al portal de clientes', key='login_back', use_container_width=True):
            st.session_state.app_mode = 'portal'
            st.query_params.clear()
            st.rerun()


def _admin_css():
    """Refinamiento visual del panel admin (coherente con el portal Pro):
    métricas como tarjetas, pestañas y separadores limpios. Se inyecta una vez."""
    st.markdown('''<style>
      /* El admin necesita más ancho que el portal (tablas, catálogo, pedidos) */
      [data-testid="stAppViewContainer"] .block-container{ max-width:1180px !important; }
      /* KPI / métricas del área principal → tarjetas premium */
      [data-testid="stMain"] [data-testid="stMetric"]{
        background:#fff; border:1px solid #e7eaef; border-radius:14px;
        padding:13px 16px; box-shadow:0 1px 2px rgba(20,60,40,.04); }
      [data-testid="stMain"] [data-testid="stMetricLabel"] p{ color:#65726b; font-weight:600; font-size:.84rem; }
      [data-testid="stMain"] [data-testid="stMetricValue"]{ color:#084a37; font-weight:800; }
      /* Pestañas admin más sobrias */
      .stTabs [data-baseweb="tab-list"]{ gap:2px; flex-wrap:wrap; }
      .stTabs [data-baseweb="tab"]{ padding:8px 14px; font-size:.92rem; }
      /* Separadores finos y con menos aire (había dobles ---) */
      [data-testid="stMain"] hr{ margin:.7rem 0; border-top:1px solid #eef2ef; }
      /* Sidebar: métricas compactas sin caja */
      [data-testid="stSidebar"] [data-testid="stMetric"]{ padding:2px 0; }
      [data-testid="stSidebar"] [data-testid="stMetricValue"]{ font-size:1.15rem; color:#084a37; }
    </style>''', unsafe_allow_html=True)


def _admin_seccion(titulo, icono=''):
    """Cabecera de sección admin uniforme (eyebrow + título), estilo Pro."""
    _t = f'{icono} {titulo}'.strip()
    st.markdown(
        '<div style="display:flex;align-items:center;gap:9px;margin:18px 0 10px">'
        '<span style="width:22px;height:2px;background:#0c6e51;border-radius:2px"></span>'
        f'<span style="font-weight:800;color:#14201a;font-size:1.12rem;letter-spacing:-.3px">{_esc(_t)}</span>'
        '</div>', unsafe_allow_html=True)


# ── Componentes "Finanzas": badge de estado, FSM timeline, timeline de eventos ──
_ESTADO_STYLE = {
    'Recibido':   ('#eef2f6', '#566472', '#dce1e8'),
    'Confirmado': ('#eef6f2', '#0b5a42', 'rgba(12,110,81,.20)'),
    'Preparando': ('#fbf6ea', '#7e560a', 'rgba(181,121,10,.22)'),
    'Enviado':    ('#eef1fb', '#3730a3', 'rgba(79,70,229,.20)'),
    'Entregado':  ('#e7f4ee', '#0b5a42', 'rgba(12,110,81,.30)'),
    'Cancelado':  ('#fcf2f1', '#992c20', 'rgba(185,28,28,.20)'),
}

def estado_badge(estado):
    """Pill de estado coloreado (estilo Finanzas)."""
    _bg, _tx, _bd = _ESTADO_STYLE.get(estado, ('#eef2f6', '#566472', '#dce1e8'))
    _ic = ESTADO_ICONS.get(estado, '📦')
    return (f'<span style="display:inline-flex;align-items:center;gap:5px;background:{_bg};color:{_tx};'
            f'border:1px solid {_bd};border-radius:999px;padding:3px 11px;font-size:.74rem;font-weight:700;'
            f'letter-spacing:.01em;white-space:nowrap">{_ic} {_esc(estado)}</span>')

def fsm_timeline_html(estado):
    """Ciclo de vida del pedido (Recibido→…→Entregado): anteriores con ✓, actual
    resaltado y pulsante, futuros en gris. Cancelado → aviso. Estilo Finanzas FSM."""
    pasos = ['Recibido', 'Confirmado', 'Preparando', 'Enviado', 'Entregado']
    if estado == 'Cancelado':
        return ('<div style="background:#fcf2f1;border:1px solid rgba(185,28,28,.20);color:#992c20;'
                'border-radius:14px;padding:12px 16px;font-size:.85rem;font-weight:600;margin:8px 0 4px">'
                '❌ Pedido cancelado</div>')
    _idx = pasos.index(estado) if estado in pasos else 0
    _seg = ''
    for _i, _p in enumerate(pasos):
        _done = _i < _idx
        _cur = _i == _idx
        if _done:
            _circ = 'background:#0c6e51;border-color:#0c6e51;color:#fff'
            _mark = '✓'
        elif _cur:
            _circ = 'background:#0c6e51;border-color:#0c6e51;color:#fff;box-shadow:0 0 0 4px rgba(12,110,81,.16)'
            _mark = ESTADO_ICONS.get(_p, '●')
        else:
            _circ = 'background:#fff;border-color:#dce1e8;color:#8b95a3'
            _mark = ''
        _lblc = '#131a21' if (_done or _cur) else '#8b95a3'
        _lblw = '700' if _cur else ('600' if _done else '500')
        _line = ('' if _i == 0 else
                 f'<div style="flex:1;height:2px;margin-top:11px;background:{"#0c6e51" if _i <= _idx else "#e7eaef"}"></div>')
        _seg += (_line +
                 '<div style="display:flex;flex-direction:column;align-items:center;flex:0 0 auto;min-width:62px">'
                 f'<div style="width:24px;height:24px;border-radius:50%;border:2px solid;display:flex;'
                 f'align-items:center;justify-content:center;font-size:11px;font-weight:800;{_circ}">{_mark}</div>'
                 f'<div style="margin-top:6px;font-size:.68rem;color:{_lblc};font-weight:{_lblw}">{_esc(_p)}</div>'
                 '</div>')
    return ('<div style="display:flex;align-items:flex-start;background:#f5f7f8;border:1px solid #e7eaef;'
            f'border-radius:14px;padding:16px 18px;margin:8px 0 4px;overflow-x:auto">{_seg}</div>')

def eventos_timeline_html(historial):
    """Timeline vertical de eventos (historial de cambios): tarjetas con icono +
    estado + fecha + usuario + nota. Estilo Finanzas historial-item."""
    if not historial:
        return '<div style="color:#8b95a3;font-size:.82rem;padding:6px 2px">Sin eventos registrados.</div>'
    _rows = ''
    for _h in reversed(historial):
        _est = _h.get('estado', '')
        _ic = ESTADO_ICONS.get(_est, '📜')
        _fe = (_h.get('fecha', '') or '')[:16].replace('T', ' ')
        _us = _h.get('usuario', '')
        _no = _h.get('nota', '')
        _no_html = f'<div style="color:#566472;font-size:.74rem;margin-top:2px">{_esc(_no)}</div>' if _no else ''
        _rows += (
            '<div style="display:flex;align-items:flex-start;gap:11px;background:#fff;border:1px solid #e7eaef;'
            'border-left:3px solid #0c6e51;border-radius:10px;padding:9px 12px;margin-bottom:7px">'
            f'<div style="font-size:1.05rem;opacity:.85;flex:0 0 auto">{_ic}</div>'
            '<div style="min-width:0">'
            f'<div style="font-weight:700;color:#131a21;font-size:.85rem">{_esc(_est)}</div>'
            f'<div style="color:#8b95a3;font-size:.72rem">{_esc(_fe)}{(" · " + _esc(_us)) if _us else ""}</div>'
            f'{_no_html}</div></div>')
    return _rows

# ─── BUSINESS LOGIC ──────────────────────────────────────────────────────
def segmentar(email, clients):
    peds = [p for p in load_pedidos() if p.get('client_email') == email]
    if not peds: return {'segmento':'Nuevo','descuento':0.0,'credito':10000,'badge':'🆕 Nuevo'}
    hoy = datetime.now()
    def _dias(p):
        try:
            return (hoy - datetime.fromisoformat(p.get('fecha') or hoy.isoformat())).days
        except (ValueError, TypeError):
            return 9999  # fecha inválida → no cuenta como reciente, pero no rompe
    p30 = [p for p in peds if _dias(p) <= 30]
    fac = sum(p.get('total_usd',0) for p in p30)
    if fac>=5000 or len(p30)>=10: return {'segmento':'VIP','descuento':0.05,'credito':50000,'badge':'⭐ VIP'}
    if len(peds)>=2: return {'segmento':'Regular','descuento':0.02,'credito':25000,'badge':'⚫ Regular'}
    return {'segmento':'Nuevo','descuento':0.0,'credito':10000,'badge':'🆕 Nuevo'}

def get_precio(codigo, destino, data):
    """Calcula precio CIF = precio_compra + flete_destino."""
    for p in data.get('products', []):
        if p.get('codigo') == codigo:
            # Support both old key names
            base = p.get('precio_cif_usd', 0) or p.get('precio_compra', 0)
            dest_val = data.get('config', {}).get('destinos', {}).get(destino, 0)
            if isinstance(dest_val, dict):
                # New format: {'moneda': 'USD', 'factor': 2.35} — factor = flete USD/caja
                flete = dest_val.get('factor', 0.0)
            elif isinstance(dest_val, (int, float)):
                # Original format: flete en USD/Kilo directamente
                flete = float(dest_val)
            else:
                flete = 0.0
            return round(base + flete, 2)
    return 0.0

def reg_cambio_precio(cod, antes, desp, motivo='Manual'):
    if antes==desp: return
    pct = ((desp-antes)/antes*100) if antes>0 else 0
    h2 = load_historial()
    h2.append({'id':f'CHG-{len(h2)+1:05d}','fecha':datetime.now().isoformat(),'producto':cod,'antes':antes,'despues':desp,'cambio_pct':round(pct,2),'usuario':st.session_state.get('user_email','sistema'),'motivo':motivo})
    save_historial(h2)
    if abs(pct)>20: st.warning(f'⚠️ Cambio >20% en {cod}: {pct:+.1f}%')

def exportar_excel(pedidos):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except: st.error('Instalar openpyxl'); return None
    wb = Workbook()
    ws1 = wb.active; ws1.title='Resumen'
    ws1['A1'] = 'EXPORT HARET - REPORTE'; ws1['A1'].font=Font(bold=True,color='FFFFFF'); ws1['A1'].fill=PatternFill(start_color='1B7A3C',end_color='1B7A3C',fill_type='solid')
    estados={}
    for p in pedidos:
        e=p.get('estado','Recibido'); estados[e]=estados.get(e,{'c':0,'t':0}); estados[e]['c']+=1; estados[e]['t']+=p.get('total_usd',0)
    ws1.append(['ESTADO','PEDIDOS','TOTAL USD'])
    for e,d in sorted(estados.items()): ws1.append([e,d['c'],round(d['t'],2)])
    ws2=wb.create_sheet('Pedidos')
    ws2.append(['ID','CLIENTE','EMAIL','ESTADO','DESTINO','TOTAL USD','FECHA'])
    for cell in ws2[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill(start_color='1B7A3C',end_color='1B7A3C',fill_type='solid')
    for p in sorted(pedidos,key=lambda x:x.get('fecha',''),reverse=True):
        ws2.append([p.get('id','').upper(),p.get('client_name',''),p.get('client_email',''),p.get('estado',''),p.get('destino',''),round(p.get('total_usd',0),2),p.get('fecha','')[:10]])
    ws3=wb.create_sheet('Productos')
    ws3.append(['CODIGO','PRODUCTO','CAJAS','PALLETS','PRECIO','TOTAL'])
    for cell in ws3[1]: cell.font=Font(bold=True,color='FFFFFF'); cell.fill=PatternFill(start_color='1B7A3C',end_color='1B7A3C',fill_type='solid')
    for p in pedidos:
        for pr in p.get('productos',[]): ws3.append([pr.get('codigo',''),pr.get('producto',''),pr.get('cajas',0),round(pr.get('pallets',0),2),round(pr.get('precio_usd',0),2),round(pr.get('cajas',0)*pr.get('precio_usd',0),2)])
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out.getvalue()

def calc_sla(pedidos):
    metas={'Recibido_Confirmado':4,'Confirmado_Preparando':2,'Preparando_Enviado':48,'Enviado_Entregado':168}
    slas=[]
    for p in pedidos:
        hist=sorted(p.get('historial_estados',[]),key=lambda x:x.get('fecha',''))
        for i in range(len(hist)-1):
            try:
                de=hist[i].get('estado',''); a=hist[i+1].get('estado','')
                t1=datetime.fromisoformat(hist[i]['fecha']); t2=datetime.fromisoformat(hist[i+1]['fecha'])
                h2=(t2-t1).total_seconds()/3600
                meta=metas.get(f'{de}_{a}')
                if meta: slas.append({'p':p.get('id',''),'tr':f'{de}→{a}','h':round(h2,1),'m':meta,'ok':h2<=meta})
            except Exception as e:
                logger.debug(f'SLA calc skip: {e}')
    tot=len(slas) or 1
    ok=sum(1 for s in slas if s['ok'])
    return slas,{'pct':round(ok/tot*100,1),'crit':tot-ok,'tot':len(slas),'prom':round(sum(s['h'] for s in slas)/tot,1)}

# ─── TAB DASHBOARD ─────────────────────────────────────────────
def render_dashboard():
    _admin_seccion('Resumen ejecutivo', '📊')
    pedidos=load_pedidos(); clients=load_clients(); data=load_data()
    c1,c2,c3,c4=st.columns(4)
    fac=sum(p.get('total_usd',0) for p in pedidos)
    vip=sum(1 for e in clients if segmentar(e,clients)['segmento']=='VIP')
    hoy_peds=len([p for p in pedidos if p.get('fecha','')[:10]==str(date.today())])
    c1.metric('📦 Pedidos',f'{len(pedidos):,}')
    c2.metric('💵 Facturación',f'${fac:,.0f}','USD')
    c3.metric('👥 Clientes',f'{len(clients):,}',f'{vip} VIP')
    c4.metric('📬 Hoy',hoy_peds,'nuevos')
    _admin_seccion('Pedidos por estado', '📋')
    ec={}
    for p in pedidos: ec[p.get('estado','Recibido')]=ec.get(p.get('estado','Recibido'),0)+1
    if ec:
        cols=st.columns(len(ORDEN_ESTADOS))
        for i,e in enumerate(ORDEN_ESTADOS): cols[i].metric(f"{ESTADO_ICONS.get(e,'')} {e}",ec.get(e,0))
    else: st.info('ℹ️ No hay pedidos. Crea uno en el tab **Hacer Pedido**.')
    _admin_seccion('Facturación mensual (USD)', '📈')
    if pedidos:
        _mes_data = {}
        for _p in pedidos:
            _fe = _p.get('fecha','')[:7]
            if _fe: _mes_data[_fe] = _mes_data.get(_fe, 0) + _p.get('total_usd', 0)
        if _mes_data:
            _meses_sorted = sorted(_mes_data.keys())[-12:]
            _df_chart = pd.DataFrame({'Mes': _meses_sorted, 'Total USD': [round(_mes_data[m], 2) for m in _meses_sorted]}).set_index('Mes')
            st.bar_chart(_df_chart, use_container_width=True, height=250)
    else: st.info('📊 Gráfico disponible cuando haya pedidos.')
    st.markdown('')
    with st.expander('⏱ SLA de Procesos', expanded=False):
        _,ss=calc_sla(pedidos)
        s1,s2,s3,s4=st.columns(4)
        s1.metric('Cumplimiento',f"{ss['pct']:.1f}%",'Meta:95%')
        s2.metric('⚠️ Críticos',ss['crit'])
        s3.metric('⏱ Prom.h',f"{ss['prom']:.1f}h")
        s4.metric('📊 Trans.',ss['tot'])
    # —— Alertas pedidos nuevos ——
    _nuevos = [p for p in pedidos if p.get('estado','') == 'Recibido']
    if _nuevos:
        _admin_seccion(f'Pedidos nuevos sin atender ({len(_nuevos)})', '🔔')
        _col_alerta = 'background:#fff3cd;border-left:4px solid #ffc107;border-radius:6px;padding:10px 14px;margin:4px 0'
        for _np in sorted(_nuevos, key=lambda x: x.get('fecha',''), reverse=True)[:5]:
            _np_id = _np.get('id','').upper()
            _np_cliente = _esc(_np.get('client_name',''))
            _np_total = _np.get('total_usd',0)
            _np_fecha = _np.get('fecha','')[:16].replace('T',' ')
            _np_dest = _np.get('destino','')
            st.markdown(f'<div style="{_col_alerta}">📦 <b>{_np_id}</b> — {_np_cliente} — <b>${_np_total:,.2f} USD</b> — {_np_dest} — <small style="color:#888">{_np_fecha}</small></div>', unsafe_allow_html=True)
        if len(_nuevos) > 5:
            st.caption(f'... y {len(_nuevos)-5} pedido(s) más en el tab 📦 Pedidos')
    _admin_seccion('Segmentación de clientes', '⭐')
    segs={'VIP':0,'Regular':0,'Nuevo':0}
    for e in clients: seg=segmentar(e,clients)['segmento']; segs[seg]=segs.get(seg,0)+1
    sg1,sg2,sg3=st.columns(3)
    sg1.metric('⭐ VIP',segs.get('VIP',0),'+5% desc.')
    sg2.metric('⚫ Regular',segs.get('Regular',0),'+2% desc.')
    sg3.metric('🆕 Nuevo',segs.get('Nuevo',0))
    if data.get('products',[]):
        st.markdown('---')
        st.markdown(f"### 📦 Productos ({len(data['products'])} activos)")
        _df_prods = pd.DataFrame([{'Código':p.get('codigo',''),'Producto':p.get('descripcion','') or p.get('producto',''),'Precio Compra':f"${p.get('precio_compra',0):.4f}",'Margen':f"{float(p.get('margen_pct',0.1) or 0.1)*100:.0f}%",'Activo':'✅' if p.get('activo',True) else '❌'} for p in data['products'][:15]])
        st.dataframe(_df_prods,use_container_width=True,hide_index=True)

# ─── TAB COTIZACION ───────────────────────────────────────────────────

# ── Mapa columnas Excel → código producto (igual que vigilar_excel.py) ──────
COL_MAP = {
     4: "F-PSG10",    5: "F-PN016",    6: "F-PPA01",
     7: "F-PSR02",    8: "F-PSR05",    9: "F-PSM09",
    10: "F-TAS04",   11: "F-GNB010",  12: "F-MPS03",
    13: "F-CCN017",  14: "F-BCC013",  15: "F-AHSS012",
    16: "F-BBB06",   17: "F-ZPT020",  18: "F-TX020",
    19: "F-UVP08",   20: "F-UVP07",
}

def parse_excel_file(xl_path):
    """
    Parsea Cotizaciones.xlsx.
    Lee hoja TABLA PRECIOS:
      - Fila 6 = cabeceras
      - Filas 7+ = productos: col2=codigo, col3=nombre, col4=kg/caja, col5=precio_compra,
        col12=flete/caja, col13=CIF base, col16-24=USD/caja a 1,2,3..9 pallets
      - Tambien lee filas 32-83 por columna (historial precios, legacy)
    Lee hoja CONFIGURACION para destinos y config.
    """
    from openpyxl import load_workbook
    import io as _io

    if isinstance(xl_path, str):
        with open(xl_path, 'rb') as f:
            wb_bytes = f.read()
    else:
        wb_bytes = xl_path.getvalue()

    wb = load_workbook(_io.BytesIO(wb_bytes), data_only=True)
    data = load_data()
    products_existing = {p['codigo']: p for p in data.get('products', [])}

    if 'TABLA PRECIOS' not in wb.sheetnames:
        raise ValueError("No se encontro la hoja 'TABLA PRECIOS'")
    ws_pr = wb['TABLA PRECIOS']

    # ── Leer tabla de productos fila a fila (filas 7+) ─────────────
    # Col: 1=blank, 2=Codigo, 3=Producto, 4=Kg/caja, 5=Precio_compra,
    #      6=Costo_caja, 7=FOBbase, 8=Merma, 9=FOB+Merma, 10=Margen%,
    #      11=Margen$, 12=PrecioFOBFinal, 13=Flete/caja, 14=CIF_USD,
    #      15=CIF_$/kg, 16=USD_1Pal, 17=USD_2Pal, ..., 24=USD_9Pal
    PLT_COL_START = 16  # columna openpyxl donde empieza USD 1 Pal
    PLT_COL_END = 38    # columna openpyxl donde termina USD 23 Pal (inclusive)
    N_PALLETS = PLT_COL_END - PLT_COL_START + 1  # = 23

    row_prices = {}  # codigo -> {precios_plt:[...], precio_compra:..., flete_caja:...}
    for r in range(7, 35):
        cod_val = ws_pr.cell(r, 2).value
        nom_val = ws_pr.cell(r, 3).value
        if not cod_val or not isinstance(cod_val, str) or not cod_val.strip():
            continue
        cod = str(cod_val).strip()
        nom = str(nom_val or '').strip()
        pc = ws_pr.cell(r, 5).value   # Precio compra
        fob_final = ws_pr.cell(r, 12).value  # FOB Final (col 12)
        fl = ws_pr.cell(r, 13).value  # Flete/caja
        cif_base = ws_pr.cell(r, 14).value  # CIF USD ref
        kg = ws_pr.cell(r, 4).value   # kg/caja
        precios_plt = []
        for c in range(PLT_COL_START, PLT_COL_END + 1):
            v = ws_pr.cell(r, c).value
            precios_plt.append(round(float(v), 4) if isinstance(v, (int, float)) and v > 0 else None)
        row_prices[cod] = {
            'nombre': nom,
            'precios_plt': precios_plt,
            'precio_compra': round(float(pc), 4) if isinstance(pc, (int, float)) else 0.0,
            'flete_ref': round(float(fl), 4) if isinstance(fl, (int, float)) else 0.0,
            'fob_final': round(float(fob_final), 4) if isinstance(fob_final, (int, float)) else 0.0,
            'cif_base': round(float(cif_base), 4) if isinstance(cif_base, (int, float)) else 0.0,
            'kg_caja': round(float(kg), 3) if isinstance(kg, (int, float)) else 0.0,
        }

    # ── Leer historial precios (filas 32-83, por columna) — legacy ──
    latest_prices = {}
    for col, cod in COL_MAP.items():
        last = None
        for r in range(32, 84):
            v = ws_pr.cell(row=r, column=col).value
            if isinstance(v, (int, float)) and v > 0:
                last = float(v)
        if last is not None:
            latest_prices[cod] = last

    # ── Construir lista de productos ────────────────────────────────
    products = []
    all_codes = set(list(row_prices.keys()) + list(COL_MAP.values()))
    for col, cod in sorted(COL_MAP.items()):
        existing = products_existing.get(cod, {})
        prod_data = dict(existing) if existing else {}
        rp = row_prices.get(cod, {})
        prod_data.update({
            'codigo': cod,
            'descripcion': rp.get('nombre') or existing.get('descripcion', '') or existing.get('producto', ''),
            'precio_compra': rp.get('precio_compra') or latest_prices.get(cod, 0.0),
            'precio_cif_usd': rp.get('cif_base') or (rp.get('precio_compra', 0) + rp.get('flete_ref', 0)),
        })
        if rp.get('precios_plt'):
            prod_data['precios_plt'] = rp['precios_plt']
        if rp.get('flete_ref'):
            prod_data['flete_ref'] = rp['flete_ref']
        if rp.get('fob_final'):
            prod_data['precio_fob_final'] = rp['fob_final']
        if rp.get('kg_caja'):
            prod_data['kg_caja'] = rp['kg_caja']
        if 'cajas_pallet' not in prod_data:
            prod_data['cajas_pallet'] = 200
        products.append(prod_data)

    # ── Leer destinos de CONFIGURACION ──────────────────────────────
    destinos_cfg = {}
    config_existing = data.get('config', {})
    if 'CONFIGURACION' in wb.sheetnames:
        ws_cfg = wb['CONFIGURACION']
        for row in ws_cfg.iter_rows():
            for cell in row:
                v = str(cell.value or '')
                dest = str(ws_cfg.cell(row=cell.row, column=2).value or '')
                c3 = ws_cfg.cell(row=cell.row, column=3).value
                if cell.column == 2 and dest and isinstance(c3, (int, float)) and c3 > 0:
                    if dest in config_existing.get('destinos', {}):
                        destinos_cfg[dest] = float(c3)
    if not destinos_cfg:
        destinos_cfg = config_existing.get('destinos', {})

    return products, destinos_cfg

def auto_load_excel():
    """Auto-carga Cotizaciones.xlsx al inicio si los productos no tienen precios válidos."""
    data = load_data()
    prods = data.get('products', [])
    # Re-parsear si no hay productos con precio_cif_usd válido
    has_prices = any(
        (p.get('precio_cif_usd', 0) or p.get('precio_compra', 0)) > 0
        for p in prods
    )
    if has_prices:
        return  # ya hay precios válidos
    xl_path = 'Cotizaciones.xlsx'
    if not os.path.exists(xl_path):
        return  # esperar upload manual
    try:
        products, destinos_cfg = parse_excel_file(xl_path)
        if products and any(p.get('precio_cif_usd', 0) > 0 for p in products):
            nueva = data.copy()
            nueva['products'] = products
            nueva['config'] = data.get('config', {})
            nueva['config']['destinos'] = destinos_cfg
            save_data(nueva)
    except Exception:
        pass  # ignorar errores de auto-carga

def render_catalogo():
    """Cat\u00e1logo: tabla estilo Excel con precios por pallets + destinos + importar."""
    data = load_data()
    prods = data.get('products', [])
    cfg = data.get('config', {})
    dests = cfg.get('destinos', {})
    dests_moneda = cfg.get('destinos_moneda', {})
    flete_ref = float(cfg.get('flete_ref', 2.35) or 2.35)  # flete referencia Madrid

    sub1, sub2, sub3, sub4 = st.tabs(['📊 Tabla de Precios', 'Destinos & Monedas', '📂 Importar Excel', '📦 Embalaje'])

    # ─── SUB-TAB 1: TABLA MAESTRA DE PRECIOS ─────────────────────────
    with sub1:
        _fc1, _fc2 = st.columns([3, 1])
        with _fc2:
            _show_inactive = st.toggle('Mostrar inactivos', value=False, key='cat_show_inactive',
                help='Muestra tambi\u00e9n productos desactivados')
        with _fc1:
            st.markdown('### \U0001f4cb Tabla de Precios por Pallets (USD/caja)')

        # P1-buscador: filtrar productos por nombre/c\u00f3digo/categor\u00eda
        _search_q = st.text_input('🔍 Buscar producto', placeholder='nombre, código o categoría...', key='cat_search_q')
        prods_vis = prods if _show_inactive else [p for p in prods if p.get('activo', True)]
        if _search_q:
            _sq = _search_q.lower().strip()
            prods_vis = [p for p in prods_vis if _sq in (p.get('producto','') + ' ' + p.get('codigo','') + ' ' + p.get('categoria','') + ' ' + p.get('grupo','')).lower()]
            st.caption(f'{len(prods_vis)} resultado(s) para: **{_search_q}**')
        dest_list = list(dests.keys())

        if not prods_vis:
            st.info('No hay productos activos. Importa el Excel o activa productos.')
            return

        # Selector de tipo de precio y destino
        _ct0, _dc1, _dc2 = st.columns([1.2, 2, 2])
        with _ct0:
            _cat_tipo = st.radio('\U0001f4b2 Tipo precio', ['CIF', 'FOB'], key='cat_tipo_precio', horizontal=True,
                help='FOB = precio en origen sin flete. CIF = precio incluye flete al destino.')
        with _dc1:
            _dest_sel = st.selectbox('\U0001f4cd Destino (ajuste flete)',
                dest_list if dest_list else ['Madrid/Espa\u00f1a'],
                key='cat_destino_sel',
                help='Selecciona el destino para ver los precios CIF con el flete correspondiente',
                disabled=(_cat_tipo == 'FOB')
            )
        with _dc2:
            _moneda_dest = dests_moneda.get(_dest_sel, 'USD')
            _rates_cat = get_exchange_rates()
            _rate_cat = _rates_cat.get(_moneda_dest, 1.0)
            _sym_cat = MONEDA_SIMBOLO.get(_moneda_dest, _moneda_dest)
            _dest_flete = dests.get(_dest_sel, 0)
            _dest_flete_v = float(_dest_flete.get('factor', _dest_flete) if isinstance(_dest_flete, dict) else _dest_flete if isinstance(_dest_flete, (int, float)) else 0)
            try:
                _margen_mkt = float(cfg.get('destinos_margen', {}).get(_dest_sel, 0) or 0)
            except (TypeError, ValueError):
                _margen_mkt = 0.0
            st.metric('\U0001f4b1 Moneda / Flete', f'{_moneda_dest} | ${_dest_flete_v:.2f}/Kilo')

        # ── Tabla estilo Excel: filas=pallets, columnas=productos ────
        if _cat_tipo == 'FOB':
            st.caption('Precios **FOB** (en origen, sin flete) | Todos los precios en USD/caja. A mayor volumen total del pedido, menor precio.')
        else:
            _marg_txt = f' | Margen mercado: **+{_margen_mkt:.1f}%**' if _margen_mkt else ''
            st.caption(f'Precios **CIF** hasta **{_dest_sel}** | Flete: **${_dest_flete_v:.2f} USD/Kilo**{_marg_txt} | Todos los precios en USD/caja. A mayor volumen total del pedido, menor precio.')

        # Precio de un producto a N pallets (CIF ajustado al destino + margen de
        # mercado del destino; o FOB fijo). El margen NO aplica a FOB (en origen).
        def _precio_en(_p, _n_plt):
            _precios_plt = _p.get('precios_plt', [])
            if _precios_plt:
                _idx = min(_n_plt - 1, len(_precios_plt) - 1)
                _precio_base = _precios_plt[_idx]
                if _cat_tipo == 'FOB':
                    _fob_f = float(_p.get('precio_fob_final', 0) or 0)
                    return round(_fob_f, 4) if _fob_f > 0 else None
                if _precio_base:
                    _pp = (float(_precio_base) - flete_ref + _dest_flete_v) * (1 + _margen_mkt / 100.0)
                    return round(_pp, 4)
                return None
            _pc = float(_p.get('precio_compra', 0) or 0)
            if _cat_tipo == 'FOB':
                _fob_f = float(_p.get('precio_fob_final', 0) or 0)
                return round(_fob_f, 4) if _fob_f > 0 else (round(_pc, 4) if _pc > 0 else None)
            _mg = float(_p.get('margen_pct', 0.1) or 0.1)
            return round((_pc * (1 + _mg) + _dest_flete_v) * (1 + _margen_mkt / 100.0), 4) if _pc > 0 else None

        # ── Vista FÁCIL: una FILA por fruta, columnas = tramos clave de volumen.
        # Por defecto solo los tramos donde el precio cambia de verdad (no 23 filas
        # casi iguales). Toggle para ver el detalle pallet a pallet.
        _ver_detalle = st.toggle('Ver detalle pallet a pallet', value=False, key='cat_ver_detalle',
                                 help='Por defecto se muestran los tramos clave de volumen (1, 2, 3, 5, 10, 20+)')
        _tiers = list(range(1, 24)) if _ver_detalle else [1, 2, 3, 5, 10, 20]
        _tier_lbl = {_t: (f'{_t} pal' if _t < 20 else '20+ pal') for _t in _tiers}

        _tbl_rows = []
        for _p in prods_vis:
            _name = (_p.get('producto','') or _p.get('descripcion','') or _p.get('codigo',''))
            _row = {'Fruta': _name}
            for _t in _tiers:
                _row[_tier_lbl[_t]] = _precio_en(_p, _t)
            _tbl_rows.append(_row)
        df_plt = pd.DataFrame(_tbl_rows).set_index('Fruta')

        # Color por FILA: más barato (más volumen) = verde más intenso.
        def _style_row(row):
            _vals = pd.to_numeric(row, errors='coerce').dropna()
            if len(_vals) < 2:
                return ['text-align:right'] * len(row)
            _vmin, _vmax = _vals.min(), _vals.max()
            _out = []
            for _v in row:
                _vn = pd.to_numeric(_v, errors='coerce')
                if pd.isna(_vn):
                    _out.append('text-align:right;color:#c6ccd6')
                else:
                    _pct = (_vmax - _vn) / (_vmax - _vmin) if _vmax > _vmin else 0  # 1 = el más barato
                    _out.append(f'background-color:rgba(12,110,81,{0.05 + _pct*0.22:.3f});text-align:right;font-weight:600')
            return _out

        _df_styled = df_plt.style.apply(_style_row, axis=1).format('${:.2f}', na_rep='—')
        st.dataframe(_df_styled, use_container_width=True, height=min(60 + 36*(len(prods_vis)+1), 760))

        st.caption('Verde = precio más bajo (más volumen). Cada fila es una fruta; las columnas son tramos de pallets. La app aplica automáticamente el precio del total de pallets del pedido.')

        # ── Consultor: precio a los pallets que tú quieras ─────────────
        st.markdown('---')
        _admin_seccion('Consultar precio a un volumen exacto', '🔢')
        _qc1, _qc2 = st.columns([1, 3])
        _n_custom = _qc1.number_input('Pallets a consultar', min_value=1, max_value=200, value=3, step=1,
                                      key='cat_n_custom', help='Escribe cuántos pallets del pedido quieres consultar')
        _custom_col = f'Precio a {int(_n_custom)} pal'
        _custom_rows = []
        for _p in prods_vis:
            _name = (_p.get('producto','') or _p.get('descripcion','') or _p.get('codigo',''))
            _custom_rows.append({'Fruta': _name, _custom_col: _precio_en(_p, int(_n_custom))})
        df_custom = pd.DataFrame(_custom_rows).set_index('Fruta')
        def _style_custom(col):
            _vals = pd.to_numeric(col, errors='coerce').dropna()
            if len(_vals) < 2:
                return ['text-align:right'] * len(col)
            _vmin, _vmax = _vals.min(), _vals.max()
            _out = []
            for _v in col:
                _vn = pd.to_numeric(_v, errors='coerce')
                if pd.isna(_vn):
                    _out.append('text-align:right;color:#c6ccd6')
                else:
                    _pct = (_vmax - _vn) / (_vmax - _vmin) if _vmax > _vmin else 0
                    _out.append(f'background-color:rgba(12,110,81,{0.05 + _pct*0.22:.3f});text-align:right;font-weight:700')
            return _out
        with _qc2:
            _mon_c = ('FOB (sin flete)' if _cat_tipo == 'FOB' else f'CIF hasta {_dest_sel}')
            st.caption(f'Precio USD/caja a **{int(_n_custom)} pallets** · {_mon_c}')
            st.dataframe(df_custom.style.apply(_style_custom, axis=0).format('${:.2f}', na_rep='—'),
                         use_container_width=True, height=min(60 + 36*(len(prods_vis)+1), 620))

        # ── Descarga cat\u00e1logo ─────────────────────────────────────────
        st.markdown('---')
        _dlc1, _dlc2 = st.columns(2)
        try:
            from openpyxl import Workbook as _WB
            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align
            _wb_dl = _WB()
            _ws_dl = _wb_dl.active
            _ws_dl.title = 'Precios'
            _hdr_dl = ['Fruta'] + list(df_plt.columns)
            _ws_dl.append(_hdr_dl)
            for _cell in _ws_dl[1]:
                _cell.font = _Font(bold=True, color='FFFFFF')
                _cell.fill = _Fill(start_color='1B7A3C', end_color='1B7A3C', fill_type='solid')
            for _row_d in _tbl_rows:
                _ws_dl.append([_row_d.get('Fruta','')] + [_row_d.get(c,'') for c in df_plt.columns])
            _out_dl = io.BytesIO(); _wb_dl.save(_out_dl); _out_dl.seek(0)
            _dlc1.download_button('\U0001f4e5 Descargar Excel', data=_out_dl.getvalue(),
                file_name=f'catalogo_{_dest_sel.replace("/","_")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True, key='dl_cat_excel')
        except ImportError:
            _dlc1.info('Instalar openpyxl para exportar')

        # ── Edici\u00f3n directa de precios por pallet ──────────────────
        st.markdown('---')
        st.markdown('### \u2699\ufe0f Editar Precios por Pallet')
        st.caption('Edita los precios USD/caja directamente. Cada fila = 1 pallet total del pedido. Cambios se guardan en el cat\u00e1logo.')

        # Construir DataFrame editable: filas=productos, columnas=pallets 1-9
        if 'price_edit_v' not in st.session_state: st.session_state.price_edit_v = 0
        _edit_rows = []
        for _p in prods:
            _precios_plt = _p.get('precios_plt', [None]*23)
            while len(_precios_plt) < 23: _precios_plt.append(None)
            _edit_rows.append({
                'Cod': _p.get('codigo',''),
                'Producto': _p.get('producto','') or _p.get('descripcion',''),
                'Activo': bool(_p.get('activo', True)),
                'Min Unidad': str(_p.get('min_unidad', 'Pallets') or 'Pallets'),
                'Min Cant': int(_p.get('min_cantidad', 0) or 0),
                '1 Plt': _precios_plt[0],
                '2 Plt': _precios_plt[1],
                '3 Plt': _precios_plt[2],
                '4 Plt': _precios_plt[3],
                '5 Plt': _precios_plt[4],
                '6 Plt': _precios_plt[5],
                '7 Plt': _precios_plt[6],
                '8 Plt': _precios_plt[7],
                '9 Plt': _precios_plt[8],
                '10 Plt': _precios_plt[9],
                '11 Plt': _precios_plt[10],
                '12 Plt': _precios_plt[11],
                '13 Plt': _precios_plt[12],
                '14 Plt': _precios_plt[13],
                '15 Plt': _precios_plt[14],
                '16 Plt': _precios_plt[15],
                '17 Plt': _precios_plt[16],
                '18 Plt': _precios_plt[17],
                '19 Plt': _precios_plt[18],
                '20 Plt': _precios_plt[19],
                '21 Plt': _precios_plt[20],
                '22 Plt': _precios_plt[21],
                '23 Plt': _precios_plt[22]
            })
        _df_edit = pd.DataFrame(_edit_rows)
        _plt_cols = ['1 Plt','2 Plt','3 Plt','4 Plt','5 Plt','6 Plt','7 Plt','8 Plt','9 Plt','10 Plt','11 Plt','12 Plt','13 Plt','14 Plt','15 Plt','16 Plt','17 Plt','18 Plt','19 Plt','20 Plt','21 Plt','22 Plt','23 Plt']
        _col_cfg_edit = {
            'Cod': st.column_config.TextColumn('C\u00f3digo', disabled=True, width='small'),
            'Producto': st.column_config.TextColumn('Producto', width='medium'),
            'Activo': st.column_config.CheckboxColumn('Activo', width='small',
                help='Desactiva para ocultar del portal del cliente'),
            'Min Unidad': st.column_config.SelectboxColumn('Min Unidad', options=['Pallets', 'Cajas'], width='small',
                help='Unidad del pedido mínimo: Pallets o Cajas'),
            'Min Cant': st.column_config.NumberColumn('Mínimo', min_value=0, step=1, width='small',
                help='Cantidad mínima de pedido por producto (0 = sin mínimo)'),
        }
        for _pc_name in _plt_cols:
            _col_cfg_edit[_pc_name] = st.column_config.NumberColumn(
                _pc_name, format='$%.2f', step=0.01, min_value=0,
                help=f'Precio USD/caja al ordenar {_pc_name.replace(" Plt","")} pallets totales (incluye flete ref. Madrid)'
            )
        _edited_prods = st.data_editor(
            _df_edit,
            column_config=_col_cfg_edit,
            use_container_width=True,
            num_rows='dynamic',
            key=f'edit_precios_plt_{st.session_state.price_edit_v}',
            hide_index=True,
        )
        if st.button('\U0001f4be Guardar Precios', type='primary', use_container_width=True, key='btn_save_precios_plt'):
            _old_by_cod = {p.get('codigo'): p for p in prods}
            _new_prods = []
            for _, _r in _edited_prods.iterrows():
                _cod_r = str(_r.get('Cod',''))
                _base = dict(_old_by_cod.get(_cod_r, {})) if _cod_r in _old_by_cod else {}
                _base.update({
                    'codigo': _cod_r,
                    'producto': str(_r.get('Producto','')),
                    'descripcion': str(_r.get('Producto','')),
                    'activo': bool(_r.get('Activo', True)),
                    'min_unidad': (lambda _mu: _mu if _mu in ('Pallets','Cajas') else 'Pallets')(str(_r.get('Min Unidad','Pallets')) if _r.get('Min Unidad') is not None and str(_r.get('Min Unidad')) not in ('nan','None','') else 'Pallets'),
                    'min_cantidad': (lambda _mc: int(_mc) if _mc is not None and str(_mc) not in ('nan','None','') and not (isinstance(_mc, float) and _mc != _mc) else 0)(_r.get('Min Cant')),
                    'precios_plt': [
                        (float(_r[c]) if _r[c] is not None and str(_r[c]) not in ('','nan','None') else None)
                        for c in _plt_cols
                    ],
                })
                if not _base.get('margen_pct'): _base['margen_pct'] = 0.1
                if not _base.get('grupo'): _base['grupo'] = 'A'
                if _cod_r: _new_prods.append(_base)
            data['products'] = _new_prods
            # ─── Registrar cambios de mínimos ──────────────────────────────────
            _min_log = load_min_log()
            for _np in _new_prods:
                _cod_np = _np.get("codigo", "")
                _old_p = _old_by_cod.get(_cod_np, {})
                _old_mu = _old_p.get("min_unidad", "Pallets")
                _old_mc = int(_old_p.get("min_cantidad", 0) or 0)
                _new_mu = _np.get("min_unidad", "Pallets")
                _new_mc = int(_np.get("min_cantidad", 0) or 0)
                if _old_mu != _new_mu or _old_mc != _new_mc:
                    _min_log.append({
                        "fecha_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "usuario": st.session_state.get("user_email", "admin"),
                        "codigo": _cod_np,
                        "producto": _np.get("producto", _cod_np),
                        "min_unidad_antes": _old_mu,
                        "min_cantidad_antes": _old_mc,
                        "min_unidad_despues": _new_mu,
                        "min_cantidad_despues": _new_mc,
                    })
            save_min_log(_min_log)
            # ────────────────────────────────────────────────────
            save_data(data)
            st.toast('Precios guardados \u2705', icon='\u2705')
            st.session_state.price_edit_v = st.session_state.get('price_edit_v', 0) + 1
            st.rerun()

    # ─── SUB-TAB 2: DESTINOS & MONEDAS ───────────────────────────────
    with sub2:
        st.markdown('### \U0001f30d Gestionar Destinos, Fletes y M\u00e1rgenes')
        st.caption('Flete en USD/Kilo. **Margen %** = sobreprecio de mercado por destino (p. ej. Espa\u00f1a 0 %, UK 8 %). '
                   'A\u00f1ade, edita o elimina destinos directamente en la tabla.')
        _dest_margen = cfg.get('destinos_margen', {})
        _dest_rows = []
        for _dn, _dv in dests.items():
            _fv = _dv.get('factor', _dv) if isinstance(_dv, dict) else _dv
            _mv = dests_moneda.get(_dn, 'USD')
            _gv = float(_dest_margen.get(_dn, 0) or 0)
            _dest_rows.append({'Destino': _dn, 'Flete $/cj': float(_fv), 'Margen %': _gv, 'Moneda': _mv})
        _df_dest = pd.DataFrame(_dest_rows)
        _edited_dest = st.data_editor(
            _df_dest,
            column_config={
                'Destino': st.column_config.TextColumn('Destino', width='medium',
                    help='Nombre del destino. A\u00f1ade filas para nuevos destinos.'),
                'Flete $/cj': st.column_config.NumberColumn('Flete $/cj', format='$%.4f', step=0.01,
                    help='Costo de flete USD por caja incluido en precio CIF'),
                'Margen %': st.column_config.NumberColumn('Margen %', format='%.1f %%', step=0.5,
                    help='Sobreprecio de mercado para este destino (0 = sin extra). Ej: UK 8, Espa\u00f1a 0'),
                'Moneda': st.column_config.SelectboxColumn('Moneda', options=MONEDAS,
                    help='Divisa de referencia para el cliente'),
            },
            use_container_width=True, num_rows='dynamic',
            key='edit_destinos_tabla', hide_index=True,
        )
        if st.button('\U0001f4be Guardar Destinos', type='primary', use_container_width=True, key='btn_guardar_destinos'):
            _new_d = {}; _new_m = {}; _new_g = {}
            for _, _r in _edited_dest.iterrows():
                if _r.get('Destino'):
                    _dn2 = str(_r['Destino'])
                    _new_d[_dn2] = float(_r['Flete $/cj'])
                    _new_m[_dn2] = str(_r['Moneda'])
                    try:
                        _new_g[_dn2] = float(_r.get('Margen %', 0) or 0)
                    except (TypeError, ValueError):
                        _new_g[_dn2] = 0.0
            data['config']['destinos'] = _new_d
            data['config']['destinos_moneda'] = _new_m
            data['config']['destinos_margen'] = _new_g
            save_data(data)
            st.toast('Destinos guardados \u2705', icon='\u2705')
            st.rerun()
        st.markdown('---')
        st.markdown('### \U0001f4b9 Tipos de Cambio en Tiempo Real')
        _rates2 = get_exchange_rates()
        _r2_cols = st.columns(min(len(_rates2), 9))
        for _i2, (_cur2, _rate2) in enumerate(list(_rates2.items())[:9]):
            with _r2_cols[_i2]:
                _sym2 = MONEDA_SIMBOLO.get(_cur2, _cur2)
                st.metric(f'USD/{_cur2}', f'{_sym2}{_rate2:.4f}')

    # ─── SUB-TAB 3: IMPORTAR EXCEL ────────────────────────────────────
    with sub3:
        st.markdown('### \U0001f4c2 Importar desde Excel')
        st.info('\U0001f4a1 El archivo Excel debe tener la hoja **TABLA PRECIOS** con la estructura est\u00e1ndar de Export Haret (fila 6 = cabeceras, filas 7+ = productos, columnas 16-24 = precios 1-9 pallets).')
        st.success('Los precios del Excel son la **BASE**. Los **m\u00e1rgenes de mercado por destino** '
                   '(pesta\u00f1a **Destinos & Monedas**) se gestionan en la app y **NO se sobrescriben** al reimportar.')
        _uploaded = st.file_uploader('Subir Cotizaciones.xlsx', type=['xlsx','xls'], key='excel_uploader_cat')
        if _uploaded:
            try:
                _parsed_prods, _parsed_dests = parse_excel_file(_uploaded)
                if _parsed_prods:
                    _prods_with_plt = [p for p in _parsed_prods if p.get('precios_plt')]
                    st.success(f'\u2705 {len(_parsed_prods)} productos encontrados, {len(_prods_with_plt)} con tabla de precios por pallet.')
                    _prev_map = {p.get('codigo'): p for p in prods}
                    _new_c = 0; _upd_c = 0
                    for _np in _parsed_prods:
                        _nc2 = _np.get('codigo','')
                        if _nc2 in _prev_map:
                            _prev_map[_nc2].update({k: v for k, v in _np.items() if v is not None})
                            _upd_c += 1
                        else:
                            prods.append(_np); _new_c += 1
                    data['products'] = list(_prev_map.values()) + [p for p in prods if p.get('codigo') not in _prev_map]
                    # IMPORTANTE: solo se actualizan productos y flete por destino. Los
                    # MÁRGENES de mercado (config.destinos_margen) y demás config NO se
                    # tocan → los márgenes que pusiste en la app sobreviven al reimport.
                    if _parsed_dests: data['config']['destinos'] = _parsed_dests
                    save_data(data)
                    st.success(f'Importado: {_upd_c} actualizados, {_new_c} nuevos.')
                    st.rerun()
                else:
                    st.warning('No se encontraron productos v\u00e1lidos.')
            except Exception as _xe:
                st.error(f'Error al importar: {_xe}')
                import traceback as _xtb; st.code(_xtb.format_exc())
    # ─── SUB-TAB 4: TABLA DE EMBALAJE ─────────────────────────────────────────────
    with sub4:
        st.markdown('### 📦 Tabla de Embalaje por Grupos')
        st.caption('Define los grupos de embalaje. Cada grupo determina las cajas por pallet y kg/caja de los productos asignados.')

        _emb_data = load_data()
        _emb_cfg = _emb_data.get('config', {})
        _emb_grupos = _emb_cfg.get('grupos', {})
        _emb_prods = _emb_data.get('products', [])

        # Lista de todos los productos disponibles (cod -> nombre)
        _all_prod_opts = {
            p.get('codigo',''): (p.get('producto','') or p.get('descripcion','') or p.get('codigo',''))
            for p in _emb_prods if p.get('codigo','')
        }

        # ── Tabla editable de grupos ──────────────────────────────────────────
        st.markdown('#### Grupos de Embalaje')
        _emb_rows = []
        for _grp_key in sorted(_emb_grupos.keys()):
            _gi = _emb_grupos[_grp_key]
            _emb_rows.append({
                'Grupo': _grp_key,
                'Nombre': _gi.get('nombre', ''),
                'Cj/Plt': int(_gi.get('cajas_pallet', 160) or 160),
                'Kg/Caja': float(_gi.get('kg_caja', 0) or 0),
            })

        _emb_df = pd.DataFrame(_emb_rows) if _emb_rows else pd.DataFrame(
            columns=['Grupo', 'Nombre', 'Cj/Plt', 'Kg/Caja'])

        _emb_col_cfg = {
            'Grupo': st.column_config.TextColumn('Grupo', width='small',
                help='Letra identificadora del grupo (A, B, C...)'),
            'Nombre': st.column_config.TextColumn('Nombre / Descripcion', width='large',
                help='Nombre descriptivo del grupo de embalaje'),
            'Cj/Plt': st.column_config.NumberColumn('Cj/Plt', min_value=1, step=1,
                help='Cajas por pallet para este grupo'),
            'Kg/Caja': st.column_config.NumberColumn('Kg/Caja', format='%.2f', step=0.05, min_value=0,
                help='Peso neto de producto por caja (kg)'),
        }

        _edited_emb = st.data_editor(
            _emb_df,
            column_config=_emb_col_cfg,
            use_container_width=True,
            num_rows='dynamic',
            key='edit_embalaje_grupos',
            hide_index=True,
            height=min(80 + 40 * (len(_emb_rows) + 2), 500),
        )

        if st.button('\U0001f4be Guardar Grupos', type='primary', use_container_width=True, key='btn_save_embalaje'):
            _new_grupos = {}
            for _, _er in _edited_emb.iterrows():
                _grp_k = str(_er.get('Grupo', '')).strip().upper()
                if not _grp_k: continue
                _existing = _emb_grupos.get(_grp_k, {})
                _upd = {}
                for _ek, _ev in _existing.items():
                    if _ek not in ('nombre','cajas_pallet','kg_caja'):
                        _upd[_ek] = _ev
                _upd['nombre'] = str(_er.get('Nombre', '') or '')
                _upd['cajas_pallet'] = int(_er.get('Cj/Plt', 160) or 160)
                _upd['kg_caja'] = float(_er.get('Kg/Caja', 0) or 0)
                _new_grupos[_grp_k] = _upd
            _emb_data['config']['grupos'] = _new_grupos
            save_data(_emb_data)
            st.toast('Grupos guardados', icon='✅')
            st.rerun()

        st.markdown('---')

        # ── Asignacion de productos a grupos (multiselect por grupo) ─────────
        st.markdown('#### Productos por Grupo')
        st.caption('Asigna productos a cada grupo. Puedes seleccionar varios productos por grupo. Al guardar se actualizan las cajas/pallet y kg/caja de cada producto.')

        # Current assignment
        _grp_to_cods = {}
        for _ppx in _emb_prods:
            _ppg = _ppx.get('grupo', '')
            _ppc = _ppx.get('codigo', '')
            if _ppg and _ppc:
                if _ppg not in _grp_to_cods:
                    _grp_to_cods[_ppg] = []
                _grp_to_cods[_ppg].append(_ppc)

        # Label maps
        _cod_label_map = {cod: cod + ' – ' + nom for cod, nom in _all_prod_opts.items()}
        _label_cod_map = {v: k for k, v in _cod_label_map.items()}
        _prod_names_for_select = list(_cod_label_map.values())

        _new_assignments = {}
        for _grp_k in sorted(_emb_grupos.keys()):
            _gi2 = _emb_grupos.get(_grp_k, {})
            _grp_nom2 = _gi2.get('nombre', '') or _grp_k
            _grp_cxp2 = int(_gi2.get('cajas_pallet', 160) or 160)
            _current_cods = _grp_to_cods.get(_grp_k, [])
            _current_labels = [_cod_label_map[c] for c in _current_cods if c in _cod_label_map]

            _gc1, _gc2 = st.columns([1, 3])
            with _gc1:
                st.markdown(f'**Grupo {_grp_k}** — `{_grp_cxp2}` cj/plt')
                if _grp_nom2:
                    st.caption(_grp_nom2[:60])
            with _gc2:
                _sel = st.multiselect(
                    label=f'Grupo {_grp_k}',
                    options=_prod_names_for_select,
                    default=[lbl for lbl in _current_labels if lbl in _prod_names_for_select],
                    key=f'msel_grp_{_grp_k}',
                    label_visibility='collapsed',
                    placeholder='Seleccionar productos...',
                )
                _new_assignments[_grp_k] = [_label_cod_map[lbl] for lbl in _sel if lbl in _label_cod_map]

        if st.button('\U0001f4be Guardar Asignacion de Productos', type='primary', use_container_width=True, key='btn_save_prod_grp'):
            _new_cod_grp = {}
            for _grp_k2, _cods2 in _new_assignments.items():
                for _c2 in _cods2:
                    _new_cod_grp[_c2] = _grp_k2
            _new_prods2 = []
            for _pp2 in _emb_prods:
                _c3 = _pp2.get('codigo', '')
                _pp2c = dict(_pp2)
                if _c3 in _new_cod_grp:
                    _grp3 = _new_cod_grp[_c3]
                    _pp2c['grupo'] = _grp3
                    _gi4 = _emb_grupos.get(_grp3, {})
                    if isinstance(_gi4, dict) and _gi4.get('cajas_pallet'):
                        _pp2c['cajas_pallet'] = int(_gi4['cajas_pallet'])
                    if isinstance(_gi4, dict) and _gi4.get('kg_caja'):
                        _pp2c['kg_caja'] = float(_gi4['kg_caja'])
                _new_prods2.append(_pp2c)
            _emb_data['products'] = _new_prods2
            save_data(_emb_data)
            st.toast('Asignacion de productos guardada', icon='✅')
            st.rerun()


def render_hacer_pedido():
    st.markdown('## 🛒 Crear Nuevo Pedido')
    data=load_data(); prods=data.get('products',[]); dests=data.get('config',{}).get('destinos',{})
    if not prods:
        st.warning('⚠️ No hay productos. Ve al tab **Cotización** y sube tu archivo Excel primero.')
        return
    clients=load_clients()
    # Paso 1: Cliente
    st.markdown('### 1️⃣ Datos del Cliente')
    cl1,cl2=st.columns(2)
    c_email=cl1.text_input('Email',key='hp_email')
    c_name=cl2.text_input('Nombre',placeholder='Nombre / Empresa',key='hp_nombre')
    seg=None
    if c_email and c_email in clients:
        c=clients[c_email]; seg=segmentar(c_email,clients)
        st.success(f'Cliente: {c.get("nombre","")} | {seg["badge"]} | Desc: {seg["descuento"]*100:.0f}%')
        if not c_name: c_name=c.get('nombre','')
    elif c_email: st.info('🆕 Cliente nuevo - se registrará al guardar')
    # Paso 2: Destino
    st.markdown('### 2️⃣ Destino')
    dest_opts=list(dests.keys()) if dests else ['Madrid/España','París/Francia','Londres/UK','Miami/USA']
    destino=st.selectbox('Destino',dest_opts,key='hp_dest')
    dest_v=dests.get(destino,{}) if dests else {}
    moneda=dest_v.get('moneda','USD') if isinstance(dest_v,dict) else 'USD'
    dest_factor=dest_v.get('factor',1.0) if isinstance(dest_v,dict) else (float(dest_v) if isinstance(dest_v,(int,float)) else 1.0)
    st.caption(f'Moneda destino: {moneda}')
    # Paso 3: Productos
    st.markdown('### 3️⃣ Agregar Productos')
    pc1,pc2,pc3=st.columns([3,1,1])
    prod_opts=[p.get('codigo','')+' - '+(p.get('descripcion','') or p.get('producto','')) for p in prods]
    psel=pc1.selectbox('Producto',['Seleccionar...']+prod_opts,key='hp_prod')
    cajas=pc2.number_input('Cajas',min_value=1,value=100,step=50,key='hp_cajas')
    pc3.markdown('<br>',unsafe_allow_html=True)
    if pc3.button('➕ Agregar') and psel!='Seleccionar...':
        cod=psel.split(' - ')[0]
        pd_=next((p for p in prods if p.get('codigo')==cod),{})
        _cart_pals_admin = sum(i.get('pallets',0) for i in st.session_state.carrito)
        _cxp_tmp = pd_.get('cajas_pallet', 200) or 200
        _new_pallets_tmp = round(cajas / _cxp_tmp, 2)
        _total_pals_admin = max(_cart_pals_admin + _new_pallets_tmp, 1)
        precio = get_precio_con_volumen(cod, destino, 'CIF' if destino else 'FOB', data, _total_pals_admin)
        cxp=pd_.get('cajas_pallet',200) or 200
        pallets=round(cajas/cxp,2)
        item={'codigo':cod,'producto':pd_.get('descripcion','') or pd_.get('producto',''),'cajas':cajas,'pallets':pallets,'precio_usd':precio,'total':round(cajas*precio,2)}
        ex_idx=next((i for i,x in enumerate(st.session_state.carrito) if x['codigo']==cod),None)
        if ex_idx is not None:
            st.session_state.carrito[ex_idx]['cajas']+=cajas
            st.session_state.carrito[ex_idx]['total']=round(st.session_state.carrito[ex_idx]['cajas']*precio,2)
        else: st.session_state.carrito.append(item)
        st.rerun()
    # Carrito
    if st.session_state.carrito:
        st.markdown('#### 🛒 Carrito')
        st.dataframe(pd.DataFrame(st.session_state.carrito)[['codigo','producto','cajas','pallets','precio_usd','total']],use_container_width=True,hide_index=True)
        tot=sum(i['total'] for i in st.session_state.carrito)
        tc1,tc2,tc3=st.columns(3)
        tc1.metric('📦 Cajas',f'{sum(i["cajas"] for i in st.session_state.carrito):,}')
        tc2.metric('📍 Pallets',f'{sum(i["pallets"] for i in st.session_state.carrito):.1f}')
        tc3.metric('💰 Total',f'${tot:,.2f}')
        if st.button('🗑️ Vaciar',key='vaciar'): st.session_state.carrito=[]; st.rerun()
    # Paso 4
    st.markdown('### 4️⃣ Confirmar')
    notas=st.text_area('Notas',placeholder='Instrucciones especiales...',key='hp_notas')
    ht1,ht2=st.columns(2)
    TOPH=['','Pago anticipado 100%','50% adelanto / 50% contra documentos','30% adelanto / 70% contra BL','Carta de cr\xe9dito (LC)','Pago a 30 d\xedas','Pago a 60 d\xedas','Otro']
    hp_term=ht1.selectbox('📋 Términos de pago',TOPH,key='hp_term')
    hp_ent=ht2.text_input('📅 Fecha entrega estimada',placeholder='ej: 2026-06-20',key='hp_ent')
    notas_internas=st.text_area('🔒 Notas internas (no visibles al cliente)',placeholder='Instrucciones de almacén, condiciones especiales...',key='hp_notas_int',height=60)
    if st.session_state.carrito:
        _tot_pre = sum(i['total'] for i in st.session_state.carrito)
        _plt_pre = sum(i.get('pallets',0) for i in st.session_state.carrito)
        st.markdown(
            f'<div style="background:#f0f7ff;border:1px solid #0c6e51;border-radius:8px;padding:12px 18px;margin:8px 0">'
            f'📋 <b>Resumen</b><br>'
            f'&bull; Cliente: <b>{c_name}</b> ({c_email})<br>'
            f'&bull; Destino: <b>{destino}</b><br>'
            f'&bull; Productos: <b>{len(st.session_state.carrito)}</b> &nbsp; Pallets: <b>{_plt_pre:.2f}</b><br>'
            f'&bull; 💰 Total: <b style="color:#0c6e51">${_tot_pre:,.2f} USD</b>'
            f'</div>',
            unsafe_allow_html=True
        )
    if st.button('📤 GUARDAR PEDIDO',type='primary',use_container_width=True):
        if not c_email: st.error('❌ Ingresa email del cliente')
        elif not c_name: st.error('❌ Ingresa nombre del cliente')
        elif not st.session_state.carrito: st.error('❌ Agrega productos al carrito')
        else:
            c_email=(c_email or '').strip().lower()  # clave única normalizada (sin duplicados)
            _tod_h=load_pedidos()
            _yn_h=datetime.now().strftime('%Y')
            _pc_h=[p for p in _tod_h if p.get('id','').startswith(f'PED-{_yn_h}')]
            pid=f'PED-{_yn_h}-{len(_pc_h)+1:04d}'
            tot=sum(i['total'] for i in st.session_state.carrito)
            ped={'id':pid,'client_email':c_email,'client_name':c_name,'destino':destino,'moneda':moneda,'productos':list(st.session_state.carrito),'total_usd':round(tot,2),'estado':'Recibido','fecha':datetime.now().isoformat(),'notas':notas,'notas_internas':notas_internas,'terminos_pago':hp_term,'fecha_entrega':hp_ent,'historial_estados':[{'estado':'Recibido','fecha':datetime.now().isoformat(),'usuario':st.session_state.user_email}],'creado_por':st.session_state.user_email}
            todos=load_pedidos(); todos.append(ped); save_pedidos(todos)
            sync_finanzas(ped, todos)
            if c_email not in clients: clients[c_email]={'nombre':c_name,'email':c_email,'fecha_registro':datetime.now().isoformat(),'pedidos_ids':[]}
            clients[c_email]['pedidos_ids']=clients[c_email].get('pedidos_ids',[])+[pid]
            save_clients(clients)
            try:
                send_order_email(ped)
                _email_status = 'enviado'
            except Exception:
                _email_status = 'fallido'
            el=load_email_log(); el.append({'id':f'EMAIL-{len(el)+1:05d}','destinatario':c_email,'asunto':f'Pedido {pid} recibido','tipo':'confirmacion','fecha':datetime.now().isoformat(),'estado':_email_status}); save_email_log(el)
            st.session_state.carrito=[]
            st.success(f'Pedido {pid} creado por ${tot:,.2f}')
            st.cache_data.clear()

# ─── TAB PRECIOS ─────────────────────────────────────────────────────────
def render_destinos():
    st.markdown('## Todos los Destinos - Tarifas')
    data=load_data(); dests=data.get('config',{}).get('destinos',{})
    if not dests: st.info('⚠️ Sube el Excel en Cotización para ver los destinos.'); return
    rows_d=[]
    for k,v in dests.items():
        if isinstance(v,dict): rows_d.append({'Destino':k,'Moneda':v.get('moneda','USD'),'CIF':v.get('factor',1.0)})
        else: rows_d.append({'Destino':k,'Moneda':'USD/EUR','CIF USD':round(float(v),2) if isinstance(v,(int,float)) else 0})
    st.dataframe(pd.DataFrame(rows_d),use_container_width=True,hide_index=True)

# ─── TAB GESTION PEDIDOS ──────────────────────────────────────────────
def render_gestion_pedidos():
    st.markdown('## 📦 Gestión de Pedidos')
    # Aviso de estado al cliente: resultado del email + botón WhatsApp de un clic.
    _sn = st.session_state.get('_status_notified')
    if _sn:
        _ic = ESTADO_ICONS.get(_sn.get('estado',''), '📦')
        _cli = _sn.get('cliente') or _sn.get('email') or 'el cliente'
        if _sn.get('email_ok'):
            st.success(f"{_ic} Pedido {_sn.get('pid','')} → **{_sn.get('estado','')}**. "
                       f"Avisamos a **{_cli}** por email automáticamente.")
        else:
            st.info(f"{_ic} Pedido {_sn.get('pid','')} → **{_sn.get('estado','')}**. "
                    f"El email automático no está configurado — avisa a **{_cli}** por WhatsApp 👇")
        _wc1, _wc2 = st.columns([1, 3])
        _wc1.link_button('📲 Avisar por WhatsApp', _sn.get('wa', 'https://wa.me/'), use_container_width=True)
        if _wc2.button('✓ Listo / ocultar', key='status_notify_dismiss'):
            st.session_state.pop('_status_notified', None)
            st.rerun()
    pedidos=load_pedidos()
    # KPI filter: si vienen de sidebar 'En proceso', precargar estados pendientes
    _kpi_filter = st.session_state.pop('pedidos_filter_estado', None)
    if _kpi_filter:
        st.info(f"\u26a1 Mostrando pedidos filtrados desde el KPI sidebar: **{', '.join(_kpi_filter)}**")
    f1,f2,f3=st.columns(3)
    if _kpi_filter:
        fe = f1.multiselect('Estado',ORDEN_ESTADOS,default=_kpi_filter,key='gp_e_multi')
        _use_multi = True
    else:
        fe = f1.selectbox('Estado',['Todos']+ORDEN_ESTADOS,key='gp_e')
        _use_multi = False
    if '_gp_c_pending' in st.session_state:   # viene de la ficha de cliente ("Ver sus pedidos")
        st.session_state['gp_c'] = st.session_state.pop('_gp_c_pending')
    fc=f2.text_input('🔍 Buscar (cliente, email o ID)',key='gp_c',placeholder='nombre, correo o PED-…')
    fd=f3.selectbox('Destino',['Todos']+sorted(set(p.get('destino','') for p in pedidos if p.get('destino'))),key='gp_d')
    fd1,fd2=st.columns(2)
    _fecha_desde = fd1.date_input('📅 Desde', value=date.today() - timedelta(days=90), key='gp_desde')
    _fecha_hasta = fd2.date_input('📅 Hasta', value=date.today(), key='gp_hasta')
    filt=[p for p in pedidos if
        ((_use_multi and p.get('estado') in fe) or (not _use_multi and (fe=='Todos' or p.get('estado')==fe))) and
        (not fc or fc.lower() in (p.get('client_name','')+p.get('id','')+p.get('client_email','')).lower()) and
        (fd=='Todos' or p.get('destino')==fd) and
        (str(_fecha_desde) <= p.get('fecha','')[:10] <= str(_fecha_hasta))
    ]
    _tot_filt = sum(p.get('total_usd',0) for p in filt)
    st.info(f'📦 **{len(filt)} pedidos** filtrados | 💰 Total: $**{_tot_filt:,.2f}** USD')
    if filt:
        xb=exportar_excel(filt)
        if xb: st.download_button('📥 Excel',data=xb,file_name=f'pedidos_{date.today()}.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    st.markdown('---')
    for ped in sorted(filt,key=lambda x:x.get('fecha',''),reverse=True):
        icon=ESTADO_ICONS.get(ped.get('estado',''),'📦')
        with st.expander(f"{icon} #{ped.get('id','').upper()} • {ped.get('client_name','N/A')} • {ped.get('destino','')} • ${ped.get('total_usd',0):,.2f}"):
            _est_p = ped.get('estado', 'Recibido')
            # Cabecera detallada (estilo Finanzas): badge + nº + total + ciclo de vida
            st.markdown(
                '<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin:2px 0 4px">'
                f'{estado_badge(_est_p)}'
                f'<span style="font-weight:800;color:#131a21;font-size:1.02rem">#{_esc(ped.get("id","").upper())}</span>'
                f'<span style="margin-left:auto;font-weight:800;color:#084a37;font-size:1.05rem;font-variant-numeric:tabular-nums">${ped.get("total_usd",0):,.2f} USD</span>'
                '</div>', unsafe_allow_html=True)
            st.markdown(fsm_timeline_html(_est_p), unsafe_allow_html=True)
            cl1,cl2,cl3=st.columns(3)
            cl1.markdown(f"**Cliente:** {ped.get('client_name','')}"); cl1.markdown(f"**Email:** {ped.get('client_email','')}")
            cl2.markdown(f"**Destino:** {ped.get('destino','')}"); cl2.markdown(f"**Fecha:** {ped.get('fecha','')[:10]}")
            cl3.markdown(f"**Modalidad:** {ped.get('tipo_precio','FOB')}"); cl3.markdown(f"**T. pago:** {ped.get('terminos_pago','') or '—'}")
            # Desglose de productos (detallado)
            _prod_p = ped.get('productos', [])
            if _prod_p:
                _rows_p = ''.join(
                    '<tr>'
                    f'<td style="padding:6px 10px;border-bottom:1px solid #eef2f6">{_esc(_ip.get("producto",""))}</td>'
                    f'<td style="padding:6px 10px;border-bottom:1px solid #eef2f6;text-align:right;font-variant-numeric:tabular-nums">{_ip.get("cajas",0):,} cj</td>'
                    f'<td style="padding:6px 10px;border-bottom:1px solid #eef2f6;text-align:right;font-variant-numeric:tabular-nums">{_ip.get("pallets",0):.2f} pal</td>'
                    f'<td style="padding:6px 10px;border-bottom:1px solid #eef2f6;text-align:right;font-variant-numeric:tabular-nums">${_ip.get("precio_usd",0):.2f}</td>'
                    f'<td style="padding:6px 10px;border-bottom:1px solid #eef2f6;text-align:right;font-weight:700;font-variant-numeric:tabular-nums">${_ip.get("total",0):,.2f}</td>'
                    '</tr>' for _ip in _prod_p)
                _tot_cj = sum(int(_ip.get('cajas',0)) for _ip in _prod_p)
                _tot_pl = sum(float(_ip.get('pallets',0)) for _ip in _prod_p)
                st.markdown(
                    '<div style="background:#fff;border:1px solid #e7eaef;border-radius:12px;overflow:hidden;margin:8px 0 4px">'
                    '<table style="width:100%;border-collapse:collapse;font-size:.82rem">'
                    '<thead><tr style="background:#f5f7f8;color:#566472;font-size:.7rem;text-transform:uppercase;letter-spacing:.04em">'
                    '<th style="padding:7px 10px;text-align:left">Producto</th><th style="padding:7px 10px;text-align:right">Cajas</th>'
                    '<th style="padding:7px 10px;text-align:right">Pallets</th><th style="padding:7px 10px;text-align:right">$/cj</th>'
                    '<th style="padding:7px 10px;text-align:right">Total</th></tr></thead>'
                    f'<tbody>{_rows_p}</tbody>'
                    f'<tfoot><tr style="background:#f5f7f8;font-weight:800;color:#084a37">'
                    f'<td style="padding:7px 10px">TOTAL</td>'
                    f'<td style="padding:7px 10px;text-align:right;font-variant-numeric:tabular-nums">{_tot_cj:,} cj</td>'
                    f'<td style="padding:7px 10px;text-align:right;font-variant-numeric:tabular-nums">{_tot_pl:.2f} pal</td>'
                    f'<td></td><td style="padding:7px 10px;text-align:right;font-variant-numeric:tabular-nums">${ped.get("total_usd",0):,.2f}</td>'
                    '</tr></tfoot></table></div>', unsafe_allow_html=True)
            _bl_key = f'bl_{ped.get("id","")}'
            _bl_val = ped.get('bl_numero','')
            _bl_new = st.text_input('🛳️ Nº BL / Contenedor', value=_bl_val, key=_bl_key, placeholder='ej: MSKU1234567')
            if _bl_new != _bl_val:
                _all_p2 = load_pedidos()
                for _ip2, _pp2 in enumerate(_all_p2):
                    if _pp2.get('id') == ped.get('id'): _all_p2[_ip2]['bl_numero'] = _bl_new; break
                save_pedidos(_all_p2); st.cache_data.clear()
            if ped.get('productos'):
                st.dataframe(pd.DataFrame(ped['productos'])[['codigo','producto','cajas','pallets','precio_usd','total']].rename(columns={'codigo':'Código','producto':'Producto','cajas':'Cajas','pallets':'Pallets','precio_usd':'Precio USD','total':'Total USD'}),use_container_width=True,hide_index=True)
            if ped.get('notas'): st.markdown(f"**Notas:** {ped['notas']}")
            # PEND4: Cambio rapido de estado en linea
            _est_actual = ped.get('estado','Recibido')
            _se1, _se2 = st.columns([3,1])
            with _se1:
                _nuevo_est = st.selectbox(f'🚚 Estado actual del pedido', ORDEN_ESTADOS, index=ORDEN_ESTADOS.index(_est_actual) if _est_actual in ORDEN_ESTADOS else 0, key=f"se_est_{ped.get('id','')}")
            with _se2:
                st.markdown('<br>', unsafe_allow_html=True)
                if st.button('Guardar estado', key=f"se_btn_{ped.get('id','')}", use_container_width=True, type='secondary'):
                    if _nuevo_est != _est_actual:
                        _all_pe = load_pedidos()
                        for _ipe, _ppe in enumerate(_all_pe):
                            if _ppe.get('id') == ped.get('id'):
                                _all_pe[_ipe]['estado'] = _nuevo_est
                                _hist = _all_pe[_ipe].get('historial_estados', [])
                                _hist.append({'estado': _nuevo_est, 'fecha': datetime.now().isoformat(), 'usuario': st.session_state.get('user_email','admin')})
                                _all_pe[_ipe]['historial_estados'] = _hist
                                break
                        save_pedidos(_all_pe); st.cache_data.clear()
                        sync_finanzas(_all_pe[_ipe], _all_pe)
                        # Avisar al cliente del nuevo estado (email auto + link WhatsApp)
                        _eml_ok = send_status_email(_all_pe[_ipe], _nuevo_est)
                        st.session_state['_status_notified'] = {
                            'pid': _all_pe[_ipe].get('id', ''), 'estado': _nuevo_est,
                            'email_ok': _eml_ok, 'cliente': _all_pe[_ipe].get('client_name', ''),
                            'email': _all_pe[_ipe].get('client_email', ''),
                            'wa': _client_wa_link(_all_pe[_ipe], _nuevo_est)}
                        st.rerun()
                    else:
                        st.toast('El estado no cambio', icon='ℹ')
            if REPORTLAB_OK:
                with st.expander('✒️ Editar pedido',expanded=False):
                    ec1,ec2=st.columns(2)
                    new_nom_g=ec1.text_input('Nombre',value=ped.get('client_name',''),key=f'g_nom_{ped.get("id","")}' )
                    new_eml_g=ec2.text_input('Email',value=ped.get('client_email',''),key=f'g_eml_{ped.get("id","")}')
                    ed1,ed2=st.columns(2)
                    new_dst_g=ed1.text_input('Destino',value=ped.get('destino',''),key=f'g_dst_{ped.get("id","")}' )
                    new_not_g=ed2.text_input('Notas',value=ped.get('notas',''),key=f'g_not_{ped.get("id","")}')
                    et1,et2=st.columns(2)
                    TOPTG=['','Pago anticipado 100%','50% adelanto / 50% contra documentos','30% adelanto / 70% contra BL','Carta de crédito (LC)','Pago a 30 días','Pago a 60 días','Otro']
                    cur_tg=ped.get('terminos_pago',''); t_ig=TOPTG.index(cur_tg) if cur_tg in TOPTG else 0
                    new_ter_g=et1.selectbox('Términos',TOPTG,index=t_ig,key=f'g_ter_{ped.get("id","")}')
                    new_ent_g=et2.text_input('Fecha entrega',value=ped.get('fecha_entrega',''),placeholder='ej: 2026-06-20',key=f'g_ent_{ped.get("id","")}')
                    ep_rg=[{'Cod':i.get('codigo',''),'Producto':i.get('producto',''),'Cajas':int(i.get('cajas',0)),'Precio_USD':float(i.get('precio_usd',0))} for i in ped.get('productos',[])]
                    if ep_rg:
                        ep_eg=st.data_editor(pd.DataFrame(ep_rg),column_config={'Cod':st.column_config.TextColumn('Cod',disabled=True),'Producto':st.column_config.TextColumn('Prod',disabled=True),'Cajas':st.column_config.NumberColumn('Cajas',min_value=1,step=1),'Precio_USD':st.column_config.NumberColumn('$/cj',format='$%.4f')},use_container_width=True,num_rows='dynamic',key=f'g_ep_{ped.get("id","")}',hide_index=True)
                    if st.button('💾 Guardar cambios',key=f'g_save_{ped.get("id","")}',type='primary'):
                        all_p=load_pedidos(); dd_g=load_data()
                        for _i,_p in enumerate(all_p):
                            if _p.get('id')==ped.get('id'):
                                all_p[_i].update({'client_name':new_nom_g,'client_email':new_eml_g,'destino':new_dst_g,'notas':new_not_g,'terminos_pago':new_ter_g,'fecha_entrega':new_ent_g})
                                nw_it=[]
                                for _,rr in ep_eg.iterrows():
                                    c3=int(rr['Cajas']); gp3=next((p.get('grupo','') for p in dd_g.get('products',[]) if p.get('codigo')==rr['Cod']),'')
                                    gi3=dd_g.get('config',{}).get('grupos',{}).get(gp3,{}); cx3=gi3.get('cajas_pallet',160) if isinstance(gi3,dict) else 160
                                    nw_it.append({'codigo':str(rr['Cod']),'producto':str(rr['Producto']),'cajas':c3,'pallets':round(c3/cx3,2),'precio_usd':float(rr['Precio_USD']),'total':round(c3*float(rr['Precio_USD']),2)})
                                all_p[_i]['productos']=nw_it; all_p[_i]['total_usd']=round(sum(it['total'] for it in nw_it),2)
                                all_p[_i].setdefault('historial_estados',[]).append({'estado':'Editado','fecha':datetime.now().isoformat(),'usuario':st.session_state.user_email,'nota':'Editado manualmente'})
                                break
                        save_pedidos(all_p); st.toast('Pedido actualizado',icon='✅'); st.rerun()
            hist_g=ped.get('historial_estados',[])
            if hist_g:
                with st.expander(f'📜 Timeline de eventos ({len(hist_g)})',expanded=False):
                    st.markdown(eventos_timeline_html(hist_g), unsafe_allow_html=True)
            _pdf_b, _pdf_m, _pdf_x = build_order_pdf(ped)
            st.download_button('⬇️ Albarán PDF', data=_pdf_b, file_name=f"{ped.get('id','ped')}{_pdf_x}", mime=_pdf_m, key=f'pdf_adm_{ped.get("id","")}', use_container_width=True)
            st.markdown('**Cambiar Estado — clic rápido:**')
            estado_actual = ped.get('estado','Recibido')
            qb_cols = st.columns(len(ORDEN_ESTADOS))
            for qi, qe in enumerate(ORDEN_ESTADOS):
                _icon = ESTADO_ICONS.get(qe,'')
                _is_current = (qe == estado_actual)
                if qb_cols[qi].button(f'{_icon} {qe}', key=f'qb_{ped["id"]}_{qe}', type='primary' if _is_current else 'secondary', use_container_width=True):
                    if not _is_current:
                        todos=load_pedidos()
                        for _i,_p in enumerate(todos):
                            if _p.get('id')==ped.get('id'):
                                todos[_i]['estado']=qe
                                todos[_i].setdefault('historial_estados',[]).append({'estado':qe,'fecha':datetime.now().isoformat(),'usuario':st.session_state.user_email})
                                break
                        save_pedidos(todos); st.cache_data.clear()
                        sync_finanzas(todos[_i], todos)
                        _eml_ok = send_status_email(todos[_i], qe)
                        st.session_state['_status_notified'] = {
                            'pid': todos[_i].get('id', ''), 'estado': qe, 'email_ok': _eml_ok,
                            'cliente': todos[_i].get('client_name', ''),
                            'email': todos[_i].get('client_email', ''),
                            'wa': _client_wa_link(todos[_i], qe)}
                        st.rerun()

# ─── TAB CONFIGURACION ──────────────────────────────────────────────
def render_configuracion():
    _admin_seccion('Configuración del sistema', '⚙️')
    _admin_seccion('Usuarios', '👤')
    _users_file = 'users_custom.json'
    _users_data = _load(_users_file, {})
    _all_users_display = []
    for _ue, _uv in USERS.items():
        _all_users_display.append({'Email': _ue, 'Nombre': _uv['nombre'], 'Rol': _uv['rol'], 'Tipo': '🔒 Sistema'})
    for _ue, _uv in _users_data.items():
        if _ue not in USERS:
            _all_users_display.append({'Email': _ue, 'Nombre': _uv.get('nombre',''), 'Rol': _uv.get('rol','ventas'), 'Tipo': 'Custom'})
    if _all_users_display:
        st.dataframe(pd.DataFrame(_all_users_display), use_container_width=True, hide_index=True)
    else:
        st.info('\u2139\ufe0f No hay usuarios configurados. Usa el panel de abajo para agregar uno.')
    # Alta / cambio de contraseña — formulario limpio (tarjeta + botón ancho)
    with st.expander('➕ Añadir usuario o cambiar contraseña', expanded=False):
        with st.form('cfg_user_form', clear_on_submit=False):
            _un1, _un2 = st.columns(2)
            _new_email = _un1.text_input('Email')
            _new_nombre = _un2.text_input('Nombre')
            _np1, _np2 = st.columns(2)
            _new_pwd = _np1.text_input('Contraseña', type='password')
            _new_rol = _np2.selectbox('Rol', ['admin', 'ventas'])
            st.caption('Si el email ya existe, se actualiza su contraseña y rol.')
            _save_user = st.form_submit_button('💾 Guardar usuario', type='primary', use_container_width=True)
        if _save_user:
            import re as _re_us
            _em_us = (_new_email or '').strip().lower()
            if not _re_us.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', _em_us):
                st.error('❌ Email no válido')
            elif not _new_pwd:
                st.error('❌ La contraseña es obligatoria')
            else:
                _users_data[_em_us] = {'pwd': hashlib.md5(_new_pwd.encode()).hexdigest(), 'rol': _new_rol, 'nombre': (_new_nombre or '').strip() or _em_us}
                _save(_users_file, _users_data)
                st.toast(f'Usuario {_em_us} guardado', icon='✅')
                st.rerun()
    if _users_data:
        with st.expander('🗑️ Eliminar usuario', expanded=False):
            _del_user = st.selectbox('Usuario a eliminar', list(_users_data.keys()), key='cfg_del_user')
            if st.button('🗑️ Eliminar', key='cfg_del_btn', type='secondary'):
                del _users_data[_del_user]
                _save(_users_file, _users_data)
                st.toast(f'Usuario {_del_user} eliminado')
                st.rerun()
    _admin_seccion('Log de emails', '✉️')
    elog=load_email_log()
    if elog:
        df_e=pd.DataFrame(elog[-20:][::-1])
        st.dataframe(df_e[['id','destinatario','asunto','tipo','fecha','estado']].rename(columns={'id':'ID','destinatario':'Para','asunto':'Asunto','tipo':'Tipo','fecha':'Fecha','estado':'Estado'}),use_container_width=True,hide_index=True)
    else: st.info('Sin emails registrados')
    st.markdown('---')
    st.markdown('---')
    st.markdown('### 📨 Emails de Pedidos')
    _pend_e = _load('pending_emails.json', [])
    _unsent_e = [e for e in _pend_e if not e.get('sent')]
    if _unsent_e:
        st.error(f'⚠️ {len(_unsent_e)} pedido(s) sin email enviado — configurar SMTP en Streamlit Secrets')
        for _ue in reversed(_unsent_e[-5:]):
            st.caption(f'❌ {_ue.get("id","")} | {_ue.get("fecha","")[:16]} | {_ue.get("cliente","")} | ${_ue.get("total",0):,.2f} | {_ue.get("error","")}')
    elif _pend_e:
        st.success(f'{len([e for e in _pend_e if e.get("sent")])} email(s) enviados correctamente')
    else:
        st.info('📬 Sin historial de emails aún')
    st.markdown('---')
    st.markdown('### Estado SMTP (order@exportharet.com)')
    try:
        _smtp_cfg = st.secrets.get('email', {})
        _smtp_host = _smtp_cfg.get('smtp_host', '')
        if _smtp_host:
            st.success(f'SMTP activo: {_smtp_cfg.get("smtp_user","?")} → {_smtp_host}:{_smtp_cfg.get("smtp_port",587)} | Emails van a order@exportharet.com')
        else:
            st.warning('⚠️ SMTP no configurado — agregar en Streamlit Secrets: [email] smtp_host / smtp_user / smtp_pass')
    except Exception as e:
        logger.debug(f'SMTP secrets not accessible: {e}')
        st.info('ℹ️ Configura SMTP en Streamlit Cloud → App settings → Secrets.')
    st.markdown('---')
    st.markdown('### 🔑 Copia de seguridad (Gist) y token de GitHub')
    if outbox and outbox.configurado():
        try:
            import requests as _rq
            _r = _rq.get('https://api.github.com/user',
                         headers={'Authorization': f'token {outbox._token()}'}, timeout=6)
            _exp = _r.headers.get('github-authentication-token-expiration', '')
            _scopes = _r.headers.get('x-oauth-scopes', '')
            if _r.status_code != 200:
                st.error('🔴 El token de GitHub **no funciona** (HTTP %d). Los pedidos/clientes '
                         'NO se están guardando en el Gist. Crea uno nuevo con permiso `gist` y '
                         'actualiza Secrets `[github] token`.' % _r.status_code)
            elif _exp:
                try:
                    _expd = datetime.strptime(_exp[:10], '%Y-%m-%d')
                    _dias = (_expd - datetime.now()).days
                    if _dias < 0:
                        st.error(f'🔴 El token **CADUCÓ** el {_exp[:10]}. Los pedidos/clientes nuevos '
                                 'NO se guardan. Crea uno nuevo (permiso `gist`) y actualiza Secrets '
                                 '`[github] token`.')
                    elif _dias <= 14:
                        st.warning(f'🟠 El token **caduca en {_dias} días** ({_exp[:10]}). Renuévalo '
                                   'pronto: nuevo token con permiso `gist` → Secrets `[github] token`, '
                                   'o te quedarás sin copia de seguridad.')
                    else:
                        st.success(f'✅ Copia de seguridad activa. Token (`{_scopes}`) caduca el '
                                   f'{_exp[:10]} — faltan **{_dias} días**.')
                except Exception:
                    st.info(f'Token activo. Caducidad: {_exp}')
            else:
                st.success('✅ Copia de seguridad activa. Token sin fecha de caducidad.')
        except Exception as _e:
            st.info(f'No se pudo verificar el token ahora: {_e}')
    else:
        st.warning('⚠️ Gist **no configurado**: los pedidos/clientes no se guardan de forma '
                   'durable (se pierden al reiniciarse la app). Configura `[github] token` + '
                   '`gist_id` en Secrets.')
    st.markdown('---')
    st.markdown('### 🗃️ Archivos de Datos')
    for fname in [DATA_FILE,CLIENTS_FILE,PEDIDOS_FILE,HIST_FILE,EMAIL_FILE]:
        exists=os.path.exists(fname)
        st.markdown(f"{'\u2705' if exists else '\u274C'} `{fname}`")
    st.markdown('---')
    st.markdown('### 🖼️ Logotipo de la Aplicación')
    _logo_col1, _logo_col2 = st.columns([2, 1])
    with _logo_col1:
        _logo_file = st.file_uploader('Subir nuevo logo (PNG o JPG)', type=['png','jpg','jpeg'], key='cfg_logo_upload')
        if _logo_file is not None:
            with open('logo.png', 'wb') as _lf: _lf.write(_logo_file.getbuffer())
            st.success('Logo actualizado. Recarga la página para verlo.')
    with _logo_col2:
        import os as _oscfg
        if _oscfg.path.exists('logo.png'):
            from PIL import Image as _ImgCfg
            st.image(_ImgCfg.open('logo.png'), width=140)
        else:
            st.info('Sin logo')
    st.markdown('---')
    st.markdown('### ✏️ Título de la Aplicación')
    _cur_cfg = load_app_config()
    _cur_title = _cur_cfg.get("app_title", "📊 Export Haret — Panel de Administración")
    _new_title = st.text_input('Título del panel de administración', value=_cur_title, key='cfg_app_title', help='Aparece en el header y sidebar del panel admin.')
    if st.button('💾 Guardar Título', key='cfg_save_title'):
        _cur_cfg["app_title"] = _new_title
        save_app_config(_cur_cfg)
        st.success('Título guardado. Recarga para verlo en el header y sidebar.')

    # ─── Log de Cambios de Mínimos ─────────────────────────────
    st.markdown("---")
    st.markdown("### 📝 Log de Cambios de Mínimos por Producto")
    _min_changes = load_min_log()
    if not _min_changes:
        st.info("Aún no hay cambios de mínimos registrados.")
    else:
        _df_min = pd.DataFrame(_min_changes[::-1])
        _df_min.columns = ["Fecha y Hora", "Usuario", "Código", "Producto", "Unidad Antes", "Cant. Antes", "Unidad Después", "Cant. Después"]
        _prods_u = ["Todos"] + sorted(_df_min["Producto"].unique().tolist())
        _filtro_m = st.selectbox("🔍 Filtrar por producto:", _prods_u, key="filtro_min_log")
        if _filtro_m != "Todos":
            _df_min = _df_min[_df_min["Producto"] == _filtro_m]
        st.dataframe(_df_min, use_container_width=True, hide_index=True)
        st.caption(f"Total de cambios registrados: {len(_min_changes)}")


    # ─── Historial de Accesos ─────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 🔐 Historial de Accesos de Clientes")
    _accesos = load_accesos()
    if not _accesos:
        st.info("Aún no hay registros de acceso.")
    else:
        _df_acc = pd.DataFrame(_accesos[::-1])
        _df_acc.columns = ["Fecha y Hora", "Email", "Nombre", "Rol"]
        _emails_u = ["Todos"] + sorted(_df_acc["Email"].unique().tolist())
        _filtro = st.selectbox("🔍 Filtrar por usuario:", _emails_u, key="filtro_accesos_hist")
        if _filtro != "Todos":
            _df_acc = _df_acc[_df_acc["Email"] == _filtro]
        st.dataframe(_df_acc, use_container_width=True, hide_index=True)
        st.caption(f"Total de accesos registrados: {len(_accesos)}")
# ─── TAB CLIENTES ──────────────────────────────────────────────
def _migrate_clients_swap(clients):
    """Normaliza el padrón del admin: re-clava cada ficha por su EMAIL, corrige el swap
    nombre<->email y COLAPSA duplicados del mismo cliente SIN perder pedidos."""
    import re as _re
    _email_re = _re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')

    def _fusionar(dst, src):
        out = dict(dst)
        for f in ('empresa', 'telefono', 'pais', 'razon_social', 'fecha_registro',
                  'primer_pedido', 'ultimo_pedido', 'origen'):
            if not out.get(f) and src.get(f):
                out[f] = src.get(f)
        # nombre legible: si el de dst parece email y el de src no, usar el de src
        if _email_re.match(out.get('nombre', '') or '') and src.get('nombre') and not _email_re.match(src['nombre']):
            out['nombre'] = src['nombre']
        # unir pedidos por id (sin perder ninguno)
        _ped = {}
        for ped in (dst.get('pedidos') or []) + (src.get('pedidos') or []):
            _ped[ped.get('id') if isinstance(ped, dict) else ped] = ped
        if _ped:
            out['pedidos'] = list(_ped.values())
        _pi = set(dst.get('pedidos_ids', []) or []) | set(src.get('pedidos_ids', []) or [])
        if _pi:
            out['pedidos_ids'] = sorted(_pi)
        return out

    _changed = False
    _new = {}
    for _k, _v in (clients or {}).items():
        _v = dict(_v)
        _nom = (_v.get('nombre') or '').strip()
        _eml = (_v.get('email') or '').strip()
        # email real: del campo email, de la clave, o del nombre (swap)
        real = None
        for cand in (_eml, _k or '', _nom):
            if _email_re.match(cand or ''):
                real = cand.lower()
                break
        if real:
            if (_v.get('email', '') or '').lower() != real:
                _v['email'] = real; _changed = True
            if _email_re.match(_v.get('nombre', '') or ''):   # nombre era un email → recuperar uno legible
                _v['nombre'] = (_k if (_k and not _email_re.match(_k)) else
                                (_v.get('razon_social') or _v.get('empresa') or real.split('@')[0]))
                _changed = True
            if _k != real:
                _changed = True
            _new[real] = _fusionar(_new[real], _v) if real in _new else _v
        else:
            _new[_k] = _fusionar(_new[_k], _v) if _k in _new else _v
    if _changed:
        try: save_clients(_new)
        except Exception: pass
    return _new

def _merge_client_record(existing, updates):
    """Fusiona el registro de un cliente SIN perder datos: aplica `updates` pero
    NUNCA sobreescribe un campo que ya tiene valor con uno vacío (evita borrar
    nombre/empresa/teléfono/país al guardar datos o confirmar un pedido). Conserva
    además cualquier campo extra del registro existente (p. ej. de Marketing)."""
    out = dict(existing or {})
    for _k, _v in (updates or {}).items():
        if _v not in (None, '', [], {}):
            out[_k] = _v          # valor nuevo no vacío → actualizar
        elif _k not in out:
            out[_k] = _v          # campo nuevo (aunque vacío) → crear
        # si el valor nuevo es vacío y el campo ya existía → conservar el viejo
    return out

def _dedupe_portal_clients(pc):
    """Normaliza el padrón del portal: clave email en minúsculas + sin espacios, y fusiona
    duplicados del mismo cliente (une sus listas de pedidos). Devuelve un dict sin duplicados."""
    out = {}
    for _k, _v in (pc or {}).items():
        _ek = (_k or '').strip().lower()
        if not _ek or not isinstance(_v, dict):
            continue
        _v = dict(_v); _v['email'] = _ek
        if _ek in out:
            _prev = out[_ek]
            for _f in ('nombre', 'empresa', 'telefono', 'pais', 'fecha_registro'):
                if not _prev.get(_f) and _v.get(_f):
                    _prev[_f] = _v[_f]
            try:
                _un = list(dict.fromkeys((_prev.get('pedidos') or []) + (_v.get('pedidos') or [])))
                if _un:
                    _prev['pedidos'] = _un
            except TypeError:
                pass
        else:
            out[_ek] = _v
    return out

def render_clientes():
    _admin_seccion('Clientes', '👥')
    clients=load_clients()
    clients=_migrate_clients_swap(clients)

    # ➕ Pre-registrar un cliente: al entrar al portal con su email ya estará reconocido
    with st.expander('➕ Agregar / pre-registrar cliente', expanded=False):
        st.caption('Crea la cuenta del cliente. Cuando entre al portal con su email, '
                   'ya aparecerá registrado y con sus datos cargados.')
        _ac1, _ac2 = st.columns(2)
        _ac_email = _ac1.text_input('📧 Email *', key='ac_email')
        _ac_nombre = _ac2.text_input('👤 Nombre / contacto *', key='ac_nombre')
        _ac3, _ac4 = st.columns(2)
        _ac_empresa = _ac3.text_input('🏢 Empresa', key='ac_empresa')
        _ac_tel = _ac4.text_input('📱 Teléfono / WhatsApp', key='ac_tel')
        _ac_pais = st.text_input('🌍 País', value='Spain', key='ac_pais')
        if st.button('💾 Crear cliente', type='primary', key='ac_save'):
            import re as _re_ac
            _em = (_ac_email or '').strip().lower()
            _nm = (_ac_nombre or '').strip()
            if not _re_ac.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', _em):
                st.error('❌ Email inválido')
            elif not _nm:
                st.error('❌ El nombre es obligatorio')
            else:
                _now = datetime.now().isoformat()
                _pc = load_portal_clients()
                _ya = _em in _pc
                _pc[_em] = {'nombre': _nm, 'empresa': (_ac_empresa or '').strip(),
                            'telefono': (_ac_tel or '').strip(), 'pais': (_ac_pais or '').strip(),
                            'email': _em,
                            'fecha_registro': _pc.get(_em, {}).get('fecha_registro', _now),
                            'pedidos': _pc.get(_em, {}).get('pedidos', [])}
                save_portal_clients(_pc)  # local + Gist (sobrevive a reinicios)
                _adm = load_clients()
                _adm[_em] = {'nombre': _nm, 'email': _em, 'empresa': (_ac_empresa or '').strip(),
                             'telefono': (_ac_tel or '').strip(), 'pais': (_ac_pais or '').strip(),
                             'fecha_registro': _adm.get(_em, {}).get('fecha_registro', _now),
                             'pedidos_ids': _adm.get(_em, {}).get('pedidos_ids', []),
                             'origen': 'admin_prereg'}
                save_clients(_adm)
                st.cache_data.clear()
                st.success(f'✅ Cliente {_nm} ({_em}) {"actualizado" if _ya else "creado"}. '
                           'Ya puede entrar al portal y estará registrado.')
                st.rerun()

    # Merge portal-registered clients so admin can see all — CASE-INSENSITIVE (sin duplicar).
    # El padrón del portal se normaliza (clave email en minúsculas) y se fusiona por email.
    try:
        _raw_pc = load_portal_clients()
        _pc = _dedupe_portal_clients(_raw_pc)
        if len(_pc) != len(_raw_pc):
            save_portal_clients(_pc)   # limpia los duplicados también en el padrón/Gist (una vez)
        _existing = {(k or '').strip().lower() for k in clients}
        for _pe, _pv in (_pc or {}).items():
            _k = (_pe or '').strip().lower()
            if _k and _k not in _existing:
                clients[_k] = {'nombre': _pv.get('nombre',''), 'email': _k, 'empresa': _pv.get('empresa',''), 'telefono': _pv.get('telefono',''), 'pais': _pv.get('pais',''), 'fecha_registro': _pv.get('fecha_registro',''), 'pedidos_ids': _pv.get('pedidos',[]), 'origen': 'portal_cliente'}
                _existing.add(_k)
        save_clients(clients)
    except Exception:
        pass
    pedidos=load_pedidos()
    if not clients: st.info('No hay clientes. Se crean al hacer pedidos.'); return
    _search = st.text_input('🔍 Buscar cliente', key='cli_search', placeholder='Nombre, email, empresa...')
    rows=[]
    for e,c in clients.items():
        if _search and _search.lower() not in (e + c.get('nombre','') + c.get('empresa','')).lower(): continue
        seg=segmentar(e,clients)
        mp=[p for p in pedidos if p.get('client_email')==e]
        rows.append({'Nombre':c.get('nombre',''),'Email':e,'Empresa':c.get('empresa',''),'País':c.get('pais',''),'Segmento':seg['badge'],'Pedidos':len(mp),'Facturación':f"${sum(p.get('total_usd',0) for p in mp):,.2f}",'Descuento':f"{seg['descuento']*100:.0f}%"})
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.caption(f'{len(rows)} de {len(clients)} clientes')

    # ── Ficha de cliente (Perfil 360, estilo Finanzas) ──
    _admin_seccion('Ficha de cliente', '🪪')
    _cli_keys = sorted(clients.keys(), key=lambda e: (clients[e].get('nombre','') or e).lower())
    _sel_cli = st.selectbox('Cliente', _cli_keys,
                            format_func=lambda e: f"{clients[e].get('nombre','') or '(sin nombre)'} · {e}",
                            key='cli_ficha_sel')
    if _sel_cli:
        _c = clients[_sel_cli]
        _mp = sorted([p for p in pedidos if (p.get('client_email','') or '').strip().lower() == _sel_cli.strip().lower()],
                     key=lambda x: x.get('fecha',''), reverse=True)
        _vol = sum(p.get('total_usd',0) for p in _mp)
        _np = len(_mp)
        _ticket = (_vol/_np) if _np else 0.0
        _seg = segmentar(_sel_cli, clients)
        _ult = _mp[0].get('fecha','')[:10] if _mp else '—'
        _rec = ''
        if _mp:
            try:
                _d = (datetime.now() - datetime.fromisoformat(_mp[0].get('fecha'))).days
                _rec = 'hoy' if _d <= 0 else (f'hace {_d} d' if _d < 60 else f'hace {_d//30} m')
            except Exception:
                _rec = ''
        _activos = len([p for p in _mp if p.get('estado') in ('Recibido','Confirmado','Preparando')])
        _ini = ((_c.get('nombre') or _sel_cli).strip()[:1] or '·').upper()
        _sub = ' · '.join([x for x in [_c.get('empresa',''), _c.get('pais',''), _sel_cli] if x])
        # Cabecera de ficha (avatar + nombre + segmento)
        st.markdown(
            '<div style="display:flex;align-items:center;gap:13px;background:#fff;border:1px solid #e7eaef;'
            'border-radius:14px;padding:14px 18px;margin:4px 0 10px;box-shadow:0 1px 2px rgba(18,28,42,.05)">'
            f'<div style="width:44px;height:44px;border-radius:50%;background:#0c6e51;color:#fff;font-weight:800;'
            f'font-size:1.1rem;display:flex;align-items:center;justify-content:center;flex:0 0 auto">{_esc(_ini)}</div>'
            '<div style="min-width:0;line-height:1.25">'
            f'<div style="font-weight:800;color:#131a21;font-size:1.08rem">{_esc(_c.get("nombre","") or "(sin nombre)")}</div>'
            f'<div style="color:#8b95a3;font-size:.82rem;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_esc(_sub)}</div></div>'
            f'<div style="margin-left:auto;flex:0 0 auto;background:#eef6f2;color:#0b5a42;border:1px solid rgba(12,110,81,.2);'
            f'border-radius:999px;padding:4px 12px;font-size:.76rem;font-weight:700;white-space:nowrap">{_esc(_seg["badge"])}</div>'
            '</div>', unsafe_allow_html=True)
        # KPIs (tarjetas)
        _k1,_k2,_k3,_k4 = st.columns(4)
        _k1.metric('💵 Volumen total', f'${_vol:,.0f}', 'USD')
        _k2.metric('📦 Pedidos', _np, f'{_activos} activos' if _activos else None)
        _k3.metric('🎟️ Ticket medio', f'${_ticket:,.0f}')
        _k4.metric('🕒 Último pedido', _ult, _rec or None)
        # Contacto rápido
        _cc1, _cc2, _cc3 = st.columns(3)
        _tel_d = ''.join(ch for ch in str(_c.get('telefono','') or '') if ch.isdigit())
        if _tel_d:
            _cc1.link_button('📲 WhatsApp', f'https://wa.me/{_tel_d}', use_container_width=True)
        _cc2.link_button('✉️ Email', f'mailto:{_sel_cli}', use_container_width=True)
        if _np:
            if _cc3.button('📦 Ver sus pedidos', use_container_width=True, key='cli_ficha_pedidos'):
                # clave pendiente: se aplica al buscador de Pedidos ANTES de crear el widget
                st.session_state['_gp_c_pending'] = _sel_cli
                st.rerun()
        # Historial de pedidos (con badge de estado)
        st.markdown('<div style="font-size:.7rem;text-transform:uppercase;letter-spacing:.06em;color:#8b95a3;'
                    'font-weight:700;margin:12px 0 4px">Historial de pedidos</div>', unsafe_allow_html=True)
        if _mp:
            for _p in _mp[:15]:
                _pid = (_p.get('id','') or '').upper()
                _pf = (_p.get('fecha','') or '')[:10]
                st.markdown(
                    '<div style="display:flex;align-items:center;gap:10px;background:#fff;border:1px solid #e7eaef;'
                    'border-radius:10px;padding:8px 13px;margin-bottom:6px">'
                    f'{estado_badge(_p.get("estado","Recibido"))}'
                    f'<span style="font-weight:700;color:#131a21;font-size:.85rem">#{_esc(_pid)}</span>'
                    f'<span style="color:#8b95a3;font-size:.78rem">{_esc(_pf)}</span>'
                    f'<span style="margin-left:auto;font-weight:800;color:#084a37;font-variant-numeric:tabular-nums">${_p.get("total_usd",0):,.2f}</span>'
                    '</div>', unsafe_allow_html=True)
            if _np > 15:
                st.caption(f'… y {_np-15} pedido(s) más')
        else:
            st.info('Este cliente aún no tiene pedidos.')

    with st.expander('🗑️ Eliminar Cliente', expanded=False):
        _cli_opts = list(clients.keys())
        _del_cli = st.selectbox('Cliente a eliminar', _cli_opts, format_func=lambda e: f"{clients[e].get('nombre','')} ({e})", key='cli_del_sel')
        _cli_peds = [p for p in pedidos if p.get('client_email') == _del_cli]
        if _cli_peds: st.warning(f'⚠️ Este cliente tiene {len(_cli_peds)} pedidos. Se conservarán los pedidos.')
        if st.button('🗑️ Eliminar Cliente', key='cli_del_btn', type='secondary'):
            if _del_cli in clients:
                del clients[_del_cli]; save_clients(clients)
                st.toast(f'Cliente {_del_cli} eliminado'); st.rerun()
    # Editar datos del cliente seleccionado arriba (notas internas, términos…)
    if _sel_cli:
        with st.expander('✏️ Editar datos del cliente', expanded=False):
            _ce = clients[_sel_cli]
            _fc1,_fc2=st.columns(2)
            _f_nom=_fc1.text_input('Nombre',value=_ce.get('nombre',''),key='f_nom')
            _f_emp=_fc2.text_input('Empresa',value=_ce.get('empresa',''),key='f_emp')
            _fc3,_fc4=st.columns(2)
            _f_tel=_fc3.text_input('Teléfono/WhatsApp',value=_ce.get('telefono',''),key='f_tel')
            _f_pais=_fc4.text_input('País',value=_ce.get('pais',''),key='f_pais')
            _TOPTC=['','Pago anticipado 100%','50% adelanto / 50% contra documentos','30% adelanto / 70% contra BL','Carta de crédito (LC)','Pago a 30 días','Pago a 60 días','Otro']
            _cur_tc=_ce.get('terminos_habituales',''); _tc_idx=_TOPTC.index(_cur_tc) if _cur_tc in _TOPTC else 0
            _f_term=st.selectbox('Términos habituales',_TOPTC,index=_tc_idx,key='f_term')
            _f_seg=st.text_input('📅 Próximo seguimiento',value=_ce.get('proximo_seguimiento',''),placeholder='ej: 2026-07-01',key='f_seg')
            _f_notas=st.text_area('🔒 Notas internas (solo admin)',value=_ce.get('notas_internas',''),height=90,key='f_notas',placeholder='Preferencias, condiciones especiales...')
            if st.button('💾 Guardar ficha',type='primary',key='save_ficha_cl'):
                clients[_sel_cli] = _merge_client_record(clients.get(_sel_cli, {}),
                    {'nombre':_f_nom,'empresa':_f_emp,'telefono':_f_tel,'pais':_f_pais})
                clients[_sel_cli].update({'terminos_habituales':_f_term,'proximo_seguimiento':_f_seg,'notas_internas':_f_notas})
                save_clients(clients); st.toast('Ficha guardada',icon='✅'); st.rerun()

def render_reportes():
    st.markdown('## 📊 Reportes y Analytics')
    pedidos=load_pedidos(); clients=load_clients()
    if not pedidos: st.info('No hay pedidos para analizar.'); return
    fc1,fc2=st.columns(2)
    fd=fc1.date_input('Desde',value=date.today()-timedelta(days=30),key='r_from')
    fh=fc2.date_input('Hasta',value=date.today(),key='r_to')
    pf=[p for p in pedidos if str(fd)<=p.get('fecha','')[:10]<=str(fh)]
    st.markdown(f'**{len(pf)} pedidos en el período**')
    k1,k2,k3,k4=st.columns(4)
    fac=sum(p.get('total_usd',0) for p in pf)
    k1.metric('💰 Facturación',f'${fac:,.0f}')
    k2.metric('📦 Pedidos',len(pf))
    k3.metric('🎫 Ticket Prom.',f'${fac/len(pf):,.0f}' if pf else '$0')
    k4.metric('Entregados',len([p for p in pf if p.get('estado')=='Entregado']))
    st.markdown('---')
    st.markdown('### Por Destino')
    dd={}
    for p in pf: dd[p.get('destino','Otros')]=dd.get(p.get('destino','Otros'),0)+p.get('total_usd',0)
    if dd: st.dataframe(pd.DataFrame(sorted(dd.items(),key=lambda x:x[1],reverse=True),columns=['Destino','Total USD']),use_container_width=True,hide_index=True)
    st.markdown('---')
    st.markdown('### 👑 Top Clientes')
    cf={}
    for p in pf: cf[p.get('client_email','')]=cf.get(p.get('client_email',''),0)+p.get('total_usd',0)
    if cf:
        t10=sorted(cf.items(),key=lambda x:x[1],reverse=True)[:10]
        st.dataframe(pd.DataFrame([{'Cliente':clients.get(e,{}).get('nombre',e),'Email':e,'Total':f'${v:,.2f}'} for e,v in t10]),use_container_width=True,hide_index=True)
    if pf:
        st.markdown('---')
        xb=exportar_excel(pf)
        if xb: st.download_button('📥 Exportar Reporte',data=xb,file_name=f'reporte_{fd}_{fh}.xlsx',mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',type='primary')

# ─── MAIN ─────────────────────────────────────────────────────────────────

# ─── PORTAL CLIENT FILES ─────────────────────────────────────────────────────
PORTAL_CLIENTS_FILE = 'portal_clientes.json'

def load_portal_clients():
    return _load(PORTAL_CLIENTS_FILE, {})

def save_portal_clients(c):
    c = _dedupe_portal_clients(c)  # clave email normalizada + sin duplicados, también en el Gist
    _save(PORTAL_CLIENTS_FILE, c)
    if outbox:
        try:
            outbox.publish_clients(c)  # persistir el padrón en el Gist (sobrevive a reinicios)
        except Exception as e:
            logger.warning(f'publish_clients falló: {e}')

# ── Carritos pendientes por email (para retomar el pedido al volver) ──
# Store aislado del padrón de clientes: evita crear clientes fantasma y
# mantiene el JSON de carritos pequeño. Clave = email normalizado.
PORTAL_CARTS_FILE = 'portal_carritos.json'

def load_portal_carts():
    return _load(PORTAL_CARTS_FILE, {})

def save_portal_carts(c):
    _save(PORTAL_CARTS_FILE, c)
    if outbox:
        try:
            outbox.publish_carts(c)  # durable en el Gist (sobrevive a reinicios de Cloud)
        except Exception as e:
            logger.warning(f'publish_carts falló: {e}')

def get_fob_price(codigo, data):
    for p in data.get('products', []):
        if p.get('codigo') == codigo:
            fob = p.get('precio_fob_final', 0) or p.get('precio_cif_usd', 0) or p.get('precio_compra', 0)
            return round(float(fob), 2)
    return 0.0

def get_cif_price(codigo, destino, data):
    return get_precio(codigo, destino, data)

def build_order_html(ped):
    rows=''
    for item in ped.get('productos',[]):
        cod=item.get('codigo',''); prod=item.get('producto','')
        cajas=item.get('cajas',0); pallets=item.get('pallets',0)
        precio=item.get('precio_usd',0); total=item.get('total',0)
        fob_u=item.get('fob_usd',precio); flete_u=item.get('flete_usd',0); dto=item.get('descuento_vol',0)
        rows+=(f'<tr><td style="padding:7px;border:1px solid #e0e0e0">{cod}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0">{prod}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:center">{cajas}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:center">{pallets}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:right;color:#856404">${fob_u:.4f}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:right;color:#888">${flete_u:.4f}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:center;color:#28a745">{dto*100:.0f}%</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:right;font-weight:bold;color:#0c6e51">${precio:.4f}</td>'
               f'<td style="padding:7px;border:1px solid #e0e0e0;text-align:right;font-weight:bold">${total:,.2f}</td></tr>')
    pid=ped.get('id',''); fecha=ped.get('fecha','')[:10]; nombre=ped.get('client_name','')
    email_c=ped.get('client_email',''); empresa=ped.get('empresa',''); telefono=ped.get('telefono','')
    pais=ped.get('pais',''); tipo=ped.get('tipo_precio','FOB'); destino=ped.get('destino','')
    total_usd=ped.get('total_usd',0); notas=ped.get('notas',''); estado=ped.get('estado','Recibido')
    terminos=ped.get('terminos_pago',''); f_entrega=ped.get('fecha_entrega','')
    incoterm=f'{tipo} {destino}' if tipo=='CIF' and destino else tipo
    t_row=f'<tr><td colspan="2" style="padding:6px;font-weight:bold;background:#f8f9fa">T\xe9rminos de pago:</td><td colspan="7" style="padding:6px">{terminos}</td></tr>' if terminos else ''
    e_row=f'<tr><td colspan="2" style="padding:6px;font-weight:bold;background:#f8f9fa">Entrega estimada:</td><td colspan="7" style="padding:6px">{f_entrega}</td></tr>' if f_entrega else ''
    return f'''<div style="font-family:Arial,sans-serif;max-width:750px;margin:0 auto;padding:20px">
      <div style="background:linear-gradient(135deg,#0c6e51,#0a5d44);color:white;padding:24px;border-radius:10px;margin-bottom:20px">
        <h1 style="margin:0;font-size:1.6em">🚀 Export Haret</h1>
        <p style="margin:4px 0 0;opacity:.85">Confirmaci\xf3n de Pedido | Frutas Ex\xf3ticas Premium</p>
      </div>
      <table style="width:100%;border-collapse:collapse;margin-bottom:16px">
        <tr><td style="padding:6px;font-weight:bold">N\xba Pedido:</td><td style="padding:6px">{pid}</td>
            <td style="padding:6px;font-weight:bold">Estado:</td><td style="padding:6px">{estado}</td></tr>
        <tr><td style="padding:6px;font-weight:bold">Cliente:</td><td style="padding:6px">{nombre}</td>
            <td style="padding:6px;font-weight:bold">Empresa:</td><td style="padding:6px">{empresa}</td></tr>
        <tr><td style="padding:6px;font-weight:bold">Email:</td><td style="padding:6px">{email_c}</td>
            <td style="padding:6px;font-weight:bold">Tel\xe9fono:</td><td style="padding:6px">{telefono}</td></tr>
        <tr><td style="padding:6px;font-weight:bold">Pa\xeds:</td><td style="padding:6px">{pais}</td>
            <td style="padding:6px;font-weight:bold">Fecha:</td><td style="padding:6px">{fecha}</td></tr>
        <tr><td style="padding:6px;font-weight:bold">Incoterm:</td><td style="padding:6px">{incoterm}</td>
            <td style="padding:6px;font-weight:bold">Destino:</td><td style="padding:6px">{destino}</td></tr>
        {t_row}{e_row}
      </table>
      <h3 style="border-bottom:2px solid #0c6e51;padding-bottom:6px">Detalle de Productos</h3>
      <table style="width:100%;border-collapse:collapse;font-size:.88em">
        <thead><tr style="background:#0c6e51;color:white">
          <th style="padding:8px">C\xf3d</th><th style="padding:8px">Producto</th>
          <th style="padding:8px">Cajas</th><th style="padding:8px">Pallets</th>
          <th style="padding:8px">FOB $/cj</th><th style="padding:8px">Flete $/cj</th>
          <th style="padding:8px">Dto.Vol</th><th style="padding:8px">Precio $/cj</th>
          <th style="padding:8px">Total USD</th>
        </tr></thead><tbody>{rows}</tbody>
        <tfoot><tr style="background:#e8f0fe">
          <td colspan="8" style="padding:10px;text-align:right;font-weight:bold">TOTAL:</td>
          <td style="padding:10px;font-weight:bold;font-size:1.15em;color:#0c6e51">${total_usd:,.2f} USD</td>
        </tr></tfoot></table>
      {f'<p style="margin-top:14px"><b>Notas:</b> {notas}</p>' if notas else ''}
      <p style="margin-top:20px;color:#666;font-size:.83em;border-top:1px solid #eee;padding-top:10px">
        Pedido recibido en order@exportharet.com | Export Haret \u00a9 2026</p>
    </div>'''
def build_order_pdf(ped):
    """Genera un PDF albaran del pedido con reportlab. Retorna bytes del PDF."""
    buf = io.BytesIO()
    # ── PDF translations (ES/EN) ─────────────────────────────────────────
    _pdf_lang = ped.get('lang', 'es')
    PDF_TEXTS = {
        'es': {
            'header_title': 'Export Haret',
            'header_sub': 'Sistema de Pedidos — Frutas Exóticas Premium',
            'doc_title': 'ALBARÁN / ORDEN DE PEDIDO',
            'sec_client': 'DATOS DEL CLIENTE',
            'sec_order': 'DETALLES DEL PEDIDO',
            'name': 'Nombre', 'company': 'Empresa', 'email': 'Email', 'phone': 'Teléfono',
            'incoterm': 'Incoterm', 'order_no': 'Nº Pedido', 'date': 'Fecha', 'status': 'Estado',
            'country': 'País', 'destination': 'Destino',
            'freight': 'Flete', 'fob_origin': 'FOB (en origen)',
            'shipping_from': 'Embarcamos desde Quito/Guayaquil, Ecuador',
            'product_detail': 'DETALLE DE PRODUCTOS',
            'th_code': 'Código', 'th_product': 'Producto', 'th_boxes': 'Cajas',
            'th_pallets': 'Pallets', 'th_price': 'Precio/caja', 'th_total': 'Total USD',
            'total_label': 'TOTAL:', 'notes_label': 'Notas:',
            'packing_title': 'TABLA DE EMBALAJE ESTIMADA',
            'pk_product': 'Producto', 'pk_group': 'Grupo', 'pk_pallets': 'Pallets',
            'pk_boxes': 'Cajas', 'pk_kgbox': 'Kg/Caja', 'pk_weight': 'Peso Total (kg)',
            'pk_bxpal': 'Cj/Plt', 'pk_total': 'TOTAL',
            'pk_note': '* Pesos totales son estimados. El peso real puede variar ±5% según calibre y variedad. No incluye embalaje de pallet.',
            'footer_main': 'Export Haret © 2026 | order@exportharet.com | Frutas Exóticas Premium de Ecuador',
            'footer_sub': 'Los precios USD son la divisa comercial. Precios en moneda destino son referenciales y sujetos a cotización.',
        },
        'en': {
            'header_title': 'Export Haret',
            'header_sub': 'Order System — Premium Exotic Fruits',
            'doc_title': 'ORDER / DELIVERY NOTE',
            'sec_client': 'CUSTOMER DETAILS',
            'sec_order': 'ORDER DETAILS',
            'name': 'Name', 'company': 'Company', 'email': 'Email', 'phone': 'Phone',
            'incoterm': 'Incoterm', 'order_no': 'Order No.', 'date': 'Date', 'status': 'Status',
            'country': 'Country', 'destination': 'Destination',
            'freight': 'Freight', 'fob_origin': 'FOB (at origin)',
            'shipping_from': 'Shipping from Quito/Guayaquil, Ecuador',
            'product_detail': 'PRODUCT DETAIL',
            'th_code': 'Code', 'th_product': 'Product', 'th_boxes': 'Boxes',
            'th_pallets': 'Pallets', 'th_price': 'Price/box', 'th_total': 'Total USD',
            'total_label': 'TOTAL:', 'notes_label': 'Notes:',
            'packing_title': 'ESTIMATED PACKAGING TABLE',
            'pk_product': 'Product', 'pk_group': 'Group', 'pk_pallets': 'Pallets',
            'pk_boxes': 'Boxes', 'pk_kgbox': 'Kg/Box', 'pk_weight': 'Total Weight (kg)',
            'pk_bxpal': 'Bx/Plt', 'pk_total': 'TOTAL',
            'pk_note': '* Total weights are estimates. Actual weight may vary ±5% depending on size and variety. Does not include pallet packaging.',
            'footer_main': 'Export Haret © 2026 | order@exportharet.com | Premium Exotic Fruits from Ecuador',
            'footer_sub': 'USD prices are the commercial currency. Prices in destination currency are referential and subject to quotation.',
        }
    }
    _PT = PDF_TEXTS.get(_pdf_lang, PDF_TEXTS['es'])
    pid = ped.get('id','')
    fecha = ped.get('fecha','')[:10]
    estado = ped.get('estado','Recibido')
    nombre = ped.get('client_name','')
    email_c = ped.get('client_email','')
    empresa = ped.get('empresa','')
    telefono = ped.get('telefono','')
    pais = ped.get('pais','')
    tipo = ped.get('tipo_precio','FOB')
    destino = ped.get('destino','')
    total_usd = ped.get('total_usd', 0)
    notas = ped.get('notas','')
    moneda_dest = ped.get('moneda_dest', 'USD')
    flete_usd_caja = ped.get('flete_usd_caja', 0.0)
    tasa_cambio = ped.get('tasa_cambio', 1.0)
    total_moneda_dest = ped.get('total_moneda_dest', total_usd)
    sym_dest = MONEDA_SIMBOLO.get(moneda_dest, moneda_dest)

    if not REPORTLAB_OK:
        # Fallback: devolver HTML como bytes
        return build_order_html(ped).encode('utf-8'), 'text/html', '.html'

    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Colores corporativos
    AZUL = colors.HexColor('#0c6e51')
    AZUL_LIGHT = colors.HexColor('#E8F0FA')
    GRIS = colors.HexColor('#666666')

    # --- Cabecera ---
    header_style = ParagraphStyle('header', fontSize=22, textColor=colors.white,
        fontName='Helvetica-Bold', spaceAfter=4, alignment=TA_LEFT)
    sub_style = ParagraphStyle('sub', fontSize=10, textColor=colors.HexColor('#CCDDFF'),
        fontName='Helvetica', alignment=TA_LEFT)

    # Logo en cabecera
    _logo_cell = Paragraph(f'<font color="white" size="20"><b>{_PT["header_title"]}</b></font><br/><font color="#CCDDFF" size="9">{_PT["header_sub"]}</font>', styles['Normal'])
    try:
        from reportlab.platypus import Image as RLImage
        import os as _os
        if _os.path.exists('logo.png'):
            _img = RLImage('logo.png', width=4.5*cm, height=1.5*cm)
            _logo_cell = _img
    except Exception:
        pass
    header_data = [[
        _logo_cell,
        Paragraph(f'<font color="white" size="9"><b>{_PT["doc_title"]}</b><br/>{pid}<br/>{fecha}</font>', styles['Normal'])
    ]]
    header_table = Table(header_data, colWidths=[10*cm, 7*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,0), colors.white),
        ('BACKGROUND', (1,0), (1,0), AZUL),
        ('PADDING', (0,0), (-1,-1), 12),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4*cm))

    # --- Datos cliente + pedido ---
    info_data = [
        [Paragraph(f'<b>{_PT["sec_client"]}</b>', styles['Normal']), Paragraph(f'<b>{_PT["sec_order"]}</b>', styles['Normal'])],
        [Paragraph(f'<b>{_PT["name"]}:</b> {nombre}', styles['Normal']), Paragraph(f'<b>{_PT["order_no"]}:</b> {pid}', styles['Normal'])],
        [Paragraph(f'<b>{_PT["company"]}:</b> {empresa or "-"}', styles['Normal']), Paragraph(f'<b>{_PT["date"]}:</b> {fecha}', styles['Normal'])],
        [Paragraph(f'<b>{_PT["email"]}:</b> {email_c}', styles['Normal']), Paragraph(f'<b>{_PT["status"]}:</b> {estado}', styles['Normal'])],
        [Paragraph(f'<b>{_PT["phone"]}:</b> {telefono or "-"}', styles['Normal']), Paragraph(f'<b>{_PT["country"]}:</b> {pais or "-"}', styles['Normal'])],
        [Paragraph(f'<b>{_PT["incoterm"]}:</b> {tipo}' + (f' | {_PT["freight"]}: ${flete_usd_caja:.2f} USD/Kilo' if tipo=="CIF" and flete_usd_caja>0 else ''), styles['Normal']),
         Paragraph(f'<b>{_PT["destination"]}:</b> {destino if tipo=="CIF" and destino else _PT["fob_origin"]}<br/><font color="#888888" size="8">{_PT["shipping_from"]}</font>', styles['Normal'])],
    ]
    info_table = Table(info_data, colWidths=[9*cm, 8*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), AZUL_LIGHT),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,0), (-1,0), AZUL),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('PADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.4*cm))

    # --- Tabla de productos ---
    prod_title = Paragraph(f'<b>{_PT["product_detail"]}</b>', ParagraphStyle('ptitle', fontSize=10, textColor=AZUL, fontName='Helvetica-Bold', spaceBefore=6))
    story.append(prod_title)
    story.append(Spacer(1, 0.2*cm))

    prod_header = [_PT['th_code'], _PT['th_product'], _PT['th_boxes'], _PT['th_pallets'], _PT['th_price'], _PT['th_total']]
    prod_rows = [prod_header]
    for item in ped.get('productos', []):
        prod_rows.append([
            item.get('codigo',''),
            item.get('producto',''),
            str(item.get('cajas',0)),
            str(item.get('pallets',0)),
            f'${item.get("precio_usd",0):.2f}',
            f'${item.get("total",0):,.2f}',
        ])
    # Total row
    prod_rows.append(['', '', '', '', Paragraph(f'<b>{_PT["total_label"]}</b>', styles['Normal']), Paragraph(f'<b>${total_usd:,.2f} USD</b>', styles['Normal'])])

    col_widths = [2.2*cm, 5.8*cm, 1.8*cm, 1.8*cm, 2.4*cm, 3*cm]
    prod_table = Table(prod_rows, colWidths=col_widths, repeatRows=1)
    prod_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), AZUL),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (2,0), (-1,-1), 'CENTER'),
        ('ALIGN', (4,0), (-1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#DDDDDD')),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F0F4FF')]),
        ('BACKGROUND', (0,-1), (-1,-1), AZUL_LIGHT),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('LINEABOVE', (0,-1), (-1,-1), 1.5, AZUL),
    ]))
    story.append(prod_table)
    story.append(Spacer(1, 0.4*cm))

    # --- Notas ---
    if notas:
        story.append(Paragraph(f'<b>{_PT["notes_label"]}</b> {notas}', ParagraphStyle('notas', fontSize=9, textColor=GRIS, spaceBefore=4)))
        story.append(Spacer(1, 0.2*cm))

    # PATCH 24: Tabla de embalaje / packaging table
    story.append(Spacer(1, 0.4*cm))
    packing_title = Paragraph(f'<b>{_PT["packing_title"]}</b>', ParagraphStyle('ptitle2', fontSize=10, textColor=AZUL, fontName='Helvetica-Bold', spaceBefore=6))
    story.append(packing_title)
    story.append(Spacer(1, 0.2*cm))
    packing_header = [_PT['pk_product'], _PT['pk_group'], _PT['pk_pallets'], _PT['pk_boxes'], _PT['pk_kgbox'], _PT['pk_weight'], _PT['pk_bxpal']]
    packing_rows = [packing_header]
    _total_weight = 0
    _total_pal_pk = 0
    _total_caj_pk = 0
    for _pk_item in ped.get('productos', []):
        _pk_cajas = int(_pk_item.get('cajas', 0))
        _pk_pallets = float(_pk_item.get('pallets', 0))
        _pk_name = _pk_item.get('producto', '')
        # Get kg_caja and grupo from product/grupo data
        _pk_kg = 0
        _pk_grp = ''
        _pk_cxp_grp = 0
        _pk_tipo_caja = ''
        _pk_data_all = load_data()
        for _pp in _pk_data_all.get('products', []):
            if _pp.get('codigo','') == _pk_item.get('codigo',''):
                _pk_grp = _pp.get('grupo','')
                _pk_gi = _pk_data_all.get('config',{}).get('grupos',{}).get(_pk_grp,{})
                _pk_kg = float(_pk_gi.get('kg_caja', _pp.get('kg_caja', 0) or 0) if _pk_grp and isinstance(_pk_gi, dict) else (_pp.get('kg_caja', 0) or 0))
                _pk_cxp_grp = int(_pk_gi.get('cajas_pallet', _pp.get('cajas_pallet', 160) or 160) if _pk_grp and isinstance(_pk_gi, dict) else (_pp.get('cajas_pallet', 160) or 160))
                _pk_tipo_caja = _pk_gi.get('tipo_caja', '') if isinstance(_pk_gi, dict) else ''
                break
        _pk_weight = round(_pk_cajas * _pk_kg, 1) if _pk_kg else 0
        _pk_cxp = round(_pk_cajas / _pk_pallets, 0) if _pk_pallets > 0 else 0
        _total_weight += _pk_weight
        _total_pal_pk += _pk_pallets
        _total_caj_pk += _pk_cajas
        packing_rows.append([
            _pk_name,
            _pk_grp if _pk_grp else '—',
            f'{_pk_pallets:.1f}',
            str(_pk_cajas),
            f'{_pk_kg:.2f} kg' if _pk_kg else '—',
            f'{_pk_weight:,.0f} kg' if _pk_weight else '—',
            str(_pk_cxp_grp) if _pk_cxp_grp else '—',
        ])
    packing_rows.append([_PT['pk_total'], '', f'{_total_pal_pk:.1f}', str(_total_caj_pk), '', f'{_total_weight:,.0f} kg' if _total_weight else '—', ''])
    pk_col_widths = [4.5*cm, 1.5*cm, 1.8*cm, 1.8*cm, 2*cm, 2.8*cm, 1.8*cm]
    pk_table = Table(packing_rows, colWidths=pk_col_widths, repeatRows=1)
    pk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), AZUL),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DDEEFF')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BOX', (0, 0), (-1, -1), 0.5, AZUL),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#CCCCCC')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(pk_table)
    story.append(Spacer(1, 0.2*cm))
    pk_note = Paragraph(f'<i>{_PT["pk_note"]}</i>', ParagraphStyle('pknote', fontSize=7, textColor=GRIS))
    story.append(pk_note)
    story.append(Spacer(1, 0.3*cm))

    # --- Pie de pagina ---
    story.append(HRFlowable(width='100%', thickness=1, color=AZUL))
    story.append(Spacer(1, 0.2*cm))
    footer_style = ParagraphStyle('footer', fontSize=8, textColor=GRIS, alignment=TA_CENTER)
    story.append(Paragraph(_PT['footer_main'], footer_style))
    story.append(Paragraph(_PT['footer_sub'], ParagraphStyle('footer2', fontSize=7, textColor=colors.HexColor('#AAAAAA'), alignment=TA_CENTER)))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue(), 'application/pdf', '.pdf'


def build_catalog_pdf(data):
    """Genera un PDF del catálogo de productos con precios de referencia CIF Madrid."""
    buf = io.BytesIO()
    if not REPORTLAB_OK:
        html = '<h1>Catálogo Export Haret</h1><table border=1><tr><th>Producto</th><th>Código</th></tr>'
        for p in data.get('products', []):
            if not p.get('activo', True): continue
            html += f"<tr><td>{p.get('producto','')}</td><td>{p.get('codigo','')}</td></tr>"
        html += '</table>'
        return html.encode('utf-8'), 'text/html', '.html'
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []
    AZUL = colors.HexColor('#0c6e51')
    GRIS = colors.HexColor('#666666')
    header_style = ParagraphStyle('h', fontSize=20, textColor=AZUL, fontName='Helvetica-Bold', spaceAfter=4, alignment=TA_LEFT)
    sub_style = ParagraphStyle('s', fontSize=10, textColor=GRIS, fontName='Helvetica', spaceAfter=12, alignment=TA_LEFT)
    story.append(Paragraph('Catálogo de Productos', header_style))
    story.append(Paragraph(f'Export Haret — Frutas Exóticas Premium de Ecuador  |  Generado: {datetime.now().strftime("%d/%m/%Y")}', sub_style))
    story.append(Spacer(1, 0.4*cm))
    # Tabla productos
    prods = [p for p in data.get('products', []) if p.get('activo', True)]
    if not prods:
        story.append(Paragraph('Sin productos activos.', styles['Normal']))
    else:
        rows = [['Código','Producto','Grupo','Cajas/Pal','Kg/Caja']]
        grupos = data.get('config',{}).get('grupos',{})
        for p in prods:
            g = p.get('grupo','')
            cxp = (grupos.get(g,{}) if isinstance(grupos.get(g,{}), dict) else {}).get('cajas_pallet', p.get('cajas_pallet', 160))
            kgc = (grupos.get(g,{}) if isinstance(grupos.get(g,{}), dict) else {}).get('kg_caja', p.get('kg_caja', 2.0))
            rows.append([p.get('codigo',''), p.get('producto',''), g, str(int(cxp) if cxp else ''), f'{float(kgc):.2f}' if kgc else ''])
        tbl = Table(rows, colWidths=[2.5*cm, 6*cm, 3*cm, 2.5*cm, 2.5*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), AZUL),
            ('TEXTCOLOR',(0,0),(-1,0), colors.white),
            ('FONTNAME',(0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1), 9),
            ('GRID',(0,0),(-1,-1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [colors.white, colors.HexColor('#f5f7fa')]),
            ('VALIGN',(0,0),(-1,-1), 'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1), 6),
            ('RIGHTPADDING',(0,0),(-1,-1), 6),
        ]))
        story.append(tbl)
    story.append(Spacer(1, 0.6*cm))
    story.append(Paragraph('<b>Contacto:</b> order@exportharet.com  |  +34 641 076 116', sub_style))
    story.append(Paragraph('Precios disponibles bajo solicitud. Sujeto a disponibilidad y condiciones de mercado.', ParagraphStyle('foot', fontSize=8, textColor=GRIS, alignment=TA_LEFT)))
    doc.build(story)
    buf.seek(0)
    return buf.getvalue(), 'application/pdf', '.pdf'

def get_precio_por_pallets(codigo, total_pallets, data, tipo_precio='CIF'):
    """Retorna el precio USD/caja para un producto segun el total de pallets del pedido.
    CIF: usa precios_plt (tabla de volumen del Excel, incluye flete ref Madrid), ajustado por destino.
    FOB: usa precio_fob_final directamente (col 12 Excel = FOB+Merma+Margen, sin flete).
         El precio FOB es fijo independiente del volumen.
    """
    pals = max(1, int(total_pallets))
    for p in data.get('products', []):
        if p.get('codigo') == codigo:
            if tipo_precio == 'FOB':
                # FOB: usar precio FOB Final del Excel (col 12) directamente
                fob_final = p.get('precio_fob_final', 0) or 0
                if fob_final and float(fob_final) > 0:
                    return round(float(fob_final), 4)
                # Fallback: si no hay precio_fob_final, calcular desde precio_compra + margen
                pc = float(p.get('precio_compra', 0) or 0)
                mg = float(p.get('margen_pct', 0.1) or 0.1)
                return round(pc * (1 + mg), 4) if pc > 0 else 0.0
            # CIF: usar tabla de volumen por pallets
            precios_plt = p.get('precios_plt', [])
            if precios_plt:
                idx = min(pals - 1, len(precios_plt) - 1)
                v = precios_plt[idx]
                if v and float(v) > 0:
                    return round(float(v), 4)
    # Fallback a calculo anterior si no hay tabla directa
    for p in data.get('products', []):
        if p.get('codigo') == codigo:
            pc = float(p.get('precio_compra', 0) or 0)
            mg = float(p.get('margen_pct', 0.1) or 0.1)
            return round(pc * (1 + mg), 4)
    return 0.0


def get_precio_cif_por_pallets(codigo, total_pallets, destino, data):
    """Precio CIF para un destino. precios_plt incluye el flete de referencia (Madrid);
    se ajusta al flete del destino igual que el panel admin:
        precio_destino = (precio_base - flete_ref + flete_destino) * (1 + margen_mercado)
    Así cada destino cobra su flete real Y su margen de mercado (p. ej. UK > España).
    """
    base = get_precio_por_pallets(codigo, total_pallets, data)  # incluye flete ref Madrid
    if base <= 0:
        return base
    cfg = data.get('config', {})
    flete_ref = float(cfg.get('flete_ref', 2.35) or 2.35)
    dest_val = cfg.get('destinos', {}).get(destino, flete_ref)
    if isinstance(dest_val, dict):
        dest_flete = float(dest_val.get('factor', flete_ref) or flete_ref)
    elif isinstance(dest_val, (int, float)):
        dest_flete = float(dest_val)
    else:
        dest_flete = flete_ref
    precio = base - flete_ref + dest_flete
    # Margen de mercado por destino (%): UK puede llevar más que España, etc.
    try:
        _margen_pct = float(cfg.get('destinos_margen', {}).get(destino, 0) or 0)
    except (TypeError, ValueError):
        _margen_pct = 0.0
    if _margen_pct:
        precio = precio * (1 + _margen_pct / 100.0)
    return round(precio, 4)


def get_precio_con_volumen(codigo, destino, tipo_precio, data, pallets):
    """Retorna precio usando la tabla directa por pallets.
    FOB: precio en origen (sin flete). CIF: precio incluye flete al destino.
    """
    if tipo_precio == 'CIF' and destino:
        return get_precio_cif_por_pallets(codigo, pallets, destino, data)
    else:
        # FOB: precio base sin flete (precios_plt - flete_ref)
        return get_precio_por_pallets(codigo, pallets, data, tipo_precio='FOB')


@st.cache_data(ttl=3600)
def get_exchange_rates():
    """Obtiene cotizaciones en tiempo real desde exchangerate-api.com (free tier)."""
    try:
        import urllib.request
        url = "https://open.er-api.com/v6/latest/USD"
        with urllib.request.urlopen(url, timeout=5) as r:
            rates_data = __import__("json").loads(r.read())
        if rates_data.get("result") == "success":
            return rates_data.get("rates", {})
    except Exception:
        pass
    return {"USD":1,"EUR":0.92,"GBP":0.79,"CHF":0.89,"AED":3.67,"CAD":1.36,"MXN":17.5,"BRL":4.97,"COP":3950}

@st.cache_data(ttl=3600)
def get_exchange_rates_meta():
    """Cotizaciones con metadatos: timestamp, fuente, live flag."""
    try:
        import urllib.request
        url = "https://open.er-api.com/v6/latest/USD"
        with urllib.request.urlopen(url, timeout=5) as r:
            rd = __import__("json").loads(r.read())
        if rd.get("result") == "success":
            ts_str = rd.get("time_last_update_utc", "")
            try:
                from datetime import datetime as _dt
                ts_dt = _dt.strptime(ts_str, "%a, %d %b %Y %H:%M:%S +0000")
                ts_fmt = ts_dt.strftime("%d/%m/%Y %H:%M UTC")
            except Exception:
                ts_fmt = ts_str[:16] if ts_str else datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC")
            return {"rates": rd.get("rates", {}), "ts": ts_fmt, "source": "open.er-api.com", "live": True}
    except Exception:
        pass
    ts_fmt = datetime.utcnow().strftime("%d/%m/%Y %H:%M UTC") + " (aprox.)"
    return {"rates": {"USD":1,"EUR":0.92,"GBP":0.79,"CHF":0.89,"AED":3.67,"CAD":1.36,"MXN":17.5,"BRL":4.97,"COP":3950}, "ts": ts_fmt, "source": "referencia", "live": False}

def convertir_precio(precio_usd, moneda):
    """Convierte precio USD a la moneda destino."""
    if moneda == "USD":
        return precio_usd
    rates = get_exchange_rates()
    rate = rates.get(moneda, 1)
    return round(precio_usd * rate, 4)

def log_email(destinatario, asunto, tipo_email):
    el = load_email_log()
    el.append({'id': f'EMAIL-{len(el)+1:05d}', 'destinatario': destinatario, 'asunto': asunto, 'tipo': tipo_email, 'fecha': datetime.now().isoformat(), 'estado': 'simulado'})
    save_email_log(el)

# ─── PORTAL PÚBLICO DE PEDIDOS ───────────────────────────────────────────────
def send_order_email(ped):
    """Envia el pedido por email a order@exportharet.com.
    Requiere en .streamlit/secrets.toml:
    [email]
    smtp_host = 'smtp.gmail.com'
    smtp_port = 587
    smtp_user = 'tu@gmail.com'
    smtp_pass = 'app_password'
    from_addr = 'tu@gmail.com'
    """
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    DEST = 'order@exportharet.com'
    pid = ped.get('id','')
    nombre = ped.get('client_name','')
    email_c = ped.get('client_email','')
    empresa = ped.get('empresa','')
    telefono = ped.get('telefono','')
    pais = ped.get('pais','')
    tipo = ped.get('tipo_precio','FOB')
    destino = ped.get('destino','')
    total_usd = ped.get('total_usd',0)
    fecha = ped.get('fecha','')[:10]
    notas = ped.get('notas','')
    rows_html = ''
    for item in ped.get('productos', []):
        rows_html += (f'<tr><td style="padding:6px 10px;border:1px solid #ddd">{_esc(item.get("codigo",""))}</td>'
                      f'<td style="padding:6px 10px;border:1px solid #ddd">{_esc(item.get("producto",""))}</td>'
                      f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{item.get("cajas",0)}</td>'
                      f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:center">{item.get("pallets",0)}</td>'
                      f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:right">${item.get("precio_usd",0):.2f}</td>'
                      f'<td style="padding:6px 10px;border:1px solid #ddd;text-align:right;font-weight:bold">${item.get("total",0):,.2f}</td></tr>')
    dest_str = f'{_esc(tipo)} \u2192 {_esc(destino)}' if tipo == 'CIF' and destino else _esc(tipo)
    # Escapados para HTML (evita XSS contra el staff que lee el email)
    nombre_h, empresa_h, email_h = _esc(nombre), _esc(empresa), _esc(email_c)
    telefono_h, pais_h = _esc(telefono), _esc(pais)
    _notas_html = f'<p><b>Notas:</b> {_esc(notas)}</p>' if notas else ''
    html = (f'<html><body style="font-family:Arial,sans-serif;color:#333">'
            f'<div style="background:#0c6e51;padding:16px 24px;border-radius:8px">'
            f'<h2 style="color:white;margin:0">🚀 Export Haret \u2014 Nueva Orden Recibida</h2>'
            f'</div><div style="padding:16px 0">'
            f'<table style="width:100%;border-collapse:collapse;font-size:14px">'
            f'<tr><td style="padding:6px"><b>N\u00ba Pedido:</b></td><td>{pid}</td>'
            f'<td style="padding:6px"><b>Fecha:</b></td><td>{fecha}</td></tr>'
            f'<tr><td style="padding:6px"><b>Cliente:</b></td><td>{nombre_h}</td>'
            f'<td style="padding:6px"><b>Empresa:</b></td><td>{empresa_h or "-"}</td></tr>'
            f'<tr><td style="padding:6px"><b>Email:</b></td><td>{email_h}</td>'
            f'<td style="padding:6px"><b>Tel\u00e9fono:</b></td><td>{telefono_h or "-"}</td></tr>'
            f'<tr><td style="padding:6px"><b>Pa\u00eds:</b></td><td>{pais_h or "-"}</td>'
            f'<td style="padding:6px"><b>Destino:</b></td><td>{dest_str}</td></tr>'
            f'</table>'
            f'<h3 style="color:#0c6e51;border-bottom:2px solid #0c6e51;padding-bottom:6px">Productos</h3>'
            f'<table style="width:100%;border-collapse:collapse;font-size:13px">'
            f'<thead><tr style="background:#0c6e51;color:white">'
            f'<th style="padding:8px">C\u00f3digo</th><th style="padding:8px">Producto</th>'
            f'<th style="padding:8px">Cajas</th><th style="padding:8px">Pallets</th>'
            f'<th style="padding:8px">Precio/caja</th><th style="padding:8px">Total</th>'
            f'</tr></thead><tbody>{rows_html}</tbody></table>'
            f'<p style="text-align:right;font-size:16px;font-weight:bold;color:#0c6e51">'
            f'TOTAL: ${total_usd:,.2f} USD</p>'
            f'{_notas_html}'
            f'</div><p style="color:#888;font-size:11px">Export Haret \u00a9 2026 | order@exportharet.com</p>'
            f'</body></html>')
    subject = f'📦 Nuevo Pedido {pid} — {nombre} ({empresa or email_c}) | ${total_usd:,.2f} USD'
    sent = False
    error_msg = ''
    try:
        cfg = {}
        try:
            cfg = st.secrets.get('email', {})
        except Exception:
            cfg = {}
        smtp_host = cfg.get('smtp_host', '')
        smtp_port = int(cfg.get('smtp_port', 587))
        smtp_user = cfg.get('smtp_user', '')
        smtp_pass = cfg.get('smtp_pass', '')
        from_addr = cfg.get('from_addr', smtp_user) or smtp_user
        if smtp_host and smtp_user and smtp_pass:
            msg = MIMEMultipart('mixed')
            msg['Subject'] = subject
            msg['From'] = from_addr
            msg['To'] = DEST
            msg['Reply-To'] = email_c
            msg.attach(MIMEText(html, 'html', 'utf-8'))
            # PATCH 23: Attach PDF to email
            try:
                from email.mime.base import MIMEBase
                from email import encoders as _enc
                _pdf_bytes, _pdf_mime, _pdf_ext = build_order_pdf(ped)
                _pdf_part = MIMEBase('application', 'octet-stream')
                _pdf_part.set_payload(_pdf_bytes)
                _enc.encode_base64(_pdf_part)
                _pdf_part.add_header('Content-Disposition', f'attachment; filename="{pid}{_pdf_ext}"')
                msg.attach(_pdf_part)
            except Exception as _pe2:
                log_email(DEST, subject, f'pdf_attach_error:{str(_pe2)[:100]}')
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls()
                server.ehlo()
                server.login(smtp_user, smtp_pass)
                server.sendmail(from_addr, [DEST], msg.as_string())
            log_email(DEST, subject, 'smtp_enviado')
            sent = True
        else:
            _missing = []
            if not smtp_host: _missing.append('smtp_host')
            if not smtp_user: _missing.append('smtp_user')
            if not smtp_pass: _missing.append('smtp_pass')
            error_msg = f'SMTP no configurado. Faltan: {", ".join(_missing)}. Ir a Streamlit Cloud → App settings → Secrets y agregar [email] smtp_host / smtp_user / smtp_pass / smtp_port / from_addr'
            log_email(DEST, subject, f'smtp_sin_config: {error_msg}')
    except Exception as e:
        error_msg = str(e)[:200]
        log_email(DEST, subject, f'smtp_error:{error_msg}')
    # Siempre guardar en pending_emails.json para auditoria
    try:
        _pf = 'pending_emails.json'
        _pe = _load(_pf, [])
        _pe.append({'id': pid, 'fecha': datetime.now().isoformat(), 'destinatario': DEST,
                    'asunto': subject, 'cliente': nombre, 'email_cliente': email_c,
                    'total': total_usd, 'sent': sent, 'error': error_msg})
        _save(_pf, _pe)
    except Exception:
        pass
    return sent  # #6: el llamador puede reflejar si el email salió de verdad


# ── Aviso de estado al CLIENTE (email automático + link WhatsApp de un clic) ──
ESTADO_MSG_CLIENTE = {
    'Recibido':   ('Pedido recibido',      'hemos recibido tu pedido {pid} y lo estamos revisando. Te confirmamos en breve.'),
    'Confirmado': ('Pedido confirmado',    '¡buenas noticias! Tu pedido {pid} ha sido CONFIRMADO. Empezamos a prepararlo.'),
    'Preparando': ('Preparando tu pedido', 'tu pedido {pid} está EN PREPARACIÓN. Te avisamos en cuanto salga.'),
    'Enviado':    ('Pedido enviado',       'tu pedido {pid} ha sido ENVIADO. En breve recibirás los detalles de la logística.'),
    'Entregado':  ('Pedido entregado',     'tu pedido {pid} ha sido ENTREGADO. ¡Gracias por tu confianza!'),
    'Cancelado':  ('Pedido cancelado',     'tu pedido {pid} ha sido CANCELADO. Si tienes cualquier duda, contáctanos.'),
}

def _client_status_text(ped, estado):
    """Devuelve (titulo, cuerpo) del aviso de estado para el cliente."""
    pid = (ped.get('id', '') or '').upper()
    nombre = (ped.get('client_name', '') or '').strip()
    _t, _b = ESTADO_MSG_CLIENTE.get(estado, ('Actualización de pedido',
                                             'el estado de tu pedido {pid} es ahora: ' + str(estado) + '.'))
    saludo = f'Hola {nombre}, ' if nombre else 'Hola, '
    return _t, saludo + _b.format(pid=pid) + ' — Export Haret'

def _client_wa_link(ped, estado):
    """Link wa.me prefilled al teléfono del cliente para avisarle en un clic."""
    import urllib.parse as _up
    tel = ''.join(ch for ch in str(ped.get('telefono', '') or '') if ch.isdigit())
    _t, _body = _client_status_text(ped, estado)
    _q = _up.quote(_body)
    return f'https://wa.me/{tel}?text={_q}' if tel else f'https://wa.me/?text={_q}'

def send_status_email(ped, estado):
    """Avisa al CLIENTE por email del nuevo estado de su pedido. Solo envía si hay
    SMTP en secrets; si no, devuelve False sin romper ni bloquear el cambio."""
    import smtplib
    from email.mime.text import MIMEText
    to_addr = (ped.get('client_email', '') or '').strip()
    if not to_addr:
        return False
    pid = (ped.get('id', '') or '').upper()
    _title, _body = _client_status_text(ped, estado)
    subject = f'Export Haret — {_title} ({pid})'
    sent = False
    try:
        try:
            cfg = st.secrets.get('email', {})
        except Exception:
            cfg = {}
        smtp_host = cfg.get('smtp_host', ''); smtp_port = int(cfg.get('smtp_port', 587))
        smtp_user = cfg.get('smtp_user', ''); smtp_pass = cfg.get('smtp_pass', '')
        from_addr = cfg.get('from_addr', smtp_user) or smtp_user
        if smtp_host and smtp_user and smtp_pass:
            _ic = ESTADO_ICONS.get(estado, '📦')
            html = (f'<div style="font-family:Arial,sans-serif;color:#19231D">'
                    f'<div style="background:#0c6e51;padding:16px 22px;border-radius:10px">'
                    f'<h2 style="margin:0;color:#fff">{_ic} {_esc(_title)}</h2></div>'
                    f'<p style="font-size:15px;line-height:1.6;margin:18px 2px">{_esc(_body)}</p>'
                    f'<p style="font-size:13px;color:#666;margin:2px">Pedido <b>{_esc(pid)}</b> · '
                    f'Dudas: order@exportharet.com</p></div>')
            msg = MIMEText(html, 'html', 'utf-8')
            msg['Subject'] = subject; msg['From'] = from_addr; msg['To'] = to_addr
            msg['Reply-To'] = 'order@exportharet.com'
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo(); server.starttls(); server.ehlo()
                server.login(smtp_user, smtp_pass)
                server.sendmail(from_addr, [to_addr], msg.as_string())
            log_email(to_addr, subject, 'status_smtp_enviado')
            sent = True
        else:
            log_email(to_addr, subject, 'status_smtp_sin_config')
    except Exception as e:
        log_email(to_addr, subject, f'status_smtp_error:{str(e)[:120]}')
    return sent

def _eh_seccion(raw, num):
    """Cabecera de sección premium: chip numerado con degradado + título.
    Sustituye los '### 1️⃣ Título' básicos."""
    t = (raw or '').replace('#', '').strip()
    i = 0
    while i < len(t) and not t[i].isalpha():
        i += 1
    t = t[i:].strip()
    st.markdown(
        f'<div class="eh-sec"><span class="eh-sec-num">{num}</span>'
        f'<span class="eh-sec-title">{t}</span></div>', unsafe_allow_html=True)


# ── Tienda visual: ilustración de línea + color de fondo por fruta ──────────────
_ART_VERDE = '#0c6e51'
_ART_NARA = '#10a37a'
# (svg_inner, stroke, bg) — SVG de línea simple por familia de fruta
_FRUIT_ART = {
    'pitahaya': ('<path d="M30 18C40 18 47 28 47 38C47 48 39 53 30 53C21 53 13 48 13 38C13 28 20 18 30 18Z"/><path d="M30 18C30 12 27 8 23 7M30 18C30 12 33 8 37 7M30 18C25 15 21 15 17 17M30 18C35 15 39 15 43 17"/>', _ART_NARA, '#FFF3DE'),
    'dragon':   ('<path d="M30 18C40 18 47 28 47 38C47 48 39 53 30 53C21 53 13 48 13 38C13 28 20 18 30 18Z"/><path d="M30 18C30 12 27 8 23 7M30 18C35 15 39 15 43 17"/>', _ART_NARA, '#FFF3DE'),
    'granadilla':('<circle cx="30" cy="34" r="18"/><path d="M30 16V9M30 9C26 9 24 7 24 5M30 9C34 9 36 7 36 5"/>', _ART_VERDE, '#EAF6E0'),
    'maracu':   ('<ellipse cx="30" cy="32" rx="17" ry="20"/><ellipse cx="30" cy="32" rx="9" ry="11"/>', _ART_NARA, '#FDEEDD'),
    'passion':  ('<ellipse cx="30" cy="32" rx="17" ry="20"/><ellipse cx="30" cy="32" rx="9" ry="11"/>', _ART_NARA, '#FDEEDD'),
    'babaco':   ('<path d="M30 12 44 24 40 50 20 50 16 24Z"/><path d="M30 12V7"/>', _ART_VERDE, '#FBF7DE'),
    'cacao':    ('<circle cx="30" cy="33" r="17"/><path d="M22 33 Q30 24 38 33 Q30 42 22 33Z"/>', _ART_NARA, '#F3E9DF'),
    'physalis': ('<circle cx="30" cy="34" r="14"/><path d="M30 20 18 12M30 20 42 12M30 20 30 9"/>', _ART_NARA, '#FBEFD6'),
    'tomate':   ('<ellipse cx="30" cy="33" rx="13" ry="18"/><path d="M30 15V8"/>', _ART_NARA, '#FBEAE0'),
    'tamarillo':('<ellipse cx="30" cy="33" rx="13" ry="18"/><path d="M30 15V8"/>', _ART_NARA, '#FBEAE0'),
    'pepino':   ('<path d="M22 22C22 16 38 16 38 22 38 40 34 50 30 50 26 50 22 40 22 22Z"/>', _ART_VERDE, '#EAF6E0'),
    'cucumber': ('<path d="M22 22C22 16 38 16 38 22 38 40 34 50 30 50 26 50 22 40 22 22Z"/>', _ART_VERDE, '#EAF6E0'),
    'lulo':     ('<circle cx="30" cy="34" r="16"/><path d="M30 18 22 11M30 18 38 11"/>', _ART_NARA, '#FBEFD6'),
    'naranjilla':('<circle cx="30" cy="34" r="16"/><path d="M30 18 22 11M30 18 38 11"/>', _ART_NARA, '#FBEFD6'),
    'melon':    ('<circle cx="30" cy="34" r="17"/><path d="M16 28 Q30 38 44 28M16 36 Q30 46 44 36"/>', _ART_VERDE, '#EAF6E0'),
    'taxo':     ('<ellipse cx="30" cy="33" rx="13" ry="19"/>', _ART_NARA, '#FDEEDD'),
}
_FRUIT_ART_DEFAULT = ('<circle cx="30" cy="33" r="17"/><path d="M30 16V9"/>', _ART_VERDE, '#EAF6E0')


def _fruit_art(nombre):
    """Devuelve (svg_html, bg_color) para la tarjeta de la fruta según su nombre."""
    n = (nombre or '').lower()
    for k, (inner, stroke, bg) in _FRUIT_ART.items():
        if k in n:
            svg = (f'<svg viewBox="0 0 60 60" fill="none" stroke="{stroke}" stroke-width="2.6" '
                   f'stroke-linecap="round" stroke-linejoin="round">{inner}</svg>')
            return svg, bg
    inner, stroke, bg = _FRUIT_ART_DEFAULT
    return (f'<svg viewBox="0 0 60 60" fill="none" stroke="{stroke}" stroke-width="2.6" '
            f'stroke-linecap="round" stroke-linejoin="round">{inner}</svg>'), bg


def render_portal_pedido():
    """Página pública para que los clientes hagan pedidos. No requiere login de staff."""
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)
    if theme:
        theme.aplicar()
    data = load_data()
    prods = [p for p in data.get('products', []) if p.get('activo', True)]
    dests = data.get('config', {}).get('destinos', {})

    # Header
    import os as _os2
    if _os2.path.exists('logo.png'):
        from PIL import Image as _Img2
        _logo2 = _Img2.open('logo.png')
        _ph1, _ph2, _ph3 = st.columns([1, 2, 1])
        with _ph2: st.image(_logo2, width=190)
        _tagline = ('Premium exotic fruits · Ecuador' if st.session_state.get('portal_lang','es') == 'en'
                    else 'Frutas exóticas premium · Ecuador')
        st.markdown(
            '<div style="text-align:center;margin:-2px 0 18px">'
            f'<span style="color:#8a978f;font-size:.74rem;font-weight:600;letter-spacing:2.4px;'
            f'text-transform:uppercase">{_tagline}</span></div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div style="background:linear-gradient(135deg,#0c6e51,#0a5d44,#10a37a);padding:20px 30px;border-radius:12px;margin-bottom:24px;text-align:center"><h1 style="color:white;margin:0;font-size:1.8em">🚀 Export Haret</h1><p style="color:rgba(255,255,255,0.85);margin:4px 0 0">{LANG_TEXTS[st.session_state.get("portal_lang","es")]["header_subtitle"].split(" | ")[0]}</p></div>',unsafe_allow_html=True)
    # Init lang EARLY so _T is available for error messages
    if 'portal_lang' not in st.session_state:
        st.session_state['portal_lang'] = 'es'

    if not prods:
        _T_early = LANG_TEXTS[st.session_state.get('portal_lang','es')]
        st.warning(_T_early['no_catalog'])
        return

    portal_clients = load_portal_clients()

    # Init session state for portal
    # -- Selector de idioma (con rerun al cambiar) --
    if 'portal_lang' not in st.session_state:
        st.session_state['portal_lang'] = 'es'
    # PATCH 5: Language selector as flag buttons (with label)
    _cur_lang = st.session_state.get('portal_lang', 'es')
    # Toggle de idioma compacto y elegante (píldoras pequeñas, sin caja tosca)
    st.markdown('''<style>
      .st-key-btn_lang_es, .st-key-btn_lang_en { display:inline-block; }
      .st-key-btn_lang_es button, .st-key-btn_lang_en button {
        min-height:34px !important; height:34px !important; min-width:42px !important;
        padding:0 8px !important; border-radius:9px !important; font-size:1.05rem !important;
        box-shadow:none !important; transition:all .15s ease; }
      .st-key-btn_lang_es button:hover, .st-key-btn_lang_en button:hover { transform:none !important; }
      .st-key-btn_lang_es button[kind="secondary"], .st-key-btn_lang_en button[kind="secondary"] {
        background:#fff !important; border-color:#e6ece8 !important; opacity:.55; }
      .st-key-btn_lang_es button[kind="primary"], .st-key-btn_lang_en button[kind="primary"] {
        background:#eef5f0 !important; border:1px solid #dce1e8 !important; opacity:1; }
      /* Móvil: mantener etiqueta + banderas en UNA fila (no apiladas) */
      @media (max-width:768px){
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es){
          flex-direction:row !important; flex-wrap:nowrap !important; align-items:center !important;
          justify-content:flex-end !important; gap:7px !important; }
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es) > div{
          flex:0 0 auto !important; width:auto !important; min-width:0 !important; }
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es) > div:first-child{ flex:1 1 auto !important; }
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es) > div:first-child > div{ padding-top:6px !important; }
      }
    </style>''', unsafe_allow_html=True)
    _lbtn_c1, _lbtn_c2, _lbtn_c3 = st.columns([9, 1, 1])
    with _lbtn_c1:
        st.markdown("<div style='text-align:right;padding-top:9px;color:#aab5ad;font-size:0.72rem;letter-spacing:.4px;text-transform:uppercase'>Idioma · Language</div>", unsafe_allow_html=True)
    with _lbtn_c2:
        _es_type = 'primary' if _cur_lang == 'es' else 'secondary'
        if st.button('🇪🇸', key='btn_lang_es', help='Español', use_container_width=False, type=_es_type):
            if _cur_lang != 'es':
                st.session_state['portal_lang'] = 'es'
                st.rerun()
    with _lbtn_c3:
        _en_type = 'primary' if _cur_lang == 'en' else 'secondary'
        if st.button('🇬🇧', key='btn_lang_en', help='English', use_container_width=False, type=_en_type):
            if _cur_lang != 'en':
                st.session_state['portal_lang'] = 'en'
                st.rerun()
    _T = LANG_TEXTS[st.session_state.portal_lang]

    for k, v in [('portal_email',''),('portal_registered',False),('portal_client_data',{}),('portal_carrito',[])]:
        if k not in st.session_state: st.session_state[k] = v

    # Pre-fill email from last confirmed order (persistence #17)
    if not st.session_state.get('portal_email') and st.session_state.get('_portal_last_confirmed_email'):
        st.session_state['portal_email'] = st.session_state['_portal_last_confirmed_email']
    # ── PASO 1: Identificación del cliente ────────────────────────────────────
    # Progress bar - steps indicator
    _step1_done = bool(st.session_state.get('portal_email',''))
    _step2_done = bool(st.session_state.get('portal_carrito',[]))
    # Stepper "Pro": barra fina blanca con estados (hecho / actual / pendiente)
    _stp_state = (['done', 'done', 'on', 'off'] if _step1_done else ['on', 'off', 'off', 'off'])
    _stp_lbls = [_T["progress_step1"], _T["progress_step2"], _T["progress_step3"], _T["progress_step4"]]
    _seg = ''
    for _i, (_lbl, _stt) in enumerate(zip(_stp_lbls, _stp_state), 1):
        _nbg = '#0c6e51' if _stt in ('on', 'done') else '#e7ece8'
        _nc = '#fff' if _stt in ('on', 'done') else '#737d77'
        _txt = '#16201b' if _stt != 'off' else '#a3aaa3'
        _bgc = '#eef6f2' if _stt == 'on' else 'transparent'
        _ico = '✓' if _stt == 'done' else str(_i)
        _br = ';border-right:1px solid #ebefec' if _i < 4 else ''
        _seg += (f'<div style="flex:1;display:flex;align-items:center;justify-content:center;gap:7px;'
                 f'padding:11px 6px;background:{_bgc};font-size:.82rem;font-weight:600;color:{_txt}{_br}">'
                 f'<span style="width:19px;height:19px;border-radius:50%;background:{_nbg};color:{_nc};'
                 f'font-size:11px;font-weight:800;display:inline-flex;align-items:center;justify-content:center">{_ico}</span>'
                 f'{_lbl}</div>')
    st.markdown(f'<div style="display:flex;border:1px solid #ebefec;border-radius:12px;overflow:hidden;'
                f'background:#fff;margin:0 0 18px">{_seg}</div>', unsafe_allow_html=True)

    # Banner comercial: invita al cliente a simular su pedido y ver precios por volumen
    _promo = {
        'es': ('Pedido mayorista', 'Realiza tu pedido',
               'Indica las cantidades — el precio por caja baja con el volumen. Sin compromiso; te confirmamos en 24 h.'),
        'en': ('Wholesale order', 'Place your order',
               'Set the quantities — the price per box drops with volume. No commitment; we confirm within 24 h.'),
    }.get(st.session_state.get('portal_lang', 'es'),
          ('Pedido mayorista', 'Realiza tu pedido',
           'Indica las cantidades — el precio por caja baja con el volumen.'))
    # Ahorro MÁXIMO por caja a mayor volumen (1 pallet -> precio más barato de la tabla)
    _ahmax = []
    for _pp in data.get('products', []):
        _plt = _pp.get('precios_plt') or []
        _vals = [v for v in _plt if v]
        if len(_plt) >= 4 and _plt[0] and _vals and _plt[0] > min(_vals):
            _ahmax.append(_plt[0] - min(_vals))
    _ahorro_max = round(sum(_ahmax) / len(_ahmax), 2) if _ahmax else 0
    _en = st.session_state.get('portal_lang', 'es') == 'en'
    # Chips de confianza (sobrios, premium) — sustituyen la caja naranja ruidosa.
    # Uno de ellos resalta el ahorro real por volumen.
    def _chip(_txt, _accent=False):
        _bd = '#dce1e8' if not _accent else '#e7d4bf'
        _bg = '#eef6f2' if not _accent else '#eef6f2'
        _cc = '#0c6e51' if not _accent else '#10a37a'
        _tc = '#3c4b42' if not _accent else '#8a4e22'
        return (f'<span style="display:inline-flex;align-items:center;gap:6px;background:{_bg};'
                f'border:1px solid {_bd};border-radius:999px;padding:6px 13px;font-size:.82rem;'
                f'font-weight:600;color:{_tc};white-space:nowrap">'
                f'<span style="color:{_cc};font-weight:800">✓</span>{_esc(_txt)}</span>')
    _chips = [
        _chip('Confirmación en 24 h' if not _en else 'Confirmed in 24 h'),
        _chip('Sin compromiso' if not _en else 'No commitment'),
    ]
    if _ahorro_max > 0:
        _sv = (f'Hasta −${_ahorro_max:.2f}/caja por volumen' if not _en
               else f'Up to −${_ahorro_max:.2f}/box by volume')
        _chips.append(_chip(_sv, _accent=True))
    _strip = ('<div style="margin-top:15px;display:flex;flex-wrap:wrap;gap:8px">'
              + ''.join(_chips) + '</div>')
    st.markdown(
        '<div style="margin:8px 0 22px">'
        '<div style="display:flex;align-items:center;gap:9px;margin-bottom:7px">'
        '<span style="width:26px;height:2px;background:#0c6e51;border-radius:2px"></span>'
        f'<span style="font-size:.72rem;letter-spacing:1.8px;text-transform:uppercase;'
        f'color:#0c6e51;font-weight:700">{_promo[0]}</span></div>'
        f'<div style="font-weight:800;color:#14201a;font-size:2rem;letter-spacing:-.8px;'
        f'margin:0 0 6px;line-height:1.05">{_promo[1]}</div>'
        f'<div style="color:#65726b;font-size:1rem;line-height:1.55;max-width:560px">{_promo[2]}</div>'
        f'{_strip}'
        '</div>', unsafe_allow_html=True)

    # ── WhatsApp flotante de dudas (siempre visible) — reduce abandono ──
    import urllib.parse as _upw
    _wa_help_txt = ('Hi, I have a question about Export Haret.' if _en
                    else 'Hola, tengo una consulta sobre Export Haret.')
    _wa_help_url = 'https://wa.me/34641076116?text=' + _upw.quote(_wa_help_txt)
    st.markdown(
        '<style>@media (max-width:768px){.eh-wa-fab{width:46px !important;height:46px !important;'
        'right:12px !important;bottom:78px !important}.eh-wa-fab svg{width:25px !important;height:25px !important}}</style>'
        f'<a class="eh-wa-fab" href="{_wa_help_url}" target="_blank" rel="noopener" '
        'aria-label="WhatsApp" title="¿Dudas? Escríbenos por WhatsApp" '
        'style="position:fixed;right:18px;bottom:84px;z-index:70;width:54px;height:54px;'
        'border-radius:50%;background:#25D366;display:flex;align-items:center;justify-content:center;'
        'box-shadow:0 6px 18px rgba(0,0,0,.22);text-decoration:none">'
        '<svg width="30" height="30" viewBox="0 0 32 32" aria-hidden="true"><path fill="#fff" '
        'd="M16 .5C7.4.5.5 7.4.5 16c0 2.8.7 5.4 2 7.8L.5 31.5l7.9-2c2.3 1.2 4.9 1.9 7.6 1.9 '
        '8.6 0 15.5-6.9 15.5-15.5S24.6.5 16 .5zm0 28.3c-2.4 0-4.7-.6-6.7-1.8l-.5-.3-4.7 1.2 '
        '1.3-4.6-.3-.5c-1.3-2.1-2-4.5-2-7 0-7.2 5.9-13.1 13.1-13.1S29.1 8.8 29.1 16 23.2 28.8 16 28.8zm7.2-9.8c-.4-.2-2.3-1.1-2.7-1.3-.4-.1-.6-.2-.9.2-.3.4-1 1.3-1.2 1.5-.2.2-.4.3-.8.1-.4-.2-1.6-.6-3.1-1.9-1.1-1-1.9-2.3-2.1-2.7-.2-.4 0-.6.2-.8.2-.2.4-.4.5-.7.2-.2.2-.4.4-.6.1-.3 0-.5 0-.7-.1-.2-.9-2.1-1.2-2.9-.3-.8-.6-.7-.9-.7h-.7c-.2 0-.6.1-.9.4-.3.4-1.2 1.2-1.2 2.9 0 1.7 1.2 3.4 1.4 3.6.2.2 2.5 3.8 6 5.3.8.4 1.5.6 2 .7.8.3 1.6.2 2.2.1.7-.1 2.3-.9 2.6-1.8.3-.9.3-1.6.2-1.8-.1-.1-.3-.2-.7-.4z"/></svg></a>',
        unsafe_allow_html=True)

    _eh_seccion(_T['step1'], 1)
    # Email form with explicit "Acceder" button
    _eml_benef = ('We use it to save your order and follow up — no spam, no commitment.'
                  if st.session_state.get('portal_lang') == 'en'
                  else 'Lo usamos para guardar tu pedido y darte seguimiento — sin spam ni compromiso.')
    with st.form('portal_email_form', clear_on_submit=False):
        _email_form_raw = st.text_input(_T['email_label'], placeholder=_T['email_ph'], key='portal_email_input', value=st.session_state.portal_email)
        st.caption('🔒 ' + _eml_benef)
        _acceder_clicked = st.form_submit_button(_T.get('btn_acceder', '🔓 Acceder'), type='primary', use_container_width=True)
    if _acceder_clicked:
        import re as _re_eml
        _eml_trim = (_email_form_raw or '').strip()
        # #7 Validación amable: si el formato no es válido, avisamos junto al campo
        # (borde rojo + mensaje) en vez de dejar pasar un correo erróneo.
        if _eml_trim and not _re_eml.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', _eml_trim):
            st.session_state['_email_invalid'] = True
        else:
            st.session_state.pop('_email_invalid', None)
            if _eml_trim and _eml_trim != st.session_state.portal_email:
                st.session_state.portal_email = _eml_trim
                st.rerun()
            elif _eml_trim:
                st.session_state.portal_email = _eml_trim
    if st.session_state.get('_email_invalid'):
        st.markdown('<style>.st-key-portal_email_input input{border-color:#dc3545 !important;'
                    'box-shadow:0 0 0 3px rgba(220,53,69,.12) !important}</style>', unsafe_allow_html=True)
        st.markdown('<div style="color:#dc3545;font-size:.85rem;margin:-6px 0 4px">'
                    + _esc(_T.get('err_email_friendly', 'Revisa el correo: parece que falta la @ o el dominio (ej. nombre@empresa.com).'))
                    + '</div>', unsafe_allow_html=True)
    # Email SIEMPRE normalizado (minúsculas + sin espacios): es la clave única del cliente.
    # Evita fichas duplicadas tipo "Demo@x" vs "demo@x" en el padrón y en la lista del admin.
    email_input = (st.session_state.portal_email or (_email_form_raw or '').strip()).strip().lower()

    client_data = {}
    is_registered = False
    show_register = False

    if email_input:
        st.session_state.portal_email = email_input
        _eml_lc = email_input.strip().lower()
        # Reconocer al cliente de forma robusta: 1) padrón (case-insensitive),
        # 2) si no está, recuperar sus datos del último pedido. Así un cliente que
        # vuelve NO tiene que reescribir nada aunque solo hubiera pedido una vez.
        _rec = None
        for _k, _v in portal_clients.items():
            if (_k or '').strip().lower() == _eml_lc and isinstance(_v, dict):
                _rec = _v
                break
        if not _rec:
            _prev_ords = [p for p in load_pedidos()
                          if (p.get('client_email', '') or '').strip().lower() == _eml_lc]
            if _prev_ords:
                _last_ord = sorted(_prev_ords, key=lambda x: x.get('fecha', ''))[-1]
                _rec = {'nombre': _last_ord.get('client_name', ''), 'empresa': _last_ord.get('empresa', ''),
                        'telefono': _last_ord.get('telefono', ''), 'pais': _last_ord.get('pais', '')}

        if _rec:
            is_registered = True
            client_data = dict(_rec)
            # Prerellenar SOLO al cambiar de email (no pisar lo que el usuario edite después)
            if st.session_state.get('portal_last_email') != email_input:
                # Cliente reconocido recién entrado: saltar directo al paso 2 (envío)
                # — ya tiene sus datos, así avanza más rápido e intuitivo.
                st.session_state['_scroll_step2'] = True
                st.session_state['portal_nombre'] = client_data.get('nombre', '') or ''
                st.session_state['portal_empresa'] = client_data.get('empresa', '') or ''
                st.session_state['portal_telefono'] = client_data.get('telefono', '') or ''
                st.session_state['portal_pais'] = (client_data.get('pais', '') or 'Spain')
                st.session_state['portal_last_email'] = email_input
                # Recordar la modalidad de envío de su último pedido (menos fricción)
                _lo_ship = [p for p in load_pedidos()
                            if (p.get('client_email', '') or '').strip().lower() == _eml_lc]
                if _lo_ship:
                    _lo = sorted(_lo_ship, key=lambda x: x.get('fecha', ''))[-1]
                    if _lo.get('tipo_precio') in ('FOB', 'CIF'):
                        st.session_state['portal_tipo'] = _lo.get('tipo_precio')
                    if _lo.get('destino'):
                        st.session_state['portal_dest'] = _lo.get('destino')
            st.success(_T['welcome_back'].format(name=client_data.get('nombre') or email_input))
        else:
            # Email nuevo/no reconocido: limpiar datos de un cliente anterior
            if st.session_state.get('portal_last_email') != email_input:
                st.session_state['portal_nombre'] = ''
                st.session_state['portal_empresa'] = ''
                st.session_state['portal_telefono'] = ''
                st.session_state['portal_pais'] = 'Spain'
                st.session_state['portal_last_email'] = email_input
            st.info(_T['not_registered'])
            show_register = True

    if email_input:
        # #10 Retomar carrito pendiente: si el cliente vuelve y aún no tiene nada
        # en el carrito de esta sesión, restaurar el que dejó guardado.
        if st.session_state.get('_cart_restored_for') != email_input:
            st.session_state['_cart_restored_for'] = email_input
            if not st.session_state.get('portal_carrito'):
                _saved_cart = load_portal_carts().get(email_input.strip().lower(), {}).get('items', [])
                if _saved_cart:
                    st.session_state.portal_carrito = _saved_cart
                    st.session_state['_cart_was_restored'] = True
                    # Sembrar los inputs (qty/unit) para que el pre-pass fresco y la
                    # lista muestren las cantidades restauradas (si no, se borrarían).
                    _code_to_idx_r = {p.get('codigo'): _ri for _ri, p in enumerate(prods)}
                    for _it_r in _saved_cart:
                        _cir = _it_r.get('codigo')
                        if _cir in _code_to_idx_r:
                            _ixr = _code_to_idx_r[_cir]
                            _iur = _it_r.get('unidad', 'Pallets')
                            _qvr = _it_r.get('pallets', 0) if _iur == 'Pallets' else _it_r.get('cajas', 0)
                            st.session_state[f'portal_qty_{_cir}_{_ixr}'] = int(round(_qvr))
                            st.session_state[f'portal_unit_{_cir}_{_ixr}'] = (_T['unit_pallets'] if _iur == 'Pallets' else _T['unit_boxes'])
        _client_orders_all = [p for p in load_pedidos() if p.get('client_email','').lower() == email_input.lower()]
        _n_orders = len(_client_orders_all)
        tab_datos, tab_historial = st.tabs([_T['tab_datos'], _T['tab_pedidos'].format(n=_n_orders)])

        with tab_datos:
            # Cliente reconocido: mostrar resumen limpio y plegar el formulario.
            # Solo se despliega si quiere editar — así llega antes a los productos.
            if is_registered and not show_register:
                _ini = (client_data.get('nombre') or email_input or '·').strip()[:1].upper()
                _sub = ' · '.join([x for x in [client_data.get('empresa',''), client_data.get('pais',''), email_input] if x])
                st.markdown(
                    '<div style="display:flex;align-items:center;gap:13px;padding:13px 15px;border:1px solid #e7eaef;'
                    'background:#eef6f2;border-radius:14px;margin:2px 0 4px">'
                    f'<div style="width:42px;height:42px;border-radius:50%;background:#0c6e51;color:#fff;font-weight:800;'
                    f'font-size:1.05rem;display:flex;align-items:center;justify-content:center;flex:0 0 auto">{_esc(_ini)}</div>'
                    '<div style="min-width:0;line-height:1.3">'
                    f'<div style="font-weight:700;color:#16201b;font-size:1.02rem">{_esc(client_data.get("nombre") or email_input)}</div>'
                    f'<div style="color:#65726b;font-size:.85rem;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_esc(_sub)}</div></div>'
                    '<div style="margin-left:auto;flex:0 0 auto;background:#eef6f2;color:#0b5a42;font-size:.72rem;font-weight:700;'
                    f'padding:5px 11px;border-radius:20px;white-space:nowrap">✓ {_esc(_T.get("client_ready","Listo para pedir"))}</div>'
                    '</div>', unsafe_allow_html=True)
                _form_host = st.expander('✏️ ' + _T.get('edit_data', 'Editar mis datos'), expanded=False)
            else:
                _form_host = st.container()
            with _form_host:
              c1, c2 = st.columns(2)
              nombre = c1.text_input(_T['nombre_label'], key='portal_nombre')
              empresa = c2.text_input(_T['empresa_label'], key='portal_empresa')
              c3, c4 = st.columns(2)
              telefono = c3.text_input(_T['telefono_label'], placeholder=_T['telefono_ph'], key='portal_telefono')
              _paises_opts = ['Afghanistan','Albania','Algeria','Andorra','Angola','Argentina','Armenia','Australia','Austria','Azerbaijan','Bahrain','Bangladesh','Belarus','Belgium','Belize','Benin','Bolivia','Bosnia and Herzegovina','Botswana','Brazil','Bulgaria','Burkina Faso','Cambodia','Cameroon','Canada','Chile','China','Colombia','Congo','Costa Rica','Croatia','Cuba','Czech Republic','Denmark','Dominican Republic','Ecuador','Egypt','El Salvador','Estonia','Ethiopia','Finland','France','Georgia','Germany','Ghana','Greece','Guatemala','Haiti','Honduras','Hungary','India','Indonesia','Iran','Iraq','Ireland','Israel','Italy','Jamaica','Japan','Jordan','Kazakhstan','Kenya','Kuwait','Latvia','Lebanon','Libya','Lithuania','Luxembourg','Madagascar','Malaysia','Mali','Malta','Mexico','Moldova','Mongolia','Morocco','Mozambique','Myanmar','Netherlands','New Zealand','Nicaragua','Nigeria','Norway','Oman','Pakistan','Panama','Paraguay','Peru','Philippines','Poland','Portugal','Qatar','Romania','Russia','Rwanda','Saudi Arabia','Senegal','Serbia','Singapore','Slovakia','Slovenia','Somalia','South Africa','South Korea','Spain','Sri Lanka','Sudan','Sweden','Switzerland','Syria','Taiwan','Tanzania','Thailand','Tunisia','Turkey','Uganda','Ukraine','United Arab Emirates','United Kingdom','United States','Uruguay','Uzbekistan','Venezuela','Vietnam','Yemen','Zambia','Zimbabwe']
              # Sanear el país guardado: si no está en la lista (p.ej. "España" en vez de
              # "Spain"), caer en Spain — evita que el selectbox (key) reviente.
              if st.session_state.get('portal_pais') not in _paises_opts:
                  st.session_state['portal_pais'] = 'Spain'
              pais = c4.selectbox(_T['pais_label'], options=_paises_opts, key='portal_pais')
              if show_register:
                  st.caption(_T['auto_register'])
              _sv_c1, _sv_c2 = st.columns([1,3])
              if _sv_c1.button(_T.get('save_data_btn','💾 Guardar datos'), key='portal_save_client_btn', type='primary', use_container_width=True):
                  import re as _re_sv
                  _eml_sv = (email_input or '').strip().lower()
                  _nm_sv = (nombre or '').strip()
                  if not _re_sv.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', _eml_sv):
                      st.error(_T.get('err_email_format','Formato de email inválido'))
                  elif not _nm_sv:
                      st.error(_T.get('err_nombre', '✏️ Ingresa tu nombre completo'))
                  else:
                      _now_sv = datetime.now().isoformat()
                      # Fusión segura: no perder campos existentes ni blanquearlos
                      portal_clients[_eml_sv] = _merge_client_record(
                          portal_clients.get(_eml_sv, {}),
                          {'nombre': _nm_sv, 'empresa': empresa, 'telefono': telefono, 'pais': pais, 'email': _eml_sv})
                      portal_clients[_eml_sv].setdefault('fecha_registro', _now_sv)
                      portal_clients[_eml_sv].setdefault('pedidos', [])
                      save_portal_clients(portal_clients)
                      _adm_sv = load_clients()
                      _adm_sv[_eml_sv] = _merge_client_record(
                          _adm_sv.get(_eml_sv, {}),
                          {'nombre': _nm_sv, 'email': _eml_sv, 'empresa': empresa, 'telefono': telefono, 'pais': pais, 'origen': 'portal_cliente'})
                      _adm_sv[_eml_sv].setdefault('fecha_registro', _now_sv)
                      _adm_sv[_eml_sv].setdefault('pedidos_ids', [])
                      save_clients(_adm_sv)
                      st.cache_data.clear()
                      st.success('' + _nm_sv)
                      st.rerun()

        with tab_historial:
            if not _client_orders_all:
                st.info(_T['no_orders'])
            else:
                st.markdown(f'#### 📋 {_n_orders} Pedido(s) realizados')
                for op in sorted(_client_orders_all, key=lambda x: x.get('fecha',''), reverse=True):
                    op_id = op.get('id','')
                    op_fecha = op.get('fecha','')[:10]
                    op_estado = op.get('estado','Recibido')
                    op_total = op.get('total_usd',0)
                    op_tipo = op.get('tipo_precio','FOB')
                    op_dest = op.get('destino','')
                    op_icon = ESTADO_ICONS.get(op_estado, '📦')
                    _clr_map = {'Recibido':'#0066cc','Confirmado':'#28a745','Preparando':'#fd7e14','Enviado':'#6f42c1','Entregado':'#20c997','Cancelado':'#dc3545'}
                    _col = _clr_map.get(op_estado,'#666')
                    with st.expander(f'{op_icon} {op_id} | {op_fecha} | {op_tipo} | ${op_total:,.2f} USD', expanded=False):
                        st.markdown(
                            '<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin:2px 0 4px">'
                            f'{estado_badge(op_estado)}'
                            f'<span style="color:#566472;font-size:.82rem">{op_tipo}{(" · " + _esc(op_dest)) if op_tipo=="CIF" and op_dest else ""} · {_esc(op_fecha)}</span>'
                            f'<span style="margin-left:auto;font-weight:800;color:#084a37;font-variant-numeric:tabular-nums">${op_total:,.2f} USD</span>'
                            '</div>', unsafe_allow_html=True)
                        # Ciclo de vida del pedido (estilo Finanzas)
                        st.markdown(fsm_timeline_html(op_estado), unsafe_allow_html=True)
                        op_hist = op.get('historial_estados', [])
                        if op_hist:
                            with st.expander(_T['full_history'], expanded=False):
                                st.markdown(eventos_timeline_html(op_hist), unsafe_allow_html=True)
                        if op.get('productos'):
                            st.markdown(f'**{_T["products_label"]}**')
                            for _pit in op.get('productos',[]):
                                st.caption(f'• {_pit.get("producto","")} — {_pit.get("cajas",0)} cj | {_pit.get("pallets",0):.2f} plt | ${_pit.get("total",0):,.2f}')
                        can_cancel = op_estado not in ['Cancelado','Entregado','Enviado']
                        # PATCH 17: Repetir pedido button
                        _rp_c1, _rp_c2, _rp_c3 = st.columns([1, 1, 3])
                        if _rp_c1.button(_T['btn_repeat'], key=f'repeat_{op_id}', help=_T['btn_repeat_help'], use_container_width=True):
                            _repeat_prods = op.get('productos', [])
                            if _repeat_prods:
                                st.session_state.portal_carrito = []
                                # Índice POR LISTA ACTIVA (prods) — el mismo que usan el
                                # bucle y el pre-pass. Usar el catálogo completo desalinearía
                                # las claves si hay productos inactivos. La unidad se siembra
                                # con la ETIQUETA traducida (en EN, 'Boxes' no 'Cajas').
                                _code_to_idx_rp = {p.get('codigo',''): _i for _i, p in enumerate(prods)}
                                _rep_loaded = 0
                                for _rp in _repeat_prods:
                                    _rp_cod = _rp.get('codigo','')
                                    _rp_cajas = int(_rp.get('cajas',0))
                                    _rp_pallets = float(_rp.get('pallets',0))
                                    _rp_unit = _rp.get('unidad','Pallets')
                                    if _rp_cajas > 0 and _rp_cod in _code_to_idx_rp:
                                        _ri = _code_to_idx_rp[_rp_cod]
                                        _rqv = _rp_pallets if _rp_unit == 'Pallets' else _rp_cajas
                                        st.session_state[f'portal_qty_{_rp_cod}_{_ri}'] = int(round(_rqv))
                                        st.session_state[f'portal_unit_{_rp_cod}_{_ri}'] = (_T['unit_pallets'] if _rp_unit == 'Pallets' else _T['unit_boxes'])
                                        _rep_loaded += 1
                                st.session_state['_cart_saved_snap'] = None  # forzar recálculo/guardado
                                if _rep_loaded:
                                    st.success(_T['order_repeat_loaded'].format(pid=op_id))
                                    st.rerun()
                                else:
                                    st.warning(_T['order_no_products'])
                            else:
                                st.warning(_T['order_no_products'])
                        if can_cancel:
                            if _rp_c2.button(_T['btn_cancel'], key=f'cancel_{op_id}', type='secondary', use_container_width=True):
                                st.session_state[f'confirm_cancel_{op_id}'] = True
                        if st.session_state.get(f'confirm_cancel_{op_id}'):
                            st.warning(_T['confirm_cancel'].format(pid=op_id).replace('<b>', '**').replace('</b>', '**'))
                            _cc1, _cc2, _ = st.columns([1,1,4])
                            if _cc1.button(_T['btn_yes_cancel'], key=f'do_cancel_{op_id}'):
                                _all_peds = load_pedidos()
                                for _tp in _all_peds:
                                    if _tp.get('id') == op_id:
                                        _tp['estado'] = 'Cancelado'
                                        _tp['historial_estados'] = _tp.get('historial_estados',[]) + [{'estado':'Cancelado','fecha':datetime.now().isoformat(),'usuario':email_input,'nota':'Cancelado por cliente via portal'}]
                                        break
                                save_pedidos(_all_peds)
                                sync_finanzas(_tp, _all_peds)
                                log_email('order@exportharet.com', f'CANCELACION {op_id} solicitada por {email_input}', 'cancelacion_cliente')
                                st.session_state[f'confirm_cancel_{op_id}'] = False
                                st.cache_data.clear()
                                st.success(_T['order_cancelled'].format(pid=op_id))
                                st.rerun()
                            if _cc2.button(_T['btn_no'], key=f'no_cancel_{op_id}'):
                                st.session_state[f'confirm_cancel_{op_id}'] = False
                                st.rerun()
        st.markdown('---')
    else:
        nombre = empresa = telefono = pais = ''
        st.info(_T['enter_email'])
        return
    # PATCH VAL: Hard validation - require valid email + name before continuing
    import re as _re_val
    _email_pat = r'^[^@\s]+@[^@\s]+\.[^@\s]+'
    _email_ok = bool(_re_val.match(_email_pat, (email_input or '').strip()))
    _nombre_val = (st.session_state.get('portal_nombre','') or '').strip()
    if not _email_ok:
        st.warning(_T['enter_valid_email'])
        return
    if not _nombre_val:
        st.warning(_T['enter_full_name'])
        return
    # ── PASO 2: Tipo de precio + Destino ─────────────────────────────────────
    # Ancla para el auto-scroll cuando un cliente reconocido entra (salta al paso 2).
    st.markdown('<div id="eh-step2-anchor" style="position:relative;top:-70px"></div>', unsafe_allow_html=True)
    _eh_seccion(_T['step2'], 2)
    # Si el cliente acaba de ser reconocido, desplazar la vista al paso 2 (una vez).
    if st.session_state.pop('_scroll_step2', False):
        import streamlit.components.v1 as _components
        _components.html(
            """<script>
            (function(){
              const tryScroll = (n) => {
                const doc = window.parent && window.parent.document;
                const el = doc && doc.getElementById('eh-step2-anchor');
                if (el) { el.scrollIntoView({behavior:'smooth', block:'start'}); }
                else if (n < 20) { setTimeout(()=>tryScroll(n+1), 120); }
              };
              setTimeout(()=>tryScroll(0), 250);
            })();
            </script>""",
            height=0,
        )
    _lang_en = st.session_state.get('portal_lang') == 'en'
    _tp_labels = ({'FOB': '🇪🇨 Price in Ecuador (you arrange transport)',
                   'CIF': '📍 Delivered to your city (freight included)'} if _lang_en
                  else {'FOB': '🇪🇨 Precio en Ecuador (tú recoges)',
                        'CIF': '📍 Puesto en tu ciudad (flete incluido)'})

    def _flete_of(_d):
        _dv = dests.get(_d, 0)
        return float(_dv) if isinstance(_dv, (int, float)) else (_dv.get('factor', 0) if isinstance(_dv, dict) else 0)

    # Resumen Pro del envío (siempre visible) + controles plegados: el cliente
    # solo despliega si quiere cambiar de modalidad. Usa el valor persistido.
    _cur_tipo = st.session_state.get('portal_tipo', 'FOB')
    _cur_dest = st.session_state.get('portal_dest', '')
    if _cur_tipo == 'CIF' and _cur_dest:
        _ship_ico = '📍'
        _ship_t = _T['ship_cif_title'].format(dest=_cur_dest)
        _ship_s = _T['ship_cif_sub'].format(flete=_flete_of(_cur_dest))
    else:
        _ship_ico = '🇪🇨'
        _ship_t = _T['ship_fob_title']
        _ship_s = _T['ship_fob_sub']
    st.markdown(
        '<div style="display:flex;align-items:center;gap:13px;padding:13px 15px;border:1px solid #e7eaef;'
        'background:#eef6f2;border-radius:14px;margin:2px 0 4px">'
        f'<div style="font-size:1.45rem;flex:0 0 auto;line-height:1">{_ship_ico}</div>'
        '<div style="min-width:0;line-height:1.3">'
        f'<div style="font-weight:700;color:#16201b;font-size:1.02rem">{_esc(_ship_t)}</div>'
        f'<div style="color:#65726b;font-size:.85rem">{_esc(_ship_s)}</div></div>'
        '</div>', unsafe_allow_html=True)

    with st.expander('🚚 ' + _T.get('edit_shipping', 'Cambiar modalidad de envío'), expanded=False):
        t1, t2 = st.columns([1, 2])
        tipo_precio = t1.radio(_T['price_type_label'], ['FOB', 'CIF'], key='portal_tipo', horizontal=False,
            format_func=lambda x: _tp_labels.get(x, x), help=_T['price_type_help'])
        destino = ''
        dest_flete = 0.0
        if tipo_precio == 'CIF':
            dest_opts = list(dests.keys()) if dests else []
            if not dest_opts:
                t2.warning(_T['no_dest'])
            else:
                destino = t2.selectbox(_T['dest_label'], dest_opts, key='portal_dest')
                dest_val = dests.get(destino, 0)
                dest_flete = float(dest_val) if isinstance(dest_val, (int, float)) else dest_val.get('factor', 0) if isinstance(dest_val, dict) else 0
                _puerto_orig = 'Quito/Guayaquil, Ecuador'
                t2.caption(_T['flete_caption'].format(flete=dest_flete, orig=_puerto_orig, dest=destino))
                t2.info(_T['cif_info'].format(dest=destino, orig=_puerto_orig))
        else:
            t2.info(_T['fob_info'])
            t2.caption(_T['fob_origin'])

    # ── PASO 3: Selección de Productos ───────────────────────────────
    _eh_seccion(_T['step3'], 3)
    if st.session_state.pop('_cart_was_restored', False):
        st.success(_T.get('cart_restored', 'Retomamos tu pedido.'), icon=None)
    st.info(_T['price_update_notice'], icon=None)
    # Toggle de moneda: ver precios en la moneda del destino (CIF) o en USD — transparencia total
    _mon_dest = data.get('config', {}).get('destinos_moneda', {}).get(destino, 'USD') if tipo_precio == 'CIF' else 'USD'
    _ver_usd = False
    if tipo_precio == 'CIF' and _mon_dest != 'USD':
        _sym_dest = MONEDA_SIMBOLO.get(_mon_dest, _mon_dest)
        _cur_opts = [f'{_sym_dest} Ver en {_mon_dest}', '$ Ver en USD']
        _mc1, _mc2 = st.columns([1.4, 2])
        with _mc1:
            _msel = st.radio('Moneda', _cur_opts, key='portal_ver_moneda',
                             horizontal=True, label_visibility='collapsed')
        _ver_usd = _msel.endswith('USD')
        with _mc2:
            st.caption('💱 Total transparencia: cambia entre tu moneda y USD. '
                       'La transacción se realiza en USD.')
    # PATCH STEP3: CSS for responsive cards, sticky header, bigger qty buttons, highlight
    st.markdown('''<style>
    /* Bigger +/- buttons on number_input (P8) */
    div[data-testid="stNumberInput"] button {
        min-width: 44px !important;
        min-height: 44px !important;
        font-size: 1.25rem !important;
        font-weight: 700 !important;
    }
    div[data-testid="stNumberInput"] input {
        min-height: 44px !important;
        font-size: 1.05rem !important;
        text-align: center !important;
    }
    /* Sticky catalog header (P11) */
    .eh-cat-header {
        position: sticky; top: 56px; z-index: 30;
        background: #fff; padding: 8px 0;
        border-bottom: 2px solid #0c6e51;
        font-weight: 700;
        display: grid; grid-template-columns: 2.4fr 2.3fr 3.5fr 2fr 1.6fr;
        gap: 8px; align-items: center;
    }
    .eh-cat-header > div { color: #0c6e51; font-size: 0.92rem; }
    /* Mobile: hide sticky header, products will stack */
    @media (max-width: 768px) {
        .eh-cat-header { display: none; }
        /* ── Móvil: cada producto = TARJETA limpia. Línea 1: icono + nombre +
              precio. Debajo: ficha + controles (cantidad + unidad) cómodos. ──
           OUTER = la fila (contiene un bloque anidado). INNER = los controles. */
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stHorizontalBlock"]) {
            display: flex !important;
            flex-direction: row !important;
            flex-wrap: wrap !important;
            align-items: center !important;
            gap: 4px 12px !important;
            padding: 13px 15px !important;
            margin: 0 0 12px !important;
            border: 1px solid #e7eaef !important;
            border-radius: 14px !important;
            background: #ffffff !important;
            box-shadow: 0 1px 2px rgba(18,28,42,.05) !important;
        }
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stHorizontalBlock"]) > div:nth-child(1) {
            flex: 0 0 auto !important; width: auto !important; min-width: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stHorizontalBlock"]) > div:nth-child(2) {
            flex: 1 1 0 !important; width: auto !important; min-width: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stHorizontalBlock"]) > div:nth-child(3) {
            flex: 0 0 auto !important; width: auto !important; min-width: 0 !important; margin-left: auto !important;
        }
        div[data-testid="stHorizontalBlock"]:has(div[data-testid="stHorizontalBlock"]) > div:nth-child(4) {
            flex: 1 1 100% !important; width: 100% !important; min-width: 100% !important; margin-top: 11px !important;
        }
        /* INNER: cantidad + unidad lado a lado (no apilados), cómodos */
        div[data-testid="stHorizontalBlock"] div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"]) {
            display: flex !important;
            flex-direction: row !important;
            flex-wrap: nowrap !important;
            gap: 9px !important;
            padding: 0 !important;
            margin: 0 !important;
            border: none !important;
            background: transparent !important;
            box-shadow: none !important;
        }
        div[data-testid="stHorizontalBlock"] div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"]) > div {
            flex: 1 1 0 !important; width: auto !important; min-width: 0 !important;
        }
        div[data-testid="stHorizontalBlock"] div[data-testid="stHorizontalBlock"]:has(div[data-testid="stNumberInput"]) > div:nth-child(1) {
            flex: 1.25 1 0 !important;
        }
        div[data-testid="stNumberInput"] button {
            min-width: 44px !important;
            min-height: 46px !important;
        }
        /* Las tarjetas ya separan; fuera divisores sueltos entre filas */
        .eh-row-div { display: none !important; }
        /* Selector de idioma: inline a la derecha (no apilado en bloques grandes) */
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es){
            flex-direction: row !important; flex-wrap: nowrap !important;
            align-items: center !important; justify-content: flex-end !important; gap: 7px !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es) > div{
            flex: 0 0 auto !important; width: auto !important; min-width: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.st-key-btn_lang_es) > div:first-child{
            flex: 1 1 auto !important;
        }
    }
    /* PEND3: Sticky Guardar Precios admin */
    div[data-testid="stButton"] button[kind="primary"]:has-text("Guardar Precios") {
        position: sticky;
        bottom: 8px;
        z-index: 50;
        box-shadow: 0 4px 12px rgba(0,62,140,0.25);
    }
    /* Cleaner product row styling */
    .eh-cat-row-marker { display:none; }
    /* Compact cart banner button alignment */
    .eh-cart-banner-actions { display:flex; gap:8px; align-items:center; justify-content:flex-end; }
    @media (max-width: 768px) {
        div[data-testid="stNumberInput"] input { min-height: 40px !important; font-size: 0.95rem !important; }
    }
    </style>''', unsafe_allow_html=True)
    # Vaciar pedido: resetear los inputs a 0 AQUÍ, antes de leerlos y de instanciar
    # los widgets (evita el error "no se puede modificar tras instanciar" y el desync
    # visual de Streamlit al borrar la clave). Lo dispara el botón "Vaciar pedido".
    if st.session_state.pop('_portal_clear_qty', False):
        for _vi, _vp in enumerate(prods):
            st.session_state[f"portal_qty_{_vp.get('codigo','')}_{_vi}"] = 0

    # ── FIX desfase: reconstruir el carrito FRESCO desde los inputs actuales ANTES
    # de pintar la barra de progreso, el resumen y la lista. Antes, esos bloques
    # leían el carrito del run anterior (p. ej. ponías 3 pallets y arriba contaba 2).
    # Al recalcular aquí y asignarlo a portal_carrito, TODO el run es consistente.
    _fresh_pal_total = 0.0
    _fresh_rows = []
    for _pi, _pp in enumerate(prods):
        _cod_p = _pp.get('codigo', '')
        _qv = st.session_state.get(f'portal_qty_{_cod_p}_{_pi}', 0) or 0
        if _qv <= 0:
            continue
        _gi_p = data.get('config', {}).get('grupos', {}).get(_pp.get('grupo', ''), {})
        _cxp_p = int(_gi_p.get('cajas_pallet', _pp.get('cajas_pallet', 160))) if isinstance(_gi_p, dict) else 160
        _uv = st.session_state.get(f'portal_unit_{_cod_p}_{_pi}', _T['unit_pallets'])
        _is_pal_p = (_uv == _T['unit_pallets'])
        _cj_p, _pal_p = cajas_y_pallets(_qv, 'Pallets' if _is_pal_p else 'Cajas', _cxp_p)
        _fresh_pal_total += _pal_p
        _fresh_rows.append({'codigo': _cod_p,
                            'producto': _pp.get('descripcion', '') or _pp.get('producto', '') or _cod_p,
                            'cajas': _cj_p, 'pallets': _pal_p, 'cxp': _cxp_p,
                            'unidad': 'Pallets' if _is_pal_p else 'Cajas'})
    for _fr in _fresh_rows:
        _pu_fr = get_precio_con_volumen(_fr['codigo'], destino, tipo_precio, data, max(_fresh_pal_total, 1)) or 0
        _fr['precio_usd'] = _pu_fr
        _fr['total'] = round(_fr['cajas'] * _pu_fr, 2)
    st.session_state.portal_carrito = _fresh_rows

    # Banner pedido mínimo
    _current_pallets = sum(i.get('pallets',0) for i in st.session_state.portal_carrito)
    # ── Indicador de volumen
    _desc_actual = get_descuento_volumen(max(_current_pallets, 1)) if _current_pallets >= 1 else 0.0
    # PATCH UX-CIF U6: Detectar cruce de tramo (toast cuando precio baja)
    if tipo_precio == 'CIF':
        _ux_prev_tramo = st.session_state.get('portal_last_tramo', None)
        _ux_curr_tramo_idx = -1
        for _ti, _tt in enumerate(TRAMOS_VOLUMEN):
            if _tt['min'] <= max(_current_pallets,1) <= _tt['max']:
                _ux_curr_tramo_idx = _ti
                break
        if _ux_prev_tramo is not None and _ux_curr_tramo_idx > _ux_prev_tramo and _current_pallets >= 1:
            try:
                st.toast(_T['unlock_better_price'].format(label=TRAMOS_VOLUMEN[_ux_curr_tramo_idx]['label']), icon='💰')
            except Exception:
                pass
        st.session_state['portal_last_tramo'] = _ux_curr_tramo_idx
    _next_tramo = None
    _pallets_para_siguiente = 0
    for _t in TRAMOS_VOLUMEN:
        if _t['descuento'] > _desc_actual:
            _next_tramo = _t
            _pallets_para_siguiente = max(0, _t['min'] - _current_pallets)
            break
    # Progreso hacia el pedido mínimo: lo muestra la barra inferior fija (una sola
    # fuente de verdad), no aquí — evita duplicar la misma información tres veces.
    _min_order = 3
    _needed = max(0.0, _min_order - _current_pallets)
    _cart_total_usd = sum(i.get('total',0) for i in st.session_state.portal_carrito)
    _cart_total_pal = sum(i.get('pallets',0) for i in st.session_state.portal_carrito)
    _cart_total_caj = sum(i.get('cajas',0) for i in st.session_state.portal_carrito)
    _cart_items = len([i for i in st.session_state.portal_carrito if i.get('cajas',0) > 0])
    if _cart_items > 0:
        # Solo "Vaciar pedido" (discreto a la derecha). El total va en la barra inferior.
        _vc_cols = st.columns([5, 2])
        with _vc_cols[1]:
            if st.button(_T.get('clear_cart','🗑️ Vaciar carrito'), key='portal_vaciar_top', use_container_width=True, type='secondary'):
                st.session_state['_portal_clear_qty'] = True   # se procesa antes de los inputs
                st.session_state.portal_carrito = []
                st.session_state['_cart_saved_snap'] = None
                st.rerun()
        # PATCH GROUP-AGG: Resumen por Grupo de Embalaje
        _grp_cfg = data.get('config',{}).get('grupos',{})
        _grp_agg = {}  # grupo -> {cajas, cxp, productos: []}
        _prod_by_code = {p.get('codigo'): p for p in (data.get('products') or [])}
        for _ci in st.session_state.portal_carrito:
            _ci_cod = _ci.get('codigo','')
            _ci_p = _prod_by_code.get(_ci_cod, {})
            _ci_grp = _ci_p.get('grupo','?')
            _ci_cxp = int(_grp_cfg.get(_ci_grp,{}).get('cajas_pallet', _ci_p.get('cajas_pallet',160)) or 160)
            if _ci_grp not in _grp_agg:
                _grp_agg[_ci_grp] = {'cajas':0, 'cxp':_ci_cxp, 'productos':[], 'nombre': _grp_cfg.get(_ci_grp,{}).get('nombre', _ci_grp)}
            _grp_agg[_ci_grp]['cajas'] += int(_ci.get('cajas',0))
            _grp_agg[_ci_grp]['productos'].append(_ci.get('producto', _ci_cod))
        if _grp_agg:
            _grp_html = ['<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:12px 16px;margin:4px 0 8px">']
            _grp_total_pal_real = 0.0
            for _gk, _gv in sorted(_grp_agg.items()):
                _g_pal_exact = _gv['cajas'] / _gv['cxp'] if _gv['cxp'] > 0 else 0
                _g_pal_fisicos = int(_gv['cajas'] / _gv['cxp']) + (1 if _gv['cajas'] % _gv['cxp'] > 0 else 0)
                _grp_total_pal_real += _g_pal_exact
                _g_full = _gv['cajas'] // _gv['cxp']
                _g_rem = _gv['cajas'] - _g_full * _gv['cxp']
                _g_breakdown = f"{_g_full} " + _T['group_full_pallets'] + (f" + {_g_rem}/{_gv['cxp']} " + _T['group_partial_boxes'] if _g_rem else "")
                _grp_html.append(
                    f'<div style="display:flex;justify-content:space-between;align-items:center;padding:6px 0;border-bottom:1px dashed #e2e8f0">'
                    f'<div><b style="color:#0c6e51">Grupo {_gk}</b> <small style="color:#666">— {" · ".join(sorted(set(_gv["productos"])))}</small></div>'
                    f'<div style="text-align:right;font-size:0.9rem">'
                    f'<b>{_gv["cajas"]:,} cj</b> / {_gv["cxp"]} cj-pal = <b style="color:#10a37a">{_g_pal_exact:.2f} pal</b>'
                    f'<br><small style="color:#888">{_g_breakdown}</small>'
                    f'</div>'
                    f'</div>'
                )
                # PATCH UX-CIF U7: sugerencia para completar pallet (solo CIF + parcial)
                if tipo_precio == 'CIF' and _g_rem > 0:
                    _ux_needed = _gv['cxp'] - _g_rem
                    # Encontrar productos del mismo grupo NO en carrito
                    _ux_codes_in = {_ci2.get('codigo','') for _ci2 in st.session_state.portal_carrito if _ci2.get('cajas',0) > 0}
                    _ux_grp_prods = [pp for pp in (data.get('products') or []) if pp.get('grupo','') == _gk and pp.get('codigo','') not in _ux_codes_in]
                    _ux_sug_txt = ''
                    if _ux_grp_prods:
                        _ux_names = ', '.join((pp.get('descripcion','') or pp.get('producto','') or pp.get('codigo','')) for pp in _ux_grp_prods[:3])
                        _ux_sug_txt = _T['group_complete_with'].format(names=_ux_names)
                    _grp_html.append(
                        f'<div style="background:#fef9c3;border-left:3px solid #ca8a04;padding:6px 10px;margin:4px 0 8px;border-radius:6px;font-size:0.82rem;color:#713f12">' +
                        (_T['group_missing'].format(n=_ux_needed, g=_gk) + _ux_sug_txt) +
                        f'</div>'
                    )
            _grp_html.append(
                f'<div style="display:flex;justify-content:space-between;align-items:center;padding-top:8px;font-weight:700;color:#0c6e51">'
                f'<div>{_T["group_total_pallets"]}</div>'
                f'<div>{_grp_total_pal_real:.2f} pal</div>'
                f'</div>'
            )
            _grp_html.append('</div>')
            with st.expander('📦 ' + _T['group_summary'], expanded=False):
                st.markdown(''.join(line.lstrip() for line in '\n'.join(_grp_html).split('\n')), unsafe_allow_html=True)
    # ── PRO: lista de frutas limpia (fila por fruta) ──
    st.markdown('''<style>
    .eh-row-ic{width:38px;height:38px;border-radius:10px;background:#eef6f2;display:flex;align-items:center;justify-content:center}
    .eh-row-ic svg{width:23px;height:23px}
    .eh-row-nm{font-weight:600;font-size:1rem;color:#16201b;line-height:1.25}
    .eh-row-sp{font-size:.78rem;color:#737d77;margin-top:1px}
    .eh-row-added{background:#eef6f2;color:#0b5a42;font-size:.66rem;font-weight:700;padding:2px 8px;border-radius:20px;margin-left:7px;white-space:nowrap}
    .eh-card-pr{line-height:1.15;white-space:nowrap;text-align:right}
    .eh-card-pr b{font-size:1.06rem}
    .eh-card-u{font-size:.7rem;color:#8a948c;margin-left:2px}
    .eh-card-kg{display:none}
    .eh-card-cj{font-size:.72rem;color:#0b5a42;font-weight:600;margin-top:3px}
    .eh-shop-h{font-size:1.05rem;font-weight:700;color:#16201b;margin:6px 0 2px;display:flex;justify-content:space-between;align-items:baseline}
    .eh-shop-h span{font-weight:500;font-size:.82rem;color:#737d77}
    .eh-row-div{border:none;border-top:1px solid #ebefec;margin:.45rem 0}
    </style>''', unsafe_allow_html=True)
    _shop_sub = (f'Puesto en {destino} · flete incluido' if (tipo_precio == 'CIF' and destino) else 'Precio en Ecuador (FOB)')
    st.markdown(f'<div class="eh-shop-h">Frutas disponibles <span>{_shop_sub}</span></div>', unsafe_allow_html=True)
    st.markdown('<hr class="eh-row-div" style="margin-top:0">', unsafe_allow_html=True)

    # Total de pallets FRESCO ya calculado arriba (mismo run, sin desfase).
    _total_pallets_fresh = _fresh_pal_total

    # Reconstruir carrito basado en los valores actuales de los inputs
    _new_carrito = []
    for idx, p in enumerate(prods):
        cod = p.get('codigo','')
        nombre_prod = p.get('descripcion','') or p.get('producto','') or cod
        _grp_x = p.get('grupo','')
        _gi_x = data.get('config',{}).get('grupos',{}).get(_grp_x,{})
        cxp = int(_gi_x.get('cajas_pallet',p.get('cajas_pallet',160))) if isinstance(_gi_x,dict) else 160
        _kg_x = float(p.get('kg_caja',0) or 0)
        _total_pallets_now = _total_pallets_fresh  # #1 fix: total fresco (mismo run)
        precio_u = get_precio_con_volumen(cod, destino, tipo_precio, data, max(_total_pallets_now, 1))
        _dv_x = data.get('config',{}).get('destinos',{}).get(destino,0)
        _fl_x = float(_dv_x.get('factor',_dv_x) if isinstance(_dv_x,dict) else _dv_x if isinstance(_dv_x,(int,float)) else 0)
        qty_key = f'portal_qty_{cod}_{idx}'
        unit_key = f'portal_unit_{cod}_{idx}'

        # Leer valor previo del carrito para pre-llenar
        _ex_item = next((x for x in st.session_state.portal_carrito if x.get('codigo') == cod), None)
        _ex_unit = _ex_item.get('unidad','Pallets') if _ex_item else 'Pallets'
        if _ex_item:
            if _ex_unit == 'Pallets':
                _ex_qty = int(_ex_item.get('pallets',0))
            else:
                _ex_qty = int(_ex_item.get('cajas',0))
        else:
            _ex_qty = 0

        # PRO: cada fruta es una FILA limpia — icono · nombre + ficha · precio · cantidad
        _svg_html, _bg_card = _fruit_art(nombre_prod)
        _kg_lbl = f'{_kg_x:.1f} {_T["unit_kg_per_box"]}'.replace('.', ',') if _kg_x else ''
        _min_cant_p = int(p.get('min_cantidad', 0) or 0)
        _min_unit_p = str(p.get('min_unidad', 'Pallets') or 'Pallets')
        _specs = ' · '.join([s for s in [_kg_lbl, (f'{cxp} cj/pallet' if cxp else ''),
                             (f'Grupo {_grp_x}' if _grp_x else '')] if s])
        _added_mark = ' <span class="eh-row-added">✓ en tu pedido</span>' if _ex_qty > 0 else ''
        _r = st.columns([0.45, 2.4, 1.35, 3.8], gap='small', vertical_alignment='center')
        _r[0].markdown(f'<div class="eh-row-ic">{_svg_html}</div>', unsafe_allow_html=True)
        _r[1].markdown(
            f'<div class="eh-row-nm">{_esc(nombre_prod)}{_added_mark}</div>'
            f'<div class="eh-row-sp">{_specs}</div>', unsafe_allow_html=True)
        # Controles: cantidad + unidad en LÍNEA (compacto). precio→col2, cajas→col1
        _ctrl = _r[3].columns([1.15, 1], gap='small', vertical_alignment='center')
        gc = [_r[1], _r[2], _ctrl[0], _ctrl[1], _r[1]]
        # Col 1: Precio por caja
        _mon_x = data.get('config',{}).get('destinos_moneda',{}).get(destino,'USD') if tipo_precio=='CIF' else 'USD'
        _rate_x = get_exchange_rates().get(_mon_x,1.0)
        _sym_x = MONEDA_SIMBOLO.get(_mon_x,_mon_x)
        _fob_x = get_fob_price(cod,data)
        # Precio segun volumen actual
        # PATCH UX-CIF U4: calcular precio en proximo tramo y ahorro (solo CIF)
        _ux_next_price = None
        _ux_next_min = None
        _ux_base_price_1 = get_precio_con_volumen(cod, destino, tipo_precio, data, 1); _ux_save_pct = round((1 - precio_u / _ux_base_price_1) * 100, 1) if (_ux_base_price_1 and precio_u and _ux_base_price_1 > precio_u) else 0; _ux_save_pct = max(_ux_save_pct, 10.0) if _total_pallets_now >= 3 else _ux_save_pct
        if True:
            # Encontrar siguiente tramo con mas pallets (FOB y CIF)
            for _tt in TRAMOS_VOLUMEN:
                if _tt['min'] > max(_total_pallets_now, 1):
                    _ux_next_min = int(_tt['min'])
                    _ux_next_price = get_precio_con_volumen(cod, destino, tipo_precio, data, _ux_next_min)
                    if _ux_next_price and precio_u and _ux_next_price < precio_u:
                        _ux_base_price_1 = get_precio_con_volumen(cod, destino, tipo_precio, data, 1); _ux_save_pct = round((1 - precio_u / _ux_base_price_1) * 100, 1) if _ux_base_price_1 and _ux_base_price_1 > precio_u else 0; _ux_save_pct = max(_ux_save_pct, 10.0) if _total_pallets_now >= 3 else _ux_save_pct
                    break
        # Precio + incentivo de volumen (móvil y escritorio igual):
        #  · si ya tiene descuento por volumen → precio a 1 pallet TACHADO + ahorro real
        #  · si aún no → gancho: "−X/caja a 3+ pallets" para incentivar
        _ux_badge_html = ''; _ux_strike_html = ''; _ux_price_color = '#0c6e51'
        _conv = (_mon_x != 'USD' and tipo_precio == 'CIF' and _rate_x != 1.0 and not _ver_usd)
        _sym_show = _sym_x if _conv else '$'
        if _ux_base_price_1:
            _p1 = round(_ux_base_price_1 * (_rate_x if _conv else 1.0), 2)   # ancla: 1 pallet
            _pu = round(precio_u * (_rate_x if _conv else 1.0), 2)           # precio actual
            if precio_u < _ux_base_price_1:  # YA con descuento por volumen
                _ux_strike_html = (f'<span style="color:#9ca3af;text-decoration:line-through;'
                                   f'font-size:0.74em;margin-right:5px">{_sym_show}{_p1:.2f}</span>')
                _ux_badge_html = (f'<div style="color:#0b5a42;font-size:0.74em;font-weight:600;'
                                  f'margin-top:1px;white-space:nowrap">−{_sym_show}{_p1 - _pu:.2f}/caja</div>')
                _ux_price_color = '#10a37a'
            elif _ux_next_min and _ux_next_price and _ux_next_price < precio_u:
                # gancho hacia el SIGUIENTE tramo real (contextual por fruta)
                _pnd = round((precio_u - _ux_next_price) * (_rate_x if _conv else 1.0), 2)
                _ux_badge_html = (f'<div style="color:#0d8a67;font-size:0.72em;font-weight:600;'
                                  f'margin-top:1px;white-space:nowrap">−{_sym_show}{_pnd:.2f}/caja a {_ux_next_min}+ pallets</div>')
            else:  # sin tramo siguiente arriba → gancho base a 3+ pallets
                _p3 = get_precio_con_volumen(cod, destino, tipo_precio, data, 3)
                if _p3 and _p3 < _ux_base_price_1:
                    _p3d = round((_ux_base_price_1 - _p3) * (_rate_x if _conv else 1.0), 2)
                    _ux_badge_html = (f'<div style="color:#0d8a67;font-size:0.72em;font-weight:600;'
                                      f'margin-top:1px;white-space:nowrap">−{_sym_show}{_p3d:.2f}/caja a 3+ pallets</div>')
        _precio_show = round(precio_u * _rate_x, 2) if _conv else precio_u
        gc[1].markdown(
            f'<div style="line-height:1.2;white-space:nowrap;text-align:right">'
            f'<b style="color:{_ux_price_color};font-size:1.1em">{_sym_show}{_precio_show:.2f}</b>'
            f'{_ux_strike_html}{_ux_badge_html}</div>', unsafe_allow_html=True)
        # Col 2: Cantidad con +/- nativo
        qty_val = gc[2].number_input(
            _T['col_qty'], min_value=0, value=_ex_qty, step=1,
            key=qty_key, label_visibility='collapsed'
        )
        # Col 3: Unidad
        _unit_opts = [_T['unit_pallets'], _T['unit_boxes']]
        _unit_default = 0 if _ex_unit == 'Pallets' else 1
        unit_sel_raw = gc[3].selectbox(
            _T['col_unit'], _unit_opts, index=_unit_default,
            key=unit_key, label_visibility='collapsed'
        )
        # Map translated label back to canonical value for internal logic
        unit_sel = 'Pallets' if unit_sel_raw == _T['unit_pallets'] else 'Cajas'
        # Col 4: Cajas calculadas (solo informacion) — PATCH 8
        if qty_val > 0:
            _n_cajas, _n_pallets = cajas_y_pallets(qty_val, unit_sel, cxp)  # fuente única
            if unit_sel == 'Pallets':
                _cajas_label = f'**{_n_cajas:,}** cj'
                _cajas_sub = f'<small style="color:#666">{int(qty_val)} pal \u00d7 {cxp} cj/pal</small>'
            else:
                _cajas_label = f'**{_n_cajas:,}** cj'
                _cajas_sub = f'<small style="color:#666">≈ {_n_pallets:.2f} pal ({cxp} cj/pal)</small>'
            gc[4].markdown(f'{_cajas_label}\n{_cajas_sub}', unsafe_allow_html=True)
            # Validar cantidad mínima
            if _min_cant_p > 0:
                _qty_in_unit = qty_val if unit_sel == _min_unit_p else (
                    round(qty_val * cxp) if unit_sel == 'Pallets' else round(qty_val / cxp, 2)
                )
                if _qty_in_unit < _min_cant_p:
                    _min_warn = _T['min_warning'].format(n=_min_cant_p, u=_min_unit_p, p=nombre_prod).replace('<b>', '**').replace('</b>', '**')
                    st.warning(_min_warn, icon=None)
            # Agregar al nuevo carrito
            _new_carrito.append({
                'codigo': cod, 'producto': nombre_prod,
                'cajas': _n_cajas, 'pallets': _n_pallets,
                'precio_usd': precio_u,
                'total': round(_n_cajas * precio_u, 2),
                'unidad': unit_sel,
                'fob_usd': _fob_x,
                'flete_usd': _fl_x,
                'descuento_vol': (lambda _b: round((_b - precio_u) / _b, 4) if _b and _b > precio_u else 0)(get_precio_con_volumen(cod, destino, tipo_precio, data, 1))
            })
        # divisor fino entre filas (lista limpia "Pro")
        if idx < len(prods) - 1:
            st.markdown('<hr class="eh-row-div">', unsafe_allow_html=True)

    # Sincronizar carrito SIN rerun forzado: el st.rerun() hacía que la pantalla
    # saltara al inicio al elegir la cantidad (el cliente perdía de vista el producto).
    # El resumen y el total inferiores se calculan del carrito ya actualizado.
    st.session_state.portal_carrito = _new_carrito

    # #10 Persistir el carrito por email SOLO cuando cambia (evita escrituras en cada
    # rerun). Así el cliente puede cerrar y retomar el pedido donde lo dejó.
    if email_input:
        _cart_snap = [(i.get('codigo'), i.get('cajas', 0), i.get('unidad', 'Pallets')) for i in _new_carrito]
        if st.session_state.get('_cart_saved_snap') != _cart_snap:
            try:
                _carts_all = load_portal_carts()
                _eml_cart = email_input.strip().lower()
                if _new_carrito:
                    _carts_all[_eml_cart] = {'items': _new_carrito, 'ts': datetime.now().isoformat()}
                else:
                    _carts_all.pop(_eml_cart, None)
                save_portal_carts(_carts_all)
                st.session_state['_cart_saved_snap'] = _cart_snap
            except Exception as _e_cart:
                logger.warning(f'guardar carrito falló: {_e_cart}')

    # ── Nube flotante sutil: descuento total por volumen + incentivo del siguiente pallet ──
    _fp = sum(i.get('pallets', 0) for i in _new_carrito)
    _bmon = _mon_dest if (tipo_precio == 'CIF' and not _ver_usd) else 'USD'
    if _bmon != 'USD':
        _brate = get_exchange_rates().get(_bmon, 1.0); _bsym = MONEDA_SIMBOLO.get(_bmon, _bmon)
    else:
        _brate = 1.0; _bsym = '$'
    # La barra mide lo CRÍTICO: progreso hacia el mínimo de pedido. Una vez
    # alcanzado, en CIF muestra el ahorro por volumen. Mensaje corto y claro.
    if _fp <= 0:
        _fb_pct = 0
        _fb_left = '📦 0 pallets'
        _fb_msg = f'Empieza tu pedido · mínimo {_min_order} pallets'
    elif _fp < _min_order:
        _fb_left = f'📦 <b>{_fp:.1f}</b>/{_min_order} pallets'
        _fb_pct = int(min(100, _fp / _min_order * 100))
        _fb_msg = f'Faltan <b>{(_min_order - _fp):.1f}</b> para poder enviar'
    elif tipo_precio != 'CIF':
        _fb_left = f'📦 <b>{_fp:.1f}</b> pallets'
        _fb_pct = 100
        _fb_msg = '✓ Listo para enviar'
    else:
        _fb_left = f'📦 <b>{_fp:.1f}</b> pallets'
        _next_p = int(_fp) + 1
        _base_sum = 0.0; _ah_total = 0.0; _ah_next = 0.0
        for _it in _new_carrito:
            _cod = _it.get('codigo', ''); _cj = _it.get('cajas', 0)
            _pcur = _it.get('precio_usd', 0)
            _p1 = get_precio_con_volumen(_cod, destino, tipo_precio, data, 1)
            _pn = get_precio_con_volumen(_cod, destino, tipo_precio, data, _next_p)
            if _p1:
                _base_sum += _cj * _p1
                _ah_total += _cj * (_p1 - _pcur)
            if _pn and _pn < _pcur:
                _ah_next += _cj * (_pcur - _pn)
        _pct_now = (_ah_total / _base_sum * 100) if _base_sum > 0 else 0
        _fb_pct = min(100, int(_pct_now / 13 * 100))
        _ah_t_d = _ah_total * _brate
        _ah_n_d = _ah_next * _brate
        if _ah_total < 1:
            _fb_msg = f'➕ Añade 1 pallet y ahorras <b>{_bsym}{_ah_n_d:,.0f}</b>'
        else:
            _fb_msg = f'💰 Ahorras <b>{_bsym}{_ah_t_d:,.0f}</b> por volumen'
    # #4 Resumen inferior más informativo: nº de frutas + total (la lista detallada
    # está justo debajo). Peek de nombres en el title para repaso sin hacer scroll.
    _fb_items = [i for i in _new_carrito if i.get('cajas', 0) > 0]
    _fb_n = len(_fb_items)
    _fb_tot_disp = sum(i.get('total', 0) for i in _new_carrito) * _brate
    _fb_peek = ' · '.join(f"{_esc(i.get('producto',''))} ({i.get('cajas',0)} cj)" for i in _fb_items[:8])
    if _fb_n:
        _fruta_w = 'fruta' if _fb_n == 1 else 'frutas'
        _fb_mid = (
            f'<span title="{_fb_peek}" style="white-space:nowrap;cursor:default">🧾 <b>{_fb_n}</b> {_fruta_w}</span>'
            f'<span style="white-space:nowrap;font-weight:800;color:#084a37;font-size:1.02rem">{_bsym}{_fb_tot_disp:,.0f}</span>')
    else:
        _fb_mid = ''
    st.markdown(
        '<div style="position:fixed;left:0;right:0;bottom:0;z-index:60;'
        'background:rgba(255,255,255,.93);backdrop-filter:blur(6px);'
        'border-top:1px solid #e0eae3;box-shadow:0 -4px 18px rgba(20,60,40,.08);padding:8px 14px">'
        '<div style="max-width:780px;margin:0 auto;display:flex;align-items:center;gap:12px;'
        'flex-wrap:wrap;font-size:.84rem;color:#1B2620">'
        f'<span style="white-space:nowrap">{_fb_left}</span>'
        f'{_fb_mid}'
        '<div style="flex:1;min-width:90px;height:7px;background:#e3ede6;border-radius:6px;overflow:hidden">'
        f'<div style="height:100%;width:{_fb_pct}%;background:linear-gradient(90deg,#10a37a,#0c6e51);'
        'transition:width .4s ease"></div></div>'
        f'<span style="color:#0b5a42;font-weight:600;white-space:nowrap">{_fb_msg}</span>'
        '</div></div>', unsafe_allow_html=True)

    # Carrito
    if st.session_state.portal_carrito:
        st.markdown('---')
        _tot_c = sum(i['total'] for i in st.session_state.portal_carrito)
        _plt_c = sum(i.get('pallets',0) for i in st.session_state.portal_carrito)
        _cj_c = sum(i.get('cajas',0) for i in st.session_state.portal_carrito)
        # Cotización en tiempo real
        _rates_meta = get_exchange_rates_meta()
        _rates_portal = _rates_meta['rates']
        _rate_ts = _rates_meta['ts']
        _rate_live = _rates_meta['live']
        _rate_src = _rates_meta['source']
        _eur_rate = _rates_portal.get('EUR', 0.92)
        # Moneda destino
        _moneda_dest = 'USD'
        if tipo_precio == 'CIF' and destino:
            _dv2 = data.get('config',{}).get('destinos',{}).get(destino,{})
            _moneda_dest = data.get('config',{}).get('destinos_moneda',{}).get(destino, _dv2.get('moneda','USD') if isinstance(_dv2,dict) else 'USD')
        _dest_rate = _rates_portal.get(_moneda_dest, 1)
        _disp_mon = _moneda_dest if _moneda_dest != 'USD' else 'EUR'; _disp_rate = _rates_portal.get(_disp_mon, _eur_rate); _disp_sym = MONEDA_SIMBOLO.get(_disp_mon, _disp_mon); _show_dest = False
        # ── Resumen del pedido (responsive: cards en móvil, tabla en desktop) ──
        # Total destacado arriba para reducir scroll en móvil
        _tot_eur = round(_tot_c * _disp_rate, 2)
        _resumen_html = ['<div class="eh-resumen-wrap">']
        # CSS local
        _resumen_html.append('''<style>
        .eh-resumen-wrap { margin: 8px 0 14px; }
        .eh-resumen-total {
            background: #0c6e51; color: #fff; border-radius: 14px; padding: 16px 20px;
            display: flex; justify-content: space-between; align-items: center;
            flex-wrap: wrap; gap: 10px; margin-bottom: 12px;
            box-shadow: 0 8px 20px -8px rgba(12,110,81,.45);
        }
        .eh-resumen-total .eh-tot-lbl { font-size: 0.7rem; opacity: 0.82; text-transform: uppercase; letter-spacing: 0.08em; font-weight: 700; }
        .eh-resumen-total .eh-tot-val { font-size: 1.55rem; font-weight: 800; line-height: 1.1; letter-spacing:-.02em; }
        .eh-resumen-total .eh-tot-eur { font-size: 0.92rem; opacity: 0.9; }
        .eh-resumen-total .eh-tot-meta { font-size: 0.8rem; opacity: 0.9; }
        .eh-resumen-card {
            background: #fff; border: 1px solid #e7eaef; border-radius: 12px;
            padding: 11px 15px; margin-bottom: 8px; display: grid;
            grid-template-columns: 1fr auto; gap: 3px 12px; align-items: center;
        }
        .eh-resumen-card .eh-prod { font-weight: 700; color: #131a21; font-size: 0.96rem; }
        .eh-resumen-card .eh-precio { color: #084a37; font-weight: 800; text-align: right; white-space: nowrap; font-variant-numeric:tabular-nums; }
        .eh-resumen-card .eh-meta { grid-column: 1 / -1; display: flex; flex-wrap: wrap; align-items: center; gap: 4px 10px; font-size: 0.78rem; color: #8b95a3; padding-top: 4px; margin-top: 2px; border-top: 1px solid #f1f4f7; }
        .eh-resumen-card .eh-meta b { color: #131a21; font-weight: 600; }
        .eh-resumen-card .eh-eur { color: #0c6e51; font-weight: 600; }
        @media (min-width: 740px) {
          .eh-resumen-card { grid-template-columns: 2.3fr 0.9fr 0.9fr 0.9fr 1.1fr 1fr; gap: 6px 10px; }
          .eh-resumen-card .eh-prod { grid-column: auto; }
          .eh-resumen-card .eh-meta { grid-column: auto; display: contents; border: none; padding: 0; margin: 0; font-size: 0.9rem; color: #566472; }
          .eh-resumen-card .eh-meta-cell { padding: 0; }
        }
        </style>''')
        # Total card destacado arriba
        _moneda_extra = ''
        if _show_dest:
            _tot_dest = round(_tot_c * _dest_rate, 2)
            _sym_dest = MONEDA_SIMBOLO.get(_moneda_dest, _moneda_dest)
            _moneda_extra = f' &nbsp;·&nbsp; <span style="opacity:0.9">{_sym_dest}{_tot_dest:,.2f} {_moneda_dest}</span>'
        _resumen_html.append(f'''
        <div class="eh-resumen-total">
          <div>
            <div class="eh-tot-lbl">📦 {_T["order_total_label"]}</div>
            <div class="eh-tot-val">${_tot_c:,.2f} USD</div>
            <div class="eh-tot-eur">≈ {_disp_sym}{_tot_eur:,.2f} {_disp_mon}{_moneda_extra}</div>
          </div>
          <div style="text-align:right">
            <div class="eh-tot-meta">📦 {_plt_c:.2f} {_T["unit_pallets"]} · {_cj_c:,} {_T["unit_boxes"]}</div>
            <div class="eh-tot-meta">🧾 {len(st.session_state.portal_carrito)} {_T["unit_products"]}</div>
          </div>
        </div>
        ''')
        # Filas por producto (cards responsive)
        for _ci, _item in enumerate(st.session_state.portal_carrito):
            _item_eur = round(_item['total'] * _disp_rate, 2)
            # PATCH UX-CIF U1: badge ahorro vs precio base (1 pal) - solo CIF
            _ux_save_html = ''
            if tipo_precio == 'CIF':
                _ux_base_price = get_precio_con_volumen(_item.get('codigo',''), destino, tipo_precio, data, 1)
                _ux_curr_price = _item.get('precio_usd', 0)
                if _ux_base_price and _ux_curr_price and _ux_base_price > _ux_curr_price:
                    _ux_diff = round(_ux_base_price - _ux_curr_price, 2)
                    _ux_pct = round((1 - _ux_curr_price / _ux_base_price) * 100, 1)
                    _ux_save_html = f'<span class="eh-meta-cell" style="color:#9ca3af;text-decoration:line-through">${_ux_base_price:.2f}/cj</span><span class="eh-meta-cell" style="color:#10a37a;font-weight:700">{_T["savings_per_box"].format(d=_ux_diff)}</span>'
            _resumen_html.append(f'''
            <div class="eh-resumen-card">
              <div class="eh-prod">{_item["producto"]}</div>
              <div class="eh-precio">${_item["total"]:,.2f}</div>
              <div class="eh-meta">
                <span class="eh-meta-cell">${_item["precio_usd"]:.2f}/cj</span>
                {_ux_save_html}
                <span class="eh-meta-cell"><b>{_item["pallets"]:.2f}</b> plt</span>
                <span class="eh-meta-cell"><b>{_item["cajas"]:,}</b> cj</span>
                <span class="eh-meta-cell eh-eur">{_disp_sym}{_item_eur:,.2f}</span>
              </div>
            </div>
            ''')
        _resumen_html.append('</div>')
        st.markdown(''.join(line.lstrip() for line in ''.join(_resumen_html).split('\n')), unsafe_allow_html=True)
        # (Bloque antiguo de tabla y fila TOTAL eliminado — ahora se muestra arriba)
        # Moneda destino adicional (si no es USD ni EUR)
        if _show_dest:
            _tot_dest = round(_tot_c * _dest_rate, 2)
            _sym_dest = MONEDA_SIMBOLO.get(_moneda_dest, _moneda_dest)
            st.markdown(
                '<div style="background:#f0fff4;border:1px solid #c3e6cb;border-radius:6px;padding:8px 14px;margin:6px 0">' + _T['price_in_dest'].format(m=_moneda_dest, sym=_sym_dest, tot=_tot_dest, rate=_dest_rate) + '</div>',
                unsafe_allow_html=True
            )
        # Nota tipo de precio y tasa EUR
        _live_lbl = '🟢 En vivo' if _rate_live else '⚪ Aprox.'
        _live_lbl_t = _T['rate_live'] if _rate_live else _T['rate_approx']
        st.markdown(
            '<div style="margin:6px 0;padding:5px 0"><small style="color:#777">' + _T['rate_info'].format(sym=_disp_sym, rate=_disp_rate, m=_disp_mon, live=_live_lbl_t, src=_rate_src, ts=_rate_ts) + '</small><br><small style="color:#999"><i>' + _T['rate_info_sub'].format(m=_disp_mon) + '</i></small></div>',
            unsafe_allow_html=True
        )
        if tipo_precio == 'FOB':
            st.caption(_T['cart_fob'])
        if st.button(_T['clear_cart'], key='portal_vaciar', use_container_width=False):
            st.session_state['_portal_clear_qty'] = True   # se procesa antes de los inputs
            st.session_state.portal_carrito = []
            st.session_state['_cart_saved_snap'] = None
            st.rerun()
    st.markdown('---')

    if st.session_state.portal_carrito:
        # ── PASO 4: Confirmar y Enviar Pedido ────────────────────────────────────
        _eh_seccion(_T['step4'], 4)
        notas = st.text_area(_T.get('notes_label','📝 Notas / instrucciones especiales'), placeholder=_T.get('notes_ph','Ej: Entrega en almacén X, condiciones especiales...'), key='portal_notas')

        _toptES=['','Pago anticipado 100%','50% adelanto / 50% contra documentos','30% adelanto / 70% contra BL','Carta de cr\xe9dito (LC)','Pago a 30 d\xedas','Pago a 60 d\xedas','Otro']
        _toptEN=['','100% advance payment','50% advance / 50% against documents','30% advance / 70% against BL','Letter of credit (LC)','30-day payment','60-day payment','Other']
        TOPT = _toptEN if st.session_state.get('portal_lang','es')=='en' else _toptES
        p_term=st.selectbox(_T.get('payment_terms', '📋 Términos de pago (opcional)'),TOPT,key='p_term')
        # PATCH 19: Full order summary before confirm
        if st.session_state.portal_carrito and email_input and nombre:
            tot_final = sum(i['total'] for i in st.session_state.portal_carrito)
            _tot_pal_fin = sum(i.get('pallets',0) for i in st.session_state.portal_carrito)
            _tot_caj_fin = sum(i.get('cajas',0) for i in st.session_state.portal_carrito)
            # Calculo de peso total del pedido
            _tot_peso_fin = 0.0
            _data_tp = load_data()
            _grupos_tp = _data_tp.get('config',{}).get('grupos',{})
            _prods_tp = {p.get('codigo',''): p for p in (_data_tp.get('products') or [])}
            for _it_tp in st.session_state.portal_carrito:
                _p_tp = _prods_tp.get(_it_tp.get('codigo',''), {})
                _g_tp = _p_tp.get('grupo','')
                _kg_tp = float(_grupos_tp.get(_g_tp,{}).get('kg_caja', _p_tp.get('kg_caja', 0)) or 0)
                _tot_peso_fin += int(_it_tp.get('cajas',0)) * _kg_tp
            # PATCH UX-CIF U8: calcular ahorro total vs precio base (1 pal) - solo CIF
            _ux_total_save = 0.0
            if tipo_precio == 'CIF':
                for _pfi2 in st.session_state.portal_carrito:
                    if _pfi2.get('cajas', 0) > 0:
                        _ux_base_pr = get_precio_con_volumen(_pfi2.get('codigo',''), destino, tipo_precio, data, 1)
                        _ux_curr_pr = _pfi2.get('precio_usd', 0)
                        if _ux_base_pr and _ux_curr_pr and _ux_base_pr > _ux_curr_pr:
                            _ux_total_save += (_ux_base_pr - _ux_curr_pr) * _pfi2.get('cajas', 0)
            _ux_total_save = round(_ux_total_save, 2)
            tipo_str = tipo_precio + (f' → {destino}' if tipo_precio == 'CIF' and destino else '')
            # Destination currency
            _fin_mon = data.get('config',{}).get('destinos_moneda',{}).get(destino,'USD') if tipo_precio=='CIF' else 'USD'
            _fin_rate = get_exchange_rates().get(_fin_mon, 1.0)
            _fin_sym = MONEDA_SIMBOLO.get(_fin_mon, _fin_mon)
            _fin_dest_total = round(tot_final * _fin_rate, 2) if _fin_mon != 'USD' and _fin_rate != 1.0 else None
            _fin_alt = f'<br><span style="font-size:0.88em;color:rgba(255,255,255,0.92);font-weight:500">≈ {_fin_sym}{_fin_dest_total:,.2f} {_fin_mon}</span>' if _fin_dest_total else ''
            # Build mobile-friendly responsive cards summary
            _prod_cards_html = ''
            _data_for_peso = load_data()
            _grupos_for_peso = _data_for_peso.get('config',{}).get('grupos',{})
            _prods_for_peso = {p.get('codigo',''): p for p in (_data_for_peso.get('products') or [])}
            for _pfi in st.session_state.portal_carrito:
                if _pfi.get('cajas', 0) > 0:
                    # Calcular peso estimado del item
                    _pfi_prod = _prods_for_peso.get(_pfi.get('codigo',''), {})
                    _pfi_grp = _pfi_prod.get('grupo','')
                    _pfi_kg_caja = float(_grupos_for_peso.get(_pfi_grp,{}).get('kg_caja', _pfi_prod.get('kg_caja', 0)) or 0)
                    _pfi_peso = int(_pfi.get('cajas',0)) * _pfi_kg_caja
                    _pfi_peso_lbl = f' · <b>{_pfi_peso:,.0f} kg</b>' if _pfi_peso > 0 else ''
                    _prod_cards_html += (
                        f'<div class="eh-cnf-card">'
                        f'<div class="eh-cnf-prod"><b>{_pfi.get("producto","")}</b><span class="eh-cnf-cod">{_pfi.get("codigo","")}</span></div>'
                        f'<div class="eh-cnf-row"><span>{_pfi.get("pallets",0):.2f} pal · {int(_pfi.get("cajas",0)):,} cj{_pfi_peso_lbl}</span><span>${_pfi.get("precio_usd",0):.2f}/cj</span><span class="eh-cnf-tot">${_pfi.get("total",0):,.2f}</span></div>'
                        f'</div>'
                    )
            _conf_html = (
                '<style>'
                '.eh-cnf-wrap { background:#f0f7ff; border:2px solid #0c6e51; border-radius:12px; padding:14px 16px; margin:10px 0; }'
                '.eh-cnf-wrap h4 { margin:0 0 10px 0; color:#0c6e51; font-size:1.1rem; }'
                '.eh-cnf-meta { font-size:0.92rem; line-height:1.5; color:#1a2540; margin-bottom:10px; }'
                '.eh-cnf-meta b { color:#0c6e51; }'
                '.eh-cnf-card { background:#fff; border:1px solid #d4dff2; border-radius:8px; padding:8px 12px; margin-bottom:6px; }'
                '.eh-cnf-prod { display:flex; justify-content:space-between; align-items:baseline; font-size:0.98rem; }'
                '.eh-cnf-cod { font-size:0.78rem; color:#6c7a93; font-weight:normal; }'
                '.eh-cnf-row { display:flex; justify-content:space-between; align-items:center; gap:8px; font-size:0.88rem; color:#3a4a6b; padding-top:4px; flex-wrap:wrap; }'
                '.eh-cnf-tot { color:#0c6e51; font-weight:700; font-size:0.98rem; margin-left:auto; }'
                '.eh-cnf-total-row { display:flex; justify-content:space-between; align-items:center; background:linear-gradient(135deg,#0c6e51 0%,#1a4f9e 100%); color:#fff; border-radius:10px; padding:12px 16px; margin-top:8px; flex-wrap:wrap; gap:8px; }'
                '.eh-cnf-total-row .eh-cnf-tl { font-size:0.82rem; opacity:0.88; }'
                '.eh-cnf-total-row .eh-cnf-tv { font-size:1.35rem; font-weight:700; line-height:1.1; }'
                '.eh-cnf-total-row .eh-cnf-tx { font-size:0.92rem; opacity:0.92; }'
                '@media (min-width:640px) { .eh-cnf-prod { font-size:1.02rem; } }'
                '</style>'
                '<div class="eh-cnf-wrap">'
                f'<h4>{_T["order_summary_title"]}</h4>'
                f'<div class="eh-cnf-meta"><b>{_T["order_lbl_client"]}</b> {_esc(nombre)} ({_esc(email_input)})<br>'
                f'<b>{_T["order_lbl_company"]}</b> {_esc(empresa) or "N/A"} &nbsp;|&nbsp; <b>{_T["order_lbl_country"]}</b> {_esc(pais) or "N/A"}<br>'
                f'<b>{_T["order_lbl_mode"]}</b> {tipo_str} &nbsp;|&nbsp; <b>{_T["order_lbl_payment"]}</b> {p_term or _T["order_lbl_pending"]}'
                + (f'<br><span style="color:#10a37a;font-weight:600">{_T["order_savings"].format(s=_ux_total_save)}</span>' if (tipo_precio == 'CIF' and _ux_total_save > 0) else '')
                + '</div>'
                f'{_prod_cards_html}'
                '<div class="eh-cnf-total-row">'
                f'<div><div class="eh-cnf-tl">{_T["order_total_label"]}</div><div class="eh-cnf-tv">${tot_final:,.2f} USD</div>{_fin_alt}</div>'
                f'<div style="text-align:right"><div class="eh-cnf-tx">{_T["order_n_pallets"].format(n=_tot_pal_fin)}</div><div class="eh-cnf-tx">{_T["order_n_boxes"].format(n=_tot_caj_fin)}</div>' + (f'<div class="eh-cnf-tx">{_T["order_weight"].format(n=_tot_peso_fin)}</div><div class="eh-cnf-tx" style="font-size:0.72rem;opacity:0.78;font-style:italic;margin-top:2px;text-align:right">{_T["order_weight_note"]}</div>' if _tot_peso_fin > 0 else '') + '</div>'
                '</div>'
                '</div>'
            )
            st.markdown(''.join(line.lstrip() for line in _conf_html.split('\n')), unsafe_allow_html=True)

        # PEND2: Calculadora moneda destino
        with st.expander(_T['calc_title'], expanded=False):
            _calc_rates = get_exchange_rates_meta()['rates']
            _calc_opts = ['USD','EUR','GBP','CAD','BRL','MXN','COP','PEN','CLP','ARS']
            _calc_opts = [m for m in _calc_opts if m == 'USD' or m in _calc_rates]
            _calc_mon = st.selectbox(_T['calc_convert_to'], _calc_opts, index=0, key='calc_mon_sel')
            _calc_rate = float(_calc_rates.get(_calc_mon, 1.0)) if _calc_mon != 'USD' else 1.0
            _calc_sym = MONEDA_SIMBOLO.get(_calc_mon, _calc_mon)
            _calc_total = tot_final * _calc_rate
            _calc_pal = (tot_final / _tot_pal_fin) if _tot_pal_fin > 0 else 0
            _calc_pal_dest = _calc_pal * _calc_rate
            st.markdown(f"""<div style='background:#f0f7ff;border:1px solid #cfe1ff;border-radius:10px;padding:14px;margin-top:8px'>
            <div style='display:flex;justify-content:space-between;font-size:1.08rem'><span>{_T['calc_total_usd']}</span><b>$ {tot_final:,.2f}</b></div>
            <div style='display:flex;justify-content:space-between;font-size:1.08rem;margin-top:4px'><span>{_T['calc_total_dest'].format(m=_calc_mon)}</span><b>{_calc_sym} {_calc_total:,.2f}</b></div>
            <div style='display:flex;justify-content:space-between;font-size:0.92em;color:#555;margin-top:6px'><span>{_T['calc_rate']}</span><span>1 USD = {_calc_rate:.4f} {_calc_mon}</span></div>
            <div style='display:flex;justify-content:space-between;font-size:0.92em;color:#555;margin-top:2px'><span>{_T['calc_per_pallet']}</span><span>{_calc_sym} {_calc_pal_dest:,.2f}</span></div>
            <small style='color:#888;display:block;margin-top:8px'>{_T['calc_note']}</small>
            </div>""", unsafe_allow_html=True)

        btn_guardar = st.button(_T['confirm_btn'], type='primary', use_container_width=True, key='portal_guardar')

        if btn_guardar:
            if not email_input:
                st.error(_T['err_email'])
            elif not nombre:
                st.error(_T['err_nombre'])
            elif not st.session_state.portal_carrito:
                st.error(_T['err_cart'])
            elif tipo_precio == 'CIF' and not destino:
                st.error(_T['err_destino'])
            elif sum(i.get('pallets',0) for i in st.session_state.portal_carrito) < 3:
                _curr_pal_v = sum(i.get('pallets',0) for i in st.session_state.portal_carrito)
                _falt_pal_v = 3 - _curr_pal_v
                st.error(_T['min_order_alert'].format(curr=_curr_pal_v, miss=_falt_pal_v))
            else:
                # C3: PID timestamp+uuid to prevent race condition
                _tod_p = load_pedidos()
                _yn_p = datetime.now().strftime('%Y')
                _ts_p = datetime.now().strftime('%m%d%H%M%S')
                _uid_p = uuid.uuid4().hex[:4].upper()
                pid = f'PED-{_yn_p}-{_ts_p}-{_uid_p}'
                _existing_ids = {p.get('id','') for p in _tod_p}
                while pid in _existing_ids:
                    _uid_p = uuid.uuid4().hex[:4].upper()
                    pid = f'PED-{_yn_p}-{_ts_p}-{_uid_p}'
                tot = sum(i['total'] for i in st.session_state.portal_carrito)
                # Ensure rates/currency vars are available (fallback if cart block didn't run)
                try:
                    _ = _rates_portal
                except NameError:
                    _rates_portal = get_exchange_rates_meta()['rates']
                try:
                    _ = _moneda_dest
                except NameError:
                    _dv_fb = data.get('config',{}).get('destinos',{}).get(destino,{})
                    _moneda_dest = _dv_fb.get('moneda','USD') if isinstance(_dv_fb,dict) else 'USD'
                ped = {
                    'id': pid,
                    'client_email': email_input,
                    'client_name': nombre,
                    'empresa': empresa,
                    'telefono': telefono,
                    'pais': pais,
                    'tipo_precio': tipo_precio,
                    'destino': destino if tipo_precio == 'CIF' else 'FOB',
                    'moneda': 'USD',
                    'moneda_dest': _moneda_dest,
                    'flete_usd_caja': dest_flete if tipo_precio == 'CIF' else 0.0,
                    'tasa_cambio': round(_rates_portal.get(_moneda_dest, 1.0), 4),
                    'total_moneda_dest': round(round(tot, 2) * _rates_portal.get(_moneda_dest, 1.0), 2),
                    'productos': list(st.session_state.portal_carrito),
                    'total_usd': round(tot, 2),
                    'estado': 'Recibido',
                    'fecha': datetime.now().isoformat(),
                    'notas':notas,'terminos_pago':p_term,'historial_estados': [{'estado': 'Recibido', 'fecha': datetime.now().isoformat(), 'usuario': 'portal'}],
                    'creado_por': 'portal',
                    'lang': st.session_state.get('portal_lang', 'es'),
                }
                # Guardar pedido
                todos = load_pedidos()
                todos.append(ped)
                save_pedidos(todos)
                sync_finanzas(ped, todos)
                # Registrar / actualizar cliente en portal — FUSIÓN segura (no pierde
                # ni blanquea datos existentes; solo añade el nuevo pedido).
                _now_cf = datetime.now().isoformat()
                portal_clients[email_input] = _merge_client_record(
                    portal_clients.get(email_input, {}),
                    {'nombre': nombre, 'empresa': empresa, 'telefono': telefono, 'pais': pais, 'email': email_input})
                portal_clients[email_input].setdefault('fecha_registro', _now_cf)
                portal_clients[email_input]['pedidos'] = list(dict.fromkeys(
                    (portal_clients.get(email_input, {}).get('pedidos', []) or []) + [pid]))
                save_portal_clients(portal_clients)
                _adm = load_clients()
                _adm[email_input] = _merge_client_record(
                    _adm.get(email_input, {}),
                    {'nombre': nombre, 'email': email_input, 'empresa': empresa, 'telefono': telefono, 'pais': pais, 'origen': 'portal_cliente'})
                _adm[email_input].setdefault('fecha_registro', _now_cf)
                _adm[email_input]['pedidos_ids'] = list(set((_adm.get(email_input, {}).get('pedidos_ids', []) or []) + [pid]))
                save_clients(_adm)
                # Log email y envio real a order@exportharet.com
                log_email(email_input, f'Confirmación pedido {pid}', 'portal_cliente')
                _email_ok = send_order_email(ped)  # #6: refleja si el email salió de verdad
                st.cache_data.clear()

                # Guardar pedido en session para acciones post-guardado
                st.session_state['ultimo_pedido'] = ped
                st.session_state.portal_carrito = []
                # #10 El pedido ya se envió: borrar el carrito pendiente guardado
                try:
                    _carts_done = load_portal_carts()
                    if _carts_done.pop((email_input or '').strip().lower(), None) is not None:
                        save_portal_carts(_carts_done)
                    st.session_state['_cart_saved_snap'] = []
                except Exception:
                    pass
                # Clear step-4 fields so next order starts clean
                for _k in ['portal_notas','p_term']: st.session_state.pop(_k, None)
                st.success(_T['order_confirmed'].format(pid=pid))
                if _email_ok:
                    st.info(_T['order_sent_email'].replace('<b>', '**').replace('</b>', '**'))
                else:
                    st.info('✅ Pedido recibido correctamente. Nuestro equipo te '
                            'contactará en 24-48h para confirmar. Síguelo en **Mis Pedidos** ↑')

    # ── Acciones post-pedido ─────────────────────────────────────────────────
    if st.session_state.get('ultimo_pedido'):
        ped_saved = st.session_state['ultimo_pedido']
        pid_saved = ped_saved.get('id','')
        _nom_post = _esc(ped_saved.get('client_name',''))
        st.markdown('---')
        # #6/#8 Confirmación Pro (verde de marca) con nº de pedido destacado + timeline
        _post_steps = _T.get('post_steps', 'Recibido|Confirmamos (24 h)|Preparación|Envío').split('|')
        _steps_html = ''
        for _si, _slbl in enumerate(_post_steps):
            _on = _si == 0
            _dot = ('background:#fff;color:#0c6e51' if _on else 'background:rgba(255,255,255,.22);color:#fff')
            _txt = ('#fff;font-weight:700' if _on else 'rgba(255,255,255,.8);font-weight:500')
            _conn = ('<div style="flex:1;height:2px;background:rgba(255,255,255,.25);margin:0 4px;min-width:10px"></div>' if _si < len(_post_steps) - 1 else '')
            _steps_html += (f'<div style="display:flex;align-items:center;gap:6px;white-space:nowrap">'
                            f'<span style="width:16px;height:16px;border-radius:50%;{_dot};font-size:.62rem;'
                            f'display:flex;align-items:center;justify-content:center;font-weight:800">{"✓" if _on else _si+1}</span>'
                            f'<span style="font-size:.74rem;color:{_txt}">{_esc(_slbl)}</span></div>{_conn}')
        st.markdown(f'''<div style="background:linear-gradient(135deg,#0c6e51,#084a37);padding:22px 24px;border-radius:16px;margin:12px 0;box-shadow:0 10px 30px rgba(15,79,41,.22)">
          <div style="display:flex;align-items:center;gap:13px">
            <div style="width:46px;height:46px;border-radius:50%;background:rgba(255,255,255,.18);display:flex;align-items:center;justify-content:center;font-size:1.5rem;flex:0 0 auto">&#x2705;</div>
            <div style="line-height:1.2">
              <div style="font-size:.7rem;letter-spacing:1.4px;text-transform:uppercase;color:rgba(255,255,255,.85);font-weight:700">{_esc(_T.get("order_received_kick","Pedido recibido"))}</div>
              <div style="color:#fff;font-size:1.45rem;font-weight:800;margin-top:1px">{_esc(pid_saved.upper())}</div>
            </div>
          </div>
          <p style="color:#eaf5ee;margin:13px 0 14px;line-height:1.5;font-size:.94rem">{_T["post_order_thanks"].format(name=_nom_post)}</p>
          <div style="display:flex;align-items:center;flex-wrap:wrap;gap:4px 2px;border-top:1px solid rgba(255,255,255,.18);padding-top:13px">{_steps_html}</div>
        </div>''', unsafe_allow_html=True)
        pdf_bytes, pdf_mime, pdf_ext = build_order_pdf(ped_saved)
        st.caption(_T.get('post_copy_hint', 'Guarda tu comprobante y, si quieres, confírmanos por WhatsApp o email:'))
        # Acciones en columnas
        ac1, ac2, ac3 = st.columns(3)
        # Descargar PDF albarán
        ac1.download_button(
            label=_T['download_pdf'],
            data=pdf_bytes,
            file_name=f'{pid_saved}{pdf_ext}',
            mime=pdf_mime,
            use_container_width=True,
            key='dl_pedido'
        )
        # WhatsApp — professional message
        tot_wa = ped_saved.get('total_usd', 0)
        _nom_wa = ped_saved.get('client_name', '')
        _emp_wa = ped_saved.get('empresa', '')
        _tipo_wa = ped_saved.get('tipo_precio', 'FOB')
        _dest_wa = ped_saved.get('destino', '')
        _pais_wa = ped_saved.get('pais', '')
        _prods_lines = '%0A'.join([' • ' + str(i.get('cajas','')) + ' cajas ' + str(i.get('producto','')) + ' ($' + str(i.get('precio_usd','')) + '/caja)' for i in ped_saved.get('productos',[])])
        _emp_line = (_T['wa_msg_company'] + ' ' + _emp_wa + '%0A') if _emp_wa else ''
        _dest_line = ('Destino: ' + _dest_wa) if _tipo_wa == 'CIF' and _dest_wa else 'Precio FOB (sin flete)'
        wa_text_lines = (
            _T['wa_msg_greeting'] + '%0A%0A'
            + _T['wa_msg_intro'] + '%0A%0A'
            + _T['wa_msg_order'] + f': {pid_saved}*%0A'
            + _T['wa_msg_client'] + f' {_nom_wa}%0A'
            + _emp_line
            + _T['wa_msg_country'] + f' {_pais_wa}%0A'
            + f'📦 {_dest_line}%0A%0A'
            + _T['wa_msg_details'] + '%0A'
            + _prods_lines
            + '%0A%0A' + _T['wa_msg_total'] + f': ${tot_wa:,.2f} USD*%0A%0A'
            + _T['wa_msg_closing'] + '%0A'
            + _T['wa_msg_regards']
        )
        wa_url = f'https://wa.me/34641076116?text={wa_text_lines}'
        ac2.link_button(_T['wa_confirm'], wa_url, use_container_width=True)
        # Email — professional with albarán note
        subject = _T['em_subj'] + f' {pid_saved} — Export Haret'
        _nom_mail = ped_saved.get('client_name','')
        _emp_mail = ped_saved.get('empresa','')
        _emp_mail_line = (_T['em_body_company'] + ' ' + _emp_mail + '%0A') if _emp_mail else ''
        _prods_mail = '%0A'.join(['- ' + str(i.get('cajas','')) + ' cajas ' + str(i.get('producto','')) for i in ped_saved.get('productos',[])])
        body = (
            f'Estimado equipo de Export Haret,%0A%0A'
            + _T['em_body_intro'] + f' {pid_saved}.%0A%0A'
            + _T['em_body_data'] + '%0A'
            + _T['em_body_client'] + f' {_nom_mail}%0A'
            + _emp_mail_line
            + _T['em_body_incoterm'] + f' {_tipo_wa}' + (f' — {_dest_wa}' if _dest_wa and _tipo_wa=='CIF' else '') + '%0A'
            + _T['em_body_total'] + f' ${tot_wa:,.2f} USD%0A%0A'
            + _T['em_body_products'] + '%0A'
            + _prods_mail
            + '%0A%0A'
            + _T['em_body_closing'] + '%0A'
            + _T['wa_msg_regards']
        )
        mailto_url = f'mailto:order@exportharet.com?subject={subject.replace(" ", "%20")}&body={body.replace(" ", "%20")}'
        ac3.link_button(_T['em_send'], mailto_url, use_container_width=True)
        # Nueva Orden
        if st.button(_T['new_order_btn'], key='nuevo_portal'):
            st.session_state['ultimo_pedido'] = None
            # Clear step-4 form fields
            for _k in ['portal_notas','p_term']: st.session_state.pop(_k, None)
            # Clear all product quantity inputs so client picks fresh quantities
            _keys_to_clear = [k for k in st.session_state.keys() if k.startswith('portal_qty_') or k.startswith('portal_unit_')]
            for _k in _keys_to_clear: st.session_state.pop(_k, None)
            st.session_state.portal_carrito = []
            st.rerun()
    st.markdown('---')
    st.markdown('---')
    with st.expander(_T['quote_expander'],expanded=False):
        st.markdown(_T['quote_intro'])
        _cc1,_cc2=st.columns(2)
        _cn=_cc1.text_input(_T['quote_name_lbl'],key='cnom',placeholder=_T['quote_name_ph'])
        _ce=_cc2.text_input(_T['email_label'],value=st.session_state.get('portal_email_input',''),key='ceml')
        _cd=_cc1.text_input(_T['dest_label'].replace('', ''),key='cdst',placeholder='ej: Madrid, España')
        _cplt=_cc2.number_input('Pallets', min_value=1,max_value=200,value=5,key='cplt')
        _cpro=st.text_area(_T['products_label'].replace(':',''),key='cpro',placeholder='ej: 3 pallets Granadilla...',height=70)
        _cmsg=st.text_area(_T['quote_msg_lbl'],key='cmsg',placeholder=_T['quote_msg_ph'],height=70)
        if st.button(_T['send_quote'], key='bcot', type='primary', use_container_width=True):
            if not _ce or not _cpro: st.error(_T['err_email'] + ' / ' + _T['products_label'])
            else:
                _cy=datetime.now().strftime('%Y');_cpv=[p for p in load_pedidos() if p.get('id','').startswith(f'COT-{_cy}')]
                _cid=f'COT-{_cy}-{len(_cpv)+1:04d}'
                _cp={'id':_cid,'tipo':'cotizacion_especial','client_name':_cn,'client_email':_ce,'destino':_cd,'pallets_aprox':_cplt,'productos_interes':_cpro,'mensaje':_cmsg,'estado':'Pendiente revisión','fecha':datetime.now().isoformat(),'total_usd':0,'productos':[],'historial_estados':[{'estado':'Recibido','fecha':datetime.now().isoformat(),'usuario':'portal'}]}
                _ct=load_pedidos();_ct.append(_cp);save_pedidos(_ct);send_order_email(_cp)
                st.success(f'{_cid} - ' + _T['send_quote'])
    st.markdown(f'<div style="text-align:center;color:#888"><small>{LANG_TEXTS[st.session_state.get("portal_lang","es")]["footer_text"]}</small></div>', unsafe_allow_html=True)

# ─── MAIN ────────────────────────────────────────────────────────────────────
def main():
    init_session()
    auto_load_excel()
    hidratar_pedidos_gist()  # #3: recupera 'Mis Pedidos' tras reinicio (disco efímero)
    if theme:
        theme.aplicar()

    # Determine mode: 'portal' (public) or 'admin' (staff)
    # Support ?view=cliente URL param to always show portal
    if 'app_mode' not in st.session_state:
        _qp = st.query_params
        _view = _qp.get('view', '')
        if _view == 'admin':
            st.session_state.app_mode = 'admin'
        else:
            st.session_state.app_mode = 'portal'
    # Allow switching to admin via URL even if session already set
    elif st.query_params.get('view', '') == 'admin' and st.session_state.app_mode == 'portal':
        st.session_state.app_mode = 'admin'
        st.rerun()

    # ── MODO PORTAL (PÚBLICO) ─────────────────────────────────────────────────
    if st.session_state.app_mode == 'portal':
        # Small admin access link in sidebar
        # Sidebar branding: logo si existe, sino texto sin emoji
        try:
            import os as _osbrand
            if _osbrand.path.exists('logo.png'):
                st.sidebar.image('logo.png', use_container_width=True)
            else:
                st.sidebar.markdown('### Export Haret')
        except Exception:
            st.sidebar.markdown('### Export Haret')
        st.sidebar.caption(LANG_TEXTS[st.session_state.get('portal_lang','es')]['sidebar_subtitle'])
        st.sidebar.markdown('---')
        st.sidebar.markdown(f'<p style="text-align:center;margin:4px 0 8px"><a href="?view=admin" target="_self" style="color:#aaa;font-size:0.75em;text-decoration:none">{LANG_TEXTS[st.session_state.get("portal_lang","es")]["admin_access"]}</a></p>', unsafe_allow_html=True)
        # PEND1: Boton descargar catalogo oficial (PDF en servidor exportharet.com)
        try:
            st.sidebar.link_button(
                LANG_TEXTS[st.session_state.get('portal_lang','es')]['download_catalog'],
                'https://exportharet.com/wp-content/uploads/2026/04/Catalog-Export-Haret.pdf',
                use_container_width=True,
            )
        except Exception:
            pass
        st.sidebar.markdown('---')
        st.sidebar.caption(LANG_TEXTS[st.session_state.get('portal_lang','es')]['sidebar_footer'])
        render_portal_pedido()
        return

    # ── MODO ADMIN (STAFF LOGIN REQUERIDO) ────────────────────────────────────
    # Reanudar sesión tras refrescar: si hay token válido en la URL (?s=...),
    # restaurar el login sin pedir credenciales (hasta 5 h de inactividad).
    if not st.session_state.get('logged_in'):
        _tok_url = st.query_params.get('s', '')
        _rec_sess = admin_session_resume(_tok_url)
        if _rec_sess:
            st.session_state.logged_in = True
            st.session_state.user_email = _rec_sess.get('email', '')
            st.session_state.user_rol = _rec_sess.get('rol', '')
            st.session_state.user_nombre = _rec_sess.get('nombre', '')
            st.session_state['_admin_tok'] = _tok_url
    if not st.session_state.logged_in:
        login_page()
        return
    # Actividad: renueva la ventana de inactividad de 5 h (máx. 1 escritura/min)
    admin_session_touch(st.session_state.get('_admin_tok'))

    # Admin panel — cabecera limpia (el logo va en el sidebar; sin duplicados)
    _admin_css()
    _app_title = load_app_config().get("app_title", "Export Haret — Panel de Administración")
    _uname = st.session_state.get('user_nombre', '') or ''
    _urol = st.session_state.get('user_rol', '') or ''
    _ini_a = (_uname or 'A').strip()[:1].upper()
    st.markdown(
        '<div style="display:flex;align-items:center;justify-content:space-between;gap:12px;'
        'background:#fff;border:1px solid #e7eaef;border-radius:16px;padding:13px 20px;margin:2px 0 18px;'
        'box-shadow:0 1px 3px rgba(20,60,40,.05)">'
        '<div style="line-height:1.15;min-width:0">'
        '<div style="font-size:.66rem;letter-spacing:1.8px;text-transform:uppercase;color:#0c6e51;font-weight:700">Panel de administración</div>'
        f'<div style="font-weight:800;color:#14201a;font-size:1.2rem;letter-spacing:-.4px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{_esc(_app_title)}</div></div>'
        '<div style="display:flex;align-items:center;gap:10px;flex:0 0 auto">'
        f'<div style="text-align:right;line-height:1.2"><div style="font-weight:700;color:#16201b;font-size:.9rem">{_esc(_uname)}</div>'
        f'<div style="color:#65726b;font-size:.76rem;text-transform:capitalize">{_esc(_urol)}</div></div>'
        f'<div style="width:38px;height:38px;border-radius:50%;background:#0c6e51;color:#fff;font-weight:800;'
        f'display:flex;align-items:center;justify-content:center;font-size:1rem;flex:0 0 auto">{_esc(_ini_a)}</div>'
        '</div></div>', unsafe_allow_html=True)
    # Sidebar branding admin: logo si existe
    try:
        import os as _osSB
        if _osSB.path.exists('logo.png'):
            st.sidebar.image('logo.png', use_container_width=True)
        else:
            st.sidebar.markdown('### Export Haret')
        st.sidebar.caption('Panel de administración')
    except Exception:
        st.sidebar.markdown(f'# {_app_title}')
    st.sidebar.markdown('---')
    pedidos = load_pedidos()
    clients = load_clients()
    _sm1, _sm2 = st.sidebar.columns(2)
    _sm1.metric('📦 Pedidos', len(pedidos))
    _sm2.metric('👥 Clientes', len(clients))
    st.sidebar.metric('💵 Facturación', f"${sum(p.get('total_usd',0) for p in pedidos):,.0f}")
    pending = len([p for p in pedidos if p.get('estado') in ['Recibido','Confirmado','Preparando']])
    # KPI clickable - filtra Pedidos por estados pendientes
    if st.sidebar.button(f'⏳ En proceso: {pending}', use_container_width=True, key='kpi_en_proceso', help='Click para filtrar Pedidos por estados activos'):
        st.session_state['pedidos_filter_estado'] = ['Recibido','Confirmado','Preparando']
        st.session_state['admin_active_tab'] = 'pedidos'
        st.rerun()
    st.sidebar.markdown('---')
    with st.sidebar.expander(LANG_TEXTS[st.session_state.get('portal_lang','es')]['share_portal_title'], expanded=False):
        _portal_url = 'https://exportharet-pedidos.streamlit.app/'
        st.caption(LANG_TEXTS[st.session_state.get('portal_lang','es')]['share_portal_caption'])
        st.code(_portal_url, language=None)
        from urllib.parse import quote as _qsh
        _msg_sh = _qsh(LANG_TEXTS[st.session_state.get('portal_lang','es')]['share_msg'] + ' ' + _portal_url)
        _c1sh, _c2sh = st.columns(2)
        with _c1sh:
            st.link_button('💬 WhatsApp', f'https://wa.me/?text={_msg_sh}', use_container_width=True)
        with _c2sh:
            st.link_button('✉️ Email', f'mailto:?subject={_qsh("Portal de Pedidos Export Haret")}&body={_msg_sh}', use_container_width=True)
        st.caption('🔓 Acceso público, sin login.')
    st.sidebar.markdown('---')
    if st.sidebar.button('🌐 Ver Portal Clientes', use_container_width=True, key='admin_go_portal'):
        admin_session_end(st.session_state.get('_admin_tok'))
        st.session_state['_admin_tok'] = None
        st.session_state.app_mode = 'portal'
        st.session_state.logged_in = False
        st.query_params.clear()
        st.rerun()
    if st.sidebar.button('🚪 Cerrar Sesión', use_container_width=True):
        admin_session_end(st.session_state.get('_admin_tok'))
        st.session_state['_admin_tok'] = None
        st.session_state.logged_in = False
        try:
            del st.query_params['s']   # quita el token de la URL (sigue en ?view=admin)
        except Exception:
            pass
        st.rerun()
    st.sidebar.caption(LANG_TEXTS[st.session_state.get('portal_lang','es')]['sidebar_footer'])

    tab1,tab2,tab3,tab4,tab5,tab6=st.tabs([
        '📊 Dashboard',
        '📄 Catálogo & Precios',
        '🛒 Hacer Pedido',
        '⚙️ Configuración',
        '📦 Pedidos',
        '👥 Clientes',
    ])
    with tab1: render_dashboard()
    with tab2: render_catalogo()
    with tab3: render_hacer_pedido()
    with tab4: render_configuracion()
    with tab5: render_gestion_pedidos()
    with tab6: render_clientes()
    st.markdown('---')
    st.markdown('<div style="text-align:center;color:#888;"><small>🚀 Export Haret © 2026 | Sistema Profesional de Gestión de Pedidos</small></div>',unsafe_allow_html=True)

if __name__=='__main__':
    main()
