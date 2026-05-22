"""
vigilar_excel.py — Export Haret
================================
Monitorea Cotizaciones.xlsx y publica cambios automáticamente en la app.

Uso:
  .venv/bin/python3 vigilar_excel.py          # modo silencioso
  .venv/bin/python3 vigilar_excel.py --verbose # muestra detalle de cambios

Al detectar una modificación en el Excel:
  1. Lee los últimos precios del historial semanal
  2. Lee parámetros de CONFIGURACION (tarifas, DUE, merma, etc.)
  3. Actualiza precios_data.json
  4. Publica en GitHub → Streamlit Cloud se actualiza en ~1 min
  5. Muestra notificación en el Mac
"""

import os, sys, time, json, io, base64, requests, traceback
from datetime import datetime
from pathlib import Path
from watchdog.observers import Observer
from watchdog.events  import FileSystemEventHandler
from openpyxl         import load_workbook

# ── Rutas ─────────────────────────────────────────────────────────────────────
BASE        = Path(__file__).parent
EXCEL_PATH  = BASE / "Documentacion" / "Precios" / "Cotizaciones.xlsx"
DATA_PATH   = BASE / "precios_data.json"
SECRETS_FILE = BASE / ".streamlit" / "secrets.toml"

GITHUB_OWNER = "expharet"
GITHUB_REPO  = "app-de-pedidos"
VERBOSE = "--verbose" in sys.argv or "-v" in sys.argv

# ── Leer token de secrets ─────────────────────────────────────────────────────
def _read_secret(key):
    if SECRETS_FILE.exists():
        for line in SECRETS_FILE.read_text().splitlines():
            if key in line and "=" in line:
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    return os.environ.get(key, "")

GITHUB_TOKEN = _read_secret("GITHUB_TOKEN")

# ── Colores para terminal ─────────────────────────────────────────────────────
GREEN  = "\033[92m"; YELLOW = "\033[93m"; RED = "\033[91m"; RESET = "\033[0m"
BOLD   = "\033[1m"

def log(msg, color=""):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"{color}[{ts}] {msg}{RESET}", flush=True)


# ── Sincronización Excel → JSON ───────────────────────────────────────────────
COL_MAP = {
     4: "F-PSG10",    # D  Granadilla
     5: "F-PN016",    # E  Lulo
     6: "F-PPA01",    # F  Amarilla P
     7: "F-PSR02",    # G  Roja P
     8: "F-PSR05",    # H  Blanca P
     9: "F-PSM09",    # I  Maracuyá
    10: "F-TAS04",    # J  Tomate de árbol
    11: "F-GNB010",   # K  Guanabana
    12: "F-MPS03",    # L  Pepino dulce
    13: "F-CCN017",   # M  Cacao
    14: "F-BCC013",   # N  Babaco
    15: "F-AHSS012",  # O  Aguacate
    16: "F-BBB06",    # P  Baby banano
    17: "F-ZPT020",   # Q  Zapote Mamey
    18: "F-TX020",    # R  Taxo
    19: "F-UVP08",    # S  Physalis
    20: "F-UVP07",    # T  Physalis - husk
}

def sync_excel(excel_bytes: bytes, data: dict) -> tuple[dict, list[str]]:
    """Devuelve (data_actualizada, lista_cambios)."""
    wb     = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws_cfg = wb["CONFIGURACION"]
    ws_pr  = wb["TABLA PRECIOS"]
    changes = []
    new_data = json.loads(json.dumps(data))   # deep copy
    cfg      = new_data["config"]

    # ── Parámetros CONFIGURACION ────────────────────────────────────────────
    for row in ws_cfg.iter_rows():
        for cell in row:
            v  = str(cell.value or "")
            c3 = ws_cfg.cell(row=cell.row, column=3).value
            if not isinstance(c3, (int, float)):
                continue
            val = float(c3)
            # Parámetros básicos
            if "Costo de la caja" in v and abs(cfg.get("costo_caja",0)-val)>1e-4:
                changes.append(f"costo_caja: {cfg.get('costo_caja')} → {val}")
                cfg["costo_caja"] = val
            elif "Merma" in v and "%" in v and abs(cfg.get("merma_pct",0)-val)>1e-6:
                changes.append(f"merma_pct: {cfg.get('merma_pct')} → {val}")
                cfg["merma_pct"] = val
            elif "DUE" in v and "fijo" in v and abs(cfg.get("due",0)-val)>0.01:
                changes.append(f"DUE: {cfg.get('due')} → {val}")
                cfg["due"] = val
            elif "Peso pallet" in v and abs(cfg.get("peso_pallet",0)-val)>0.01:
                changes.append(f"peso_pallet: {cfg.get('peso_pallet')} → {val}")
                cfg["peso_pallet"] = val
            elif "Tara de la caja" in v and abs(cfg.get("tara_caja",0)-val)>0.001:
                changes.append(f"tara_caja: {cfg.get('tara_caja')} → {val}")
                cfg["tara_caja"] = val
            elif "transporte interno" in v.lower() and "costo" in v.lower():
                if abs(cfg.get("transporte_interno",0)-val)>0.01:
                    changes.append(f"transporte_interno: {cfg.get('transporte_interno')} → {val}")
                    cfg["transporte_interno"] = val
            # Tarifas destino: columna B = nombre destino
            dest = str(ws_cfg.cell(row=cell.row, column=2).value or "")
            if cell.column == 2 and dest in cfg.get("destinos", {}):
                if abs(cfg["destinos"][dest] - val) > 0.001:
                    changes.append(f"tarifa {dest}: {cfg['destinos'][dest]} → {val}")
                    cfg["destinos"][dest] = val

    # ── Precios desde historial semanal (TABLA PRECIOS filas 32-83) ─────────
    latest = {}
    for col, codigo in COL_MAP.items():
        last = None
        for r in range(32, 84):
            v = ws_pr.cell(row=r, column=col).value
            if isinstance(v, (int, float)) and v > 0:
                last = float(v)
        if last:
            latest[codigo] = last

    for i, p in enumerate(new_data["products"]):
        if p["codigo"] in latest:
            new_price = latest[p["codigo"]]
            if abs(p["precio_compra"] - new_price) > 0.001:
                changes.append(f"{p['producto']}: ${p['precio_compra']:.2f} → ${new_price:.2f}")
                new_data["products"][i]["precio_compra"] = new_price

    return new_data, changes


# ── Publicar en GitHub ────────────────────────────────────────────────────────
def publish_to_github(data: dict) -> bool:
    if not GITHUB_TOKEN:
        log("Token GitHub no configurado — no se puede publicar", RED)
        return False

    headers = {
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept":        "application/vnd.github.v3+json",
    }
    content = json.dumps(data, indent=2, ensure_ascii=False).encode()
    b64     = base64.b64encode(content).decode()
    api_url = f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/precios_data.json"

    r_get = requests.get(api_url, headers=headers, timeout=15)
    if r_get.status_code != 200:
        log(f"Error obteniendo SHA del archivo: {r_get.status_code}", RED)
        return False
    sha = r_get.json()["sha"]

    payload = {
        "message": f"Sync automático desde Excel — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "content": b64,
        "sha":     sha,
    }
    r_put = requests.put(api_url, json=payload, headers=headers, timeout=20)
    return r_put.status_code in (200, 201)


# ── Notificación macOS ────────────────────────────────────────────────────────
def notify(title: str, msg: str):
    os.system(
        f"osascript -e 'display notification \"{msg}\" "
        f"with title \"{title}\" sound name \"Glass\"'"
    )


# ── Handler del vigilante ─────────────────────────────────────────────────────
class ExcelHandler(FileSystemEventHandler):
    def __init__(self):
        self._last_sync = 0

    def on_modified(self, event):
        if not str(event.src_path).endswith("Cotizaciones.xlsx"):
            return
        now = time.time()
        if now - self._last_sync < 5:   # debounce 5 s
            return
        self._last_sync = now
        time.sleep(2)                    # esperar a que Excel termine de escribir
        self._run_sync()

    def _run_sync(self):
        log(f"Cambio detectado en Cotizaciones.xlsx — sincronizando…", YELLOW)
        try:
            if not DATA_PATH.exists():
                log("precios_data.json no existe aún, se creará.", YELLOW)
                return

            with open(DATA_PATH) as f:
                data = json.load(f)

            excel_bytes = EXCEL_PATH.read_bytes()
            new_data, changes = sync_excel(excel_bytes, data)

            if not changes:
                log("Sin cambios detectados.", "")
                return

            log(f"{len(changes)} cambios:", GREEN)
            if VERBOSE:
                for c in changes:
                    print(f"   • {c}")

            # Guardar JSON
            DATA_PATH.write_text(
                json.dumps(new_data, indent=2, ensure_ascii=False), encoding="utf-8"
            )
            log("precios_data.json actualizado ✓", GREEN)

            # Publicar en GitHub
            log("Publicando en Streamlit Cloud…", YELLOW)
            ok = publish_to_github(new_data)
            if ok:
                log("¡Publicado! La app se actualizará en ~1 minuto. ✅", GREEN + BOLD)
                notify("Export Haret",
                       f"{len(changes)} precio(s) actualizados desde Excel")
            else:
                log("Error publicando en GitHub.", RED)
                notify("Export Haret ⚠️", "No se pudo publicar en la nube")

        except Exception:
            log(f"Error durante la sincronización:\n{traceback.format_exc()}", RED)


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"\n{BOLD}{'='*55}")
    print("   Export Haret — Vigilante automático de Excel")
    print(f"{'='*55}{RESET}")
    print(f"  Archivo:  {EXCEL_PATH}")
    print(f"  JSON:     {DATA_PATH}")
    print(f"  GitHub:   {GITHUB_OWNER}/{GITHUB_REPO}")
    print(f"  Token:    {'✓ configurado' if GITHUB_TOKEN else '✗ NO encontrado'}")
    print(f"  Verbose:  {'sí' if VERBOSE else 'no  (usa --verbose para ver cambios)'}")
    print(f"\n  Guardando el Excel → sincronización automática")
    print(f"  Ctrl+C para detener\n")

    if not EXCEL_PATH.exists():
        print(f"{RED}ERROR: No se encuentra el Excel en:\n  {EXCEL_PATH}{RESET}")
        sys.exit(1)

    if not GITHUB_TOKEN:
        print(f"{YELLOW}AVISO: Sin token GitHub — los cambios se guardarán localmente")
        print(f"       pero NO se publicarán en la nube.{RESET}\n")

    observer = Observer()
    handler  = ExcelHandler()
    observer.schedule(handler, path=str(EXCEL_PATH.parent), recursive=False)
    observer.start()

    log("Vigilante activo — esperando cambios…", GREEN)

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        log("Vigilante detenido.", "")
    observer.join()
